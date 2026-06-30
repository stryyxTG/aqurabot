from __future__ import annotations
import asyncio
import hashlib
import html
import inspect
import io
import json
import logging
import math
import re
import shutil
import aiohttp
import tempfile
import uuid
import zipfile
from contextlib import closing
from datetime import datetime, timedelta, timezone
from pathlib import Path

from aiogram import Bot, Dispatcher, F, BaseMiddleware
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.exceptions import TelegramAPIError, TelegramBadRequest, TelegramForbiddenError, TelegramRetryAfter
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import CallbackQuery, ChatJoinRequest, ChatMemberUpdated, FSInputFile, InlineKeyboardButton, InlineKeyboardMarkup, Message, TelegramObject

from .db import (
    accept_user_agreement,
    add_balance,
    add_catalog_country,
    add_product_group_to_cart,
    add_product_to_cart,
    add_topup_reviewer,
    approve_service_order,
    approve_topup_request,
    claim_product_for_admin,
    count_product_departments,
    count_available_products_in_department,
    count_products_in_department,
    count_products,
    count_user_purchases,
    create_service_order_with_charge,
    create_topup_request,
    create_catalog_department,
    create_product,
    clear_cart,
    delete_sold_product_with_history,
    force_remove_product,
    get_all_users_with_purchases,
    get_available_countries,
    get_catalog_country,
    get_balance,
    get_crypto_invoice,
    get_db_conn,
    find_existing_product_identity,
    get_product,
    get_product_department,
    get_product_group,
    get_channel_join_request,
    get_app_meta,
    get_service_order,
    get_stats,
    get_stuck_products,
    get_topup_request,
    get_user,
    has_user_accepted_agreement,
    init_db,
    is_topup_reviewer,
    list_catalog_countries,
    list_cart_items,
    list_country_counts,
    list_user_ids,
    list_available_products_in_department,
    list_products,
    list_products_in_department,
    list_product_departments,
    list_product_session_references,
    list_product_session_paths,
    list_sold_products_for_manual_cleanup,
    list_user_batch_purchases,
    list_user_purchase_groups,
    list_topup_reviewers,
    list_user_purchases,
    mark_product_session_cleanup_disabled,
    mark_crypto_invoice_status,
    process_crypto_topup,
    purchase_cart,
    reject_topup_request,
    reject_service_order,
    reset_revenue_stats,
    record_channel_member,
    record_channel_join_request,
    record_crypto_invoice,
    rename_catalog_country,
    remove_catalog_country,
    remove_product_department,
    remove_product_from_cart,
    remove_topup_reviewer,
    update_product_info,
    update_product_group_info,
    purchase_product,
    remove_product,
    reset_stats,
    return_product_to_catalog,
    update_product_session_path,
    update_product_status,
    update_channel_join_status,
    upsert_user,
    set_app_meta,
    search_products,
)
from .config import Settings, load_settings
from .keyboards import (
    agreement_kb,
    admin_catalog_kb,
    admin_clean_confirm_kb,
    admin_clean_kb,
    admin_countries_available_kb,
    admin_country_kb,
    admin_country_remove_confirm_kb,
    admin_product_group_kb,
    admin_product_group_remove_confirm_kb,
    admin_home_kb,
    admin_product_detail_kb,
    admin_product_kb,
    admin_claim_confirm_kb,
    admin_product_search_results_kb,
    admin_products_by_country_kb,
    admin_terminate_sessions_step1_kb,
    admin_terminate_sessions_step2_kb,
    admin_scan_confirm_kb,
    admin_scan_settings_kb,
    admin_stats_kb,
    admin_sold_history_kb,
    back_to_main_kb,
    BTN_ICON_ADMIN,
    BTN_ICON_BACK,
    BTN_ICON_CANCEL,
    BTN_ICON_CHECK,
    BTN_ICON_PAY,
    batch_download_confirm_kb,
    cart_kb,
    catalog_sections_kb,
    help_menu_kb,
    cancel_flow_kb,
    catalog_home_kb,
    code_keypad_kb,
    code_received_kb,
    country_select_kb,
    main_menu_kb,
    menu_only_kb,
    open_cart_kb,
    product_detail_kb,
    product_group_cart_kb,
    product_group_detail_kb,
    premium_periods_kb,
    purchase_history_detail_kb,
    purchase_batch_kb,
    purchases_nav_kb,
    product_list_kb,
    proxy_menu_kb,
    purchase_success_kb,
    purchase_waiting_kb,
    stuck_products_kb,
    topup_confirm_kb,
    drop_manage_kb,
    drops_menu_kb,
    topup_methods_kb,
    topup_other_kb,
    topup_receipt_kb,
    topup_review_kb,
    admin_user_manage_kb,
    support_kb,
    service_detail_kb,
    service_order_confirm_kb,
    service_order_review_kb,
    service_recipient_cancel_kb,
    stars_packages_kb,
    subscription_kb,
)
from .proxy_store import load_global_proxy, save_global_proxy
from .proxy_utils import check_proxy_connectivity, format_proxy_summary, parse_proxy_input
from .paths import product_session_base_path, SESSIONS_DIR, ROOT_DIR
from .session_flow import LoginExpiredError, ShopSessionManager
from .session_metadata import (
    load_metadata_file,
    metadata_match_keys,
    normalize_session_metadata,
    session_json_path,
    write_session_metadata,
)
from .states import AdminAddProductStates, AdminBroadcastStates, AdminCardsStates, AdminCatalogStates, AdminCleanStates, AdminProxyStates, AdminTopUpStates, AdminEditProductStates, AdminEditProductGroupStates, AdminSearchUserStates, AdminSearchProductStates, AdminDropsStates, UserTopUpStates, UserCartStates, UserCatalogStates, ServiceOrderStates, AdminScanStates


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    handlers=[logging.FileHandler("bot.log", encoding="utf-8"), logging.StreamHandler()],
)
logger = logging.getLogger("shopbot")

settings: Settings = load_settings()
PAGE_SIZE = 7
PURCHASE_PAGE_SIZE = 7
ADMIN_SOLD_PAGE_SIZE = 21
JOIN_REQUEST_PENDING_TTL = timedelta(hours=24)
LOG_CHANNEL_ID = settings.log_channel_id
RU_TOPUP_CHAT_ID = settings.ru_topup_chat_id
UA_TOPUP_CHAT_ID = settings.ua_topup_chat_id
STARS_RATE_RUB = 1.4
STARS_PACKAGES = [50, 75, 100, 150, 200, 250, 300, 500, 1000]
PREMIUM_PRICES_RUB = {
    3: 1175.0,
    6: 1425.0,
    12: 2350.0,
}
FEATURE_TOPUP_RU = False
FEATURE_TOPUP_UA = True
FEATURE_TOPUP_CRYPTO = True
FEATURE_TOPUP_OTHER = False
FEATURE_CATALOG_PREMIUM = False
FEATURE_CATALOG_STARS = False
MIN_CRYPTO_TOPUP_AMOUNT = 0.1
FALLBACK_USD_TO_UAH_RATE = 41.5
UA_CARDS_META_KEY = "topup_cards_ua"
MONEY_EPSILON = 0.000001
TG_EMOJI_RE = re.compile(r'<tg-emoji\s+emoji-id=(["\'])(\d{5,32})\1>([^<>]{0,32})</tg-emoji>')
TG_EMOJI_START_RE = re.compile(r'^\s*<tg-emoji\s+emoji-id=(["\'])(\d{5,32})\1>([^<>]{0,32})</tg-emoji>\s*')
ICON_SHOP = '<tg-emoji emoji-id="5920332557466997677">🏪</tg-emoji>'
ICON_COIN = '<tg-emoji emoji-id="5904462880941545555">🪙</tg-emoji>'
ICON_TAG = '<tg-emoji emoji-id="5890883384057533697">🏷</tg-emoji>'
ICON_BRIEFCASE = '<tg-emoji emoji-id="5938492039971737551">💼</tg-emoji>'
ICON_FOLDER = '<tg-emoji emoji-id="6037475557082403885">📁</tg-emoji>'
ICON_CRYPTO = '<tg-emoji emoji-id="6037083366438737901">💎</tg-emoji>'
ICON_UA = '<tg-emoji emoji-id="5264782095531661663">🇺🇦</tg-emoji>'
ICON_RU = '<tg-emoji emoji-id="5424670808700114602">🇷🇺</tg-emoji>'
ICON_HELP = '<tg-emoji emoji-id="5774077015388852135">❌</tg-emoji>'
ICON_COUNTRY = '<tg-emoji emoji-id="5870718740236079262">🌐</tg-emoji>'
ICON_SPARKLE = '<tg-emoji emoji-id="5453980026305274747">✨</tg-emoji>'
ICON_WALLET = '<tg-emoji emoji-id="5769126056262898415">👛</tg-emoji>'
ICON_CART = '<tg-emoji emoji-id="5935991549528570744">🛒</tg-emoji>'
ICON_PARTY = '<tg-emoji emoji-id="6041731551845159060">🎉</tg-emoji>'
ICON_PURCHASE_TAG = '<tg-emoji emoji-id="5886285355279193209">🏷</tg-emoji>'
ICON_SUCCESS = '<tg-emoji emoji-id="5938252440926163756">✅</tg-emoji>'
ICON_KEYBOARD = '<tg-emoji emoji-id="5873204392429096339">⌨</tg-emoji>'
ICON_CHECK = '<tg-emoji emoji-id="5870633910337015697">✅</tg-emoji>'
ICON_BLOCK = '<tg-emoji emoji-id="5872988737826197458">⛔️</tg-emoji>'
ICON_HEART = '<tg-emoji emoji-id="5872863028428410654">❤</tg-emoji>'
ICON_AUTO_SPARKLE = '<tg-emoji emoji-id="5890925363067886150">✨</tg-emoji>'
ICON_BANK = '<tg-emoji emoji-id="5312110302966343332">🏦</tg-emoji>'
ICON_CARD = '<tg-emoji emoji-id="5240450298345956614">💳</tg-emoji>'
ICON_TG_ACCOUNTS = '<tg-emoji emoji-id="6028346797368283073">✈️</tg-emoji>'
ICON_CATALOG_SECTIONS = '<tg-emoji emoji-id="5766994197705921104">🗂</tg-emoji>'
ICON_TG_PREMIUM = '<tg-emoji emoji-id="6028338546736107668">⭐️</tg-emoji>'
ICON_TG_STARS = '<tg-emoji emoji-id="5463289097336405244">⭐️</tg-emoji>'
ICON_STAR_RATE = '<tg-emoji emoji-id="5435957248314579621">⭐️</tg-emoji>'
ICON_TIME = '<tg-emoji emoji-id="5983150113483134607">⏰️</tg-emoji>'
ICON_NOTICE = '<tg-emoji emoji-id="6030563507299160824">❗️</tg-emoji>'
bot = Bot(settings.bot_token, default=DefaultBotProperties(parse_mode=ParseMode.HTML, link_preview_is_disabled=True))
dp = Dispatcher(storage=MemoryStorage())
session_manager = ShopSessionManager(settings)


def topup_methods_keyboard() -> InlineKeyboardMarkup:
    return topup_methods_kb(
        bool(settings.cryptopay_token) and FEATURE_TOPUP_CRYPTO,
        ru_enabled=FEATURE_TOPUP_RU,
        ua_enabled=FEATURE_TOPUP_UA,
        other_enabled=FEATURE_TOPUP_OTHER,
    )


def catalog_sections_keyboard() -> InlineKeyboardMarkup:
    return catalog_sections_kb(
        premium_enabled=FEATURE_CATALOG_PREMIUM,
        stars_enabled=FEATURE_CATALOG_STARS,
    )


def is_topup_method_enabled(method: str) -> bool:
    return {
        "ru": FEATURE_TOPUP_RU,
        "ua": FEATURE_TOPUP_UA,
        "crypto": FEATURE_TOPUP_CRYPTO and bool(settings.cryptopay_token),
    }.get(method, False)


def subscription_keyboard() -> InlineKeyboardMarkup:
    return subscription_kb(settings.required_channel_url or settings.required_channel)


async def remember_required_channel(chat_id: int) -> None:
    await set_app_meta("required_channel_chat_id", str(chat_id))


async def get_required_channel_chat_id() -> str | None:
    if settings.required_channel:
        return settings.required_channel
    return await get_app_meta("required_channel_chat_id")


async def has_subscription_access(user_id: int) -> tuple[bool, bool]:
    """Returns (allowed, allowed_by_join_request)."""
    required_channel = await get_required_channel_chat_id()
    if required_channel:
        try:
            member = await bot.get_chat_member(chat_id=required_channel, user_id=user_id)
            if member.status not in ("left", "kicked"):
                await update_channel_join_status(user_id, "joined")
                return True, False
            if member.status == "kicked":
                await update_channel_join_status(user_id, "kicked")
                return False, False
        except Exception as exc:
            logger.debug("Could not check channel member %s in %s: %s", user_id, required_channel, exc)

    join_request = await get_channel_join_request(user_id)
    if join_request:
        chat_id = int(join_request["chat_id"])
        stored_status = join_request["status"] or "pending"
        requested_at = parse_iso_datetime(join_request["requested_at"])
        if stored_status == "pending" and requested_at and datetime.now(timezone.utc) - requested_at > JOIN_REQUEST_PENDING_TTL:
            await update_channel_join_status(user_id, "expired")
            return False, False

        try:
            member = await bot.get_chat_member(chat_id=chat_id, user_id=user_id)
            if member.status not in ("left", "kicked"):
                await update_channel_join_status(user_id, "joined")
                return True, False
            if stored_status == "joined":
                await update_channel_join_status(user_id, member.status)
                return False, False
        except Exception as exc:
            logger.debug("Could not check stored channel member %s in %s: %s", user_id, chat_id, exc)
            if stored_status == "joined":
                return False, False

        if stored_status == "pending":
            return True, True

    return False, False


class SubscriptionMiddleware(BaseMiddleware):
    async def __call__(self, handler, event: TelegramObject, data: dict):
        user = data.get("event_from_user")
        if not user or user.is_bot:
            return await handler(event, data)

        # Админов не мучаем проверкой
        if is_admin(user.id):
            return await handler(event, data)

        # Разрешаем кнопку проверки подписки
        if isinstance(event, CallbackQuery) and event.data == "check_sub":
            return await handler(event, data)

        if isinstance(event, CallbackQuery) and (event.data or "").startswith("topup_"):
            chat_id = event.message.chat.id if event.message and event.message.chat else None
            if chat_id in {RU_TOPUP_CHAT_ID, UA_TOPUP_CHAT_ID}:
                return await handler(event, data)

        allowed, _ = await has_subscription_access(user.id)
        if not allowed:
            text = (
                f"{ICON_NOTICE} <b>Доступ ограничен</b>\n\n"
                f"Вступите в канал, затем нажмите «Проверить»."
            )
            if isinstance(event, Message):
                await event.answer(text, reply_markup=subscription_keyboard())
            elif isinstance(event, CallbackQuery):
                await safe_edit(event.message, text, subscription_keyboard())
                await event.answer()
            return

        return await handler(event, data)


class AgreementMiddleware(BaseMiddleware):
    async def __call__(self, handler, event: TelegramObject, data: dict):
        user = data.get("event_from_user")
        if not user or user.is_bot or is_admin(user.id):
            return await handler(event, data)

        if isinstance(event, CallbackQuery) and event.data in {"agreement_accept", "check_sub"}:
            return await handler(event, data)

        await ensure_known_user(event)
        if await has_user_accepted_agreement(user.id):
            return await handler(event, data)

        await show_agreement(event)
        if isinstance(event, CallbackQuery):
            await event.answer()
        return


FSM_CALLBACK_WHITELIST = {
    AdminAddProductStates.waiting_add_method.state: ("admin_add_by_phone", "admin_add_by_session"),
    AdminAddProductStates.waiting_session_count.state: ("session_count_single", "session_count_bulk"),
    AdminAddProductStates.waiting_bulk_sessions.state: ("bulk_sessions_done",),
    AdminAddProductStates.waiting_code.state: ("code_digit:", "code_backspace", "code_clear", "code_submit"),
    AdminAddProductStates.waiting_country.state: ("add_country:",),
    AdminAddProductStates.waiting_department.state: ("add_department:", "add_department_new", "add_department_back"),
    ServiceOrderStates.waiting_recipient.state: ("service_order_cancel",),
    UserTopUpStates.waiting_receipt.state: ("cancel_topup_receipt",),
    AdminBroadcastStates.waiting_text.state: ("admin_broadcast_send", "broadcast_confirm"),
    AdminCleanStates.waiting_action.state: ("clean_refresh:", "clean_confirm:"),
    AdminCleanStates.waiting_confirm.state: ("clean_execute:", "clean_back:"),
}


def is_fsm_callback_allowed(current_state: str, callback_data: str) -> bool:
    if callback_data == "agreement_accept":
        return True
    if callback_data.startswith("cancel_flow:"):
        return True
    for allowed in FSM_CALLBACK_WHITELIST.get(current_state, ()):
        if allowed.endswith(":"):
            if callback_data.startswith(allowed):
                return True
        elif callback_data == allowed or callback_data.startswith(f"{allowed}:"):
            return True
    return False


class CallbackFSMGuardMiddleware(BaseMiddleware):
    async def __call__(self, handler, event: TelegramObject, data: dict):
        if not isinstance(event, CallbackQuery):
            return await handler(event, data)

        state: FSMContext | None = data.get("state")
        if state is None:
            return await handler(event, data)

        current_state = await state.get_state()
        if not current_state:
            return await handler(event, data)

        callback_data = event.data or ""
        if is_fsm_callback_allowed(current_state, callback_data):
            return await handler(event, data)

        await event.answer("Сначала завершите текущий шаг или нажмите «Отменить».", show_alert=True)


@dp.callback_query(F.data == "check_sub")
async def check_sub_callback(query: CallbackQuery):
    allowed, by_request = await has_subscription_access(query.from_user.id)
    if allowed:
        text = "✅ Доступ открыт." if by_request else "✅ Подписка подтверждена."
        await query.answer(text, show_alert=True)
        await start_logic(query)
    else:
        await query.answer("❌ Подписка не найдена.", show_alert=True)


@dp.callback_query(F.data == "agreement_accept")
async def agreement_accept(query: CallbackQuery):
    await ensure_known_user(query)
    await accept_user_agreement(query.from_user.id)
    await query.answer("✅ Условия приняты.", show_alert=True)
    await show_home(query)


@dp.chat_join_request()
async def channel_join_request_handler(request: ChatJoinRequest):
    required_username = settings.required_channel.lstrip("@").lower()
    request_username = (request.chat.username or "").lower()
    if required_username and request_username and request_username != required_username:
        return

    invite_link = request.invite_link.invite_link if request.invite_link else None
    expected_invite = (settings.required_channel_url or "").rstrip("/")
    actual_invite = (invite_link or "").rstrip("/")
    if settings.required_channel and expected_invite and actual_invite and actual_invite != expected_invite:
        return
    if expected_invite and not actual_invite and not settings.required_channel:
        logger.info("Join request without invite link recorded: user=%s chat=%s", request.from_user.id, request.chat.id)

    await record_channel_join_request(
        user_id=request.from_user.id,
        chat_id=request.chat.id,
        chat_username=request.chat.username,
        invite_link=invite_link,
    )
    await remember_required_channel(request.chat.id)
    await upsert_user(request.from_user.id, request.from_user.username, request.from_user.first_name)
    logger.info("Channel join request recorded: user=%s chat=%s", request.from_user.id, request.chat.id)


@dp.chat_member()
async def channel_member_update_handler(update: ChatMemberUpdated):
    user = update.new_chat_member.user
    if not user or user.is_bot:
        return

    join_request = await get_channel_join_request(user.id)
    if join_request and int(join_request["chat_id"]) != update.chat.id:
        return

    status = update.new_chat_member.status
    if status in ("left", "kicked"):
        if join_request:
            await update_channel_join_status(user.id, status)
        logger.info("Channel access revoked: user=%s chat=%s status=%s", user.id, update.chat.id, status)
    else:
        if join_request:
            await update_channel_join_status(user.id, "joined")
        else:
            await record_channel_member(user.id, update.chat.id, update.chat.username, "joined")
        await remember_required_channel(update.chat.id)
        logger.info("Channel access confirmed: user=%s chat=%s status=%s", user.id, update.chat.id, status)


def is_admin(user_id: int) -> bool:
    return user_id in settings.admin_ids


async def require_admin(target: Message | CallbackQuery) -> bool:
    if is_admin(target.from_user.id):
        return True
    text = "Доступ запрещен."
    if isinstance(target, CallbackQuery):
        await target.answer("Нет доступа.", show_alert=True)
        if target.message:
            await safe_edit(target.message, text)
    else:
        await target.answer(text)
    return False


def fmt_money(value: float) -> str:
    return f"{value:.2f} {settings.currency}"


def parse_iso_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def short_date(value: object) -> str:
    if not value:
        return "—"
    return str(value)[:10]


def extract_custom_emoji_id(message: Message) -> str | None:
    for entity in message.entities or []:
        if entity.type == "custom_emoji" and entity.custom_emoji_id:
            return entity.custom_emoji_id
    text = message.text or ""
    match = re.search(r'emoji-id\\?=["\']?(\d+)', text)
    if match:
        return match.group(1)
    digits = re.sub(r"\D", "", text)
    return digits if len(digits) >= 10 else None


def topup_method_title(method: str) -> str:
    return {
        "ru": "Российская карта",
        "ua": "Украинская карта",
        "crypto": "Crypto Bot",
    }.get(method, method)


def topup_method_icon(method: str) -> str:
    return {
        "ru": ICON_RU,
        "ua": ICON_UA,
        "crypto": ICON_CRYPTO,
    }.get(method, "")


def topup_currency(method: str) -> str:
    return "UAH" if method == "ua" else settings.currency


def topup_chat_id(method: str) -> int:
    return UA_TOPUP_CHAT_ID if method == "ua" else RU_TOPUP_CHAT_ID


async def get_usd_to_uah_rate() -> tuple[float, str]:
    url = "https://open.er-api.com/v6/latest/USD"
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url, timeout=8) as resp:
                data = await resp.json()
        rate = float((data.get("rates") or {}).get("UAH") or 0)
        if rate > 0:
            return rate, "курс API"
    except Exception as exc:
        logger.warning("Could not fetch USD->UAH rate, using fallback: %s", exc)
    return FALLBACK_USD_TO_UAH_RATE, "резервный курс"


async def get_ua_cards_text() -> str:
    return (await get_app_meta(UA_CARDS_META_KEY) or "").strip()


async def set_ua_cards_text(value: str) -> None:
    await set_app_meta(UA_CARDS_META_KEY, (value or "").strip())


def message_html_text(message: Message) -> str:
    value = getattr(message, "html_text", None) or message.text or message.caption or ""
    return str(value).strip()


def cards_command_payload(message: Message) -> str:
    html_text = message_html_text(message)
    return re.sub(r"^/cards(?:@\w+)?(?:\s+)?", "", html_text, count=1, flags=re.IGNORECASE).strip()


async def build_topup_quote(method: str, credit_amount: float) -> dict:
    if method == "ua":
        rate, source = await get_usd_to_uah_rate()
        payment_amount = math.ceil(credit_amount * rate * 100) / 100
        return {
            "credit_amount": credit_amount,
            "payment_amount": payment_amount,
            "payment_currency": "UAH",
            "rate": rate,
            "rate_source": source,
        }
    return {
        "credit_amount": credit_amount,
        "payment_amount": credit_amount,
        "payment_currency": settings.currency,
        "rate": 1.0,
        "rate_source": "без конвертации",
    }


async def manual_topup_requisites(method: str, quote: dict) -> str:
    credit_amount = float(quote["credit_amount"])
    payment_amount = float(quote["payment_amount"])
    if method == "ua":
        cards_text = await get_ua_cards_text()
        cards_block = f"\n\n{cards_text}\n\n" if cards_text else "\n\n"
        return (
            f"{ICON_UA} <b>Пополнение украинской картой</b>\n\n"
            f"<b>К зачислению:</b> {fmt_money(credit_amount)}\n"
            f"<b>К оплате:</b> {payment_amount:.2f} UAH\n"
            f"<b>Курс:</b> 1 USD = {float(quote['rate']):.4f} UAH\n"
            f"{cards_block}"
            f"{ICON_NOTICE} <b>После оплаты отправьте чек в формате PDF или скриншот PDF-файла</b>"
        )
    return (
        f"{ICON_RU} <b>Пополнение российской картой</b>\n\n"
        f"<b>К зачислению:</b> {fmt_money(credit_amount)}\n"
        f"<b>К оплате:</b> {fmt_money(payment_amount)}\n\n"
        f"{ICON_BANK} <b>Ozon Bank</b>\n"
        "По номеру телефона:\n"
        "<code>+79216521675</code>\n"
        "<b>Alexander Golovchanskiy</b>\n\n"
        f"{ICON_NOTICE} <b>После оплаты отправьте чек в формате PDF или скриншот PDF-файла</b>"
    )


def user_profile_label(user) -> str:
    if user.username:
        return f"@{user.username}"
    name = " ".join(part for part in [user.first_name, user.last_name] if part)
    return name or str(user.id)


def user_profile_html(*, user_id: int, username: str | None, label: str) -> str:
    if username:
        return f"<a href=\"https://t.me/{html.escape(username)}\">{html.escape(label)}</a>"
    return html.escape(label)


def normalize_recipient_username(value: str) -> str | None:
    text = (value or "").strip()
    text = re.sub(r"^https?://t\.me/", "", text, flags=re.IGNORECASE).strip()
    text = text.lstrip("@").strip()
    if not re.fullmatch(r"[A-Za-z0-9_]{5,32}", text):
        return None
    return f"@{text}"


def premium_service_label(months: int) -> str:
    return f"Premium на {premium_period_label(months)}"


def premium_period_label(months: int) -> str:
    if months == 12:
        return "1 год"
    if months in {3, 4}:
        return f"{months} месяца"
    return f"{months} месяцев"


def service_order_caption_from_row(order) -> str:
    username = order["username"] or ""
    first_name = order["first_name"] or ""
    profile_name = f"@{username}" if username else (first_name or str(order["user_id"]))
    status_map = {
        "pending": "ожидает выдачи",
        "delivered": "выдано",
        "rejected": "отказано",
    }
    return (
        f"<b>Заявка на выдачу #{order['order_id']}</b>\n\n"
        f"<b>Услуга:</b> {html.escape(order['service_label'])}\n"
        f"<b>Получатель:</b> <code>{html.escape(order['recipient'])}</code>\n"
        f"<b>Сумма:</b> {fmt_money(float(order['amount']))}\n\n"
        f"<b>Клиент:</b> {user_profile_html(user_id=int(order['user_id']), username=username or None, label=profile_name)}\n"
        f"<b>User ID:</b> <code>{order['user_id']}</code>\n"
        f"<b>Статус:</b> {status_map.get(order['status'], html.escape(order['status']))}"
    )


async def send_service_order_log(
    *,
    order_id: int,
    user,
    service_label: str,
    recipient: str,
    amount: float,
    new_balance: float,
) -> None:
    username = user.username or None
    profile_label = user_profile_label(user)
    text = (
        f"<b>Заявка на выдачу #{order_id}</b>\n\n"
        f"<b>Услуга:</b> {html.escape(service_label)}\n"
        f"<b>Получатель:</b> <code>{html.escape(recipient)}</code>\n"
        f"<b>Сумма:</b> {fmt_money(amount)}\n\n"
        f"<b>Клиент:</b> {user_profile_html(user_id=user.id, username=username, label=profile_label)}\n"
        f"<b>User ID:</b> <code>{user.id}</code>\n"
        f"<b>Баланс после списания:</b> {fmt_money(new_balance)}\n\n"
        "Статус: ожидает выдачи."
    )
    await bot.send_message(LOG_CHANNEL_ID, text, reply_markup=service_order_review_kb(order_id, username))


def topup_request_caption(request_id: int, *, user, method: str, payment_amount: float, currency: str, credit_amount: float) -> str:
    profile_label = user_profile_label(user)
    return (
        f"<b>Заявка на пополнение #{request_id}</b>\n\n"
        f"Метод: {topup_method_icon(method)} <b>{html.escape(topup_method_title(method))}</b>\n"
        f"<b>К оплате:</b> {payment_amount:.2f} {html.escape(currency)}\n"
        f"<b>К зачислению:</b> {fmt_money(credit_amount)}\n"
        f"Пользователь: {user_profile_html(user_id=user.id, username=user.username, label=profile_label)}\n"
        f"User ID: <code>{user.id}</code>\n\n"
        "Проверьте чек и выберите действие."
    )


def topup_request_caption_from_row(request) -> str:
    username = request["username"] or ""
    first_name = request["first_name"] or ""
    profile_name = f"@{username}" if username else (first_name or str(request["user_id"]))
    return (
        f"<b>Заявка на пополнение #{request['request_id']}</b>\n\n"
        f"Метод: {topup_method_icon(request['method'])} <b>{html.escape(topup_method_title(request['method']))}</b>\n"
        f"<b>К оплате:</b> {float(request['amount']):.2f} {html.escape(request['currency'])}\n"
        f"<b>К зачислению:</b> {fmt_money(float(request['credit_amount'] or request['amount']))}\n"
        f"Пользователь: {user_profile_html(user_id=int(request['user_id']), username=username or None, label=profile_name)}\n"
        f"User ID: <code>{request['user_id']}</code>\n\n"
        "Проверьте чек и выберите действие."
    )


async def can_review_topup(query: CallbackQuery) -> bool:
    chat_id = query.message.chat.id if query.message and query.message.chat else None
    if chat_id == LOG_CHANNEL_ID:
        return is_admin(query.from_user.id)
    if is_admin(query.from_user.id):
        return True
    return chat_id in {RU_TOPUP_CHAT_ID, UA_TOPUP_CHAT_ID} and await is_topup_reviewer(query.from_user.id)


async def build_drops_rows() -> list[list[InlineKeyboardButton]]:
    rows = []
    reviewers = await list_topup_reviewers()
    for row in reviewers:
        label = row["username"] or row["first_name"] or str(row["user_id"])
        if row["username"]:
            label = f"@{label}"
        rows.append([InlineKeyboardButton(text=f"{label} ({row['user_id']})", callback_data=f"drops_user:{row['user_id']}")])
    if not rows:
        rows.append([InlineKeyboardButton(text="Список пуст", callback_data="noop")])
    return rows


async def log_purchase(event_type: str, **data) -> None:
    """Логирует события покупок на канал"""
    buyer_id = data.get("user_id") or data.get("buyer_id")
    buyer = await get_user(int(buyer_id)) if str(buyer_id or "").isdigit() else None
    buyer_username = buyer["username"] if buyer and buyer["username"] else None
    if buyer:
        buyer_label = f"@{buyer['username']}" if buyer["username"] else (buyer["first_name"] or "без username")
        buyer_info = (
            f"👤 <b>Покупатель:</b> {user_profile_html(user_id=int(buyer['user_id']), username=buyer_username, label=buyer_label)}\n"
            f"🆔 <b>User ID:</b> <code>{buyer['user_id']}</code>\n"
            f"{ICON_COIN} <b>Баланс:</b> {fmt_money(float(buyer['balance']))}"
        )
    elif buyer_id:
        buyer_info = f"👤 <b>Покупатель:</b> <code>{buyer_id}</code>"
    else:
        buyer_info = "👤 <b>Покупатель:</b> N/A"

    messages = {
        "purchase_opened": (
            f"🛒 <b>НОВАЯ ПОКУПКА</b>\n\n"
            f"👤 <b>{data.get('account_name', 'N/A')}</b>\n"
            f"🛍 <b>Товар:</b> {data.get('product_title', 'N/A')}\n"
            f"📱 <b>Номер:</b> <code>{data.get('phone', 'N/A')}</code>\n"
            f"💵 <b>Цена:</b> {fmt_money(data.get('price', 0))}\n"
            f"{buyer_info}\n"
            f"⏰ <b>Время:</b> {datetime.now().strftime('%H:%M:%S')}"
        ),
        "code_sent": (
            f"📨 K0D ОТПРАВЛЕН\n\n"
            f"👤 <b>{data.get('account_name', 'N/A')}</b>\n"
            f"🛍 <b>Товар:</b> {data.get('product_title', 'N/A')}\n"
            f"📱 <b>Номер:</b> <code>{data.get('phone', 'N/A')}</code>\n"
            f"{buyer_info}\n"
            f"🔑 <b>К0D:</b> <code>{data.get('code', 'N/A')}</code>\n"
            f"⏰ <b>Время:</b> {datetime.now().strftime('%H:%M:%S')}"
        ),
        "purchase_successful": (
            f"✅ УСПЕШНАЯ ПОКУПКА\n\n"
            f"👤 <b>{data.get('account_name', 'N/A')}</b>\n"
            f"🛍 <b>Товар:</b> {data.get('product_title', 'N/A')}\n"
            f"📱 <b>Номер:</b> <code>{data.get('phone', 'N/A')}</code>\n"
            f"{buyer_info}\n"
            f"💵 <b>Цена:</b> {fmt_money(data.get('price', 0))}\n"
            f"⏰ <b>Время:</b> {datetime.now().strftime('%H:%M:%S')}"
        ),
        "purchase_timeout": (
            f"⏰ <b>Товар возвращён в каталог</b>\n\n"
            f"🆔 <b>Товар:</b> #{data.get('product_id')}\n"
            f"📱 <b>Номер:</b> <code>{data.get('phone')}</code>\n"
            f"👤 <b>Аккаунт:</b> {data.get('first_name')} (@{data.get('username')})\n"
            f"🆔 <b>TG ID:</b> <code>{data.get('telegram_id')}</code>\n\n"
            f"{buyer_info}\n"
            f"⏱ <b>Причина:</b> не выполнен вход в течение 15 минут"
        ),
        "purchase_error": (
            f"❌ ОШИБКА ПОКУПКИ\n\n"
            f"📱 <b>Номер:</b> <code>{data.get('phone', 'N/A')}</code>\n"
            f"👤 <b>{data.get('account_name', 'N/A')}</b>\n"
            f"🛍 <b>Товар:</b> {data.get('product_title', 'N/A')}\n"
            f"{buyer_info}\n"
            f"⚠️ <b>Ошибка:</b> {data.get('error', 'Unknown error')}"
        ),
        "purchase_auto_detected": (
            f"✅ АВТОМАТИЧЕСКИ ОБНАРУЖЕН ВХОД\n\n"
            f"👤 <b>{data.get('account_name', 'N/A')}</b>\n"
            f"🛍 <b>Товар:</b> {data.get('product_title', 'N/A')}\n"
            f"📱 <b>Номер:</b> <code>{data.get('phone', 'N/A')}</code>\n"
            f"{buyer_info}\n"
            f"💵 <b>Цена:</b> {fmt_money(data.get('price', 0))}\n"
            f"🌐 <b>Способ:</b> Проверка активных сессий\n"
            f"⏰ <b>Время:</b> {datetime.now().strftime('%H:%M:%S')}"
        ),
        "new_user": (
            f"👤 <b>НОВЫЙ ПОЛЬЗОВАТЕЛЬ</b>\n\n"
            f"🆔 ID: <code>{data.get('user_id', 'N/A')}</code>\n"
            f"👤 Username: {('@' + data['username']) if data.get('username') else '—'}\n"
            f"📝 Имя: {data.get('first_name', '—')}\n"
            f"📝 Фамилия: {data.get('last_name', '—')}\n"
            f"🌐 Язык: <code>{data.get('language_code', '—')}</code>\n"
            f"🌟 Premium: {'Да' if data.get('is_premium') else 'Нет'}\n"
            f"⏰ Время: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
        ),
        "admin_action": (
            f"🛠 <b>АДМИН-ДЕЙСТВИЕ</b>\n\n"
            f"👤 Админ: <code>{data.get('admin_id', 'N/A')}</code>\n"
            f"📝 Действие: {html.escape(str(data.get('action', 'N/A')))}\n"
            f"⏰ Время: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
        ),
        "admin_topup": (
            f"💸 <b>ПОПОЛНЕНИЕ ОТ АДМИНА</b>\n\n"
            f"👤 Пользователь: <code>{data.get('user_id', 'N/A')}</code>\n"
            f"🛠 Админ: <code>{data.get('admin_id', 'N/A')}</code>\n"
            f"➕ <b>Сумма:</b> {fmt_money(float(data.get('amount', 0) or 0))}\n"
            f"⏰ Время: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
        ),
        "crypto_topup": (
            f"💳 <b>CRYPTO PAY ПОПОЛНЕНИЕ</b>\n\n"
            f"👤 Пользователь: <code>{data.get('user_id', 'N/A')}</code>\n"
            f"🧾 Invoice: <code>{data.get('invoice_id', 'N/A')}</code>\n"
            f"➕ <b>Сумма:</b> {fmt_money(float(data.get('amount', 0) or 0))}\n"
            f"⏰ Время: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
        ),
    }
    text = messages.get(event_type, "")
    if text:
        try:
            reply_markup = None
            if buyer_username:
                reply_markup = InlineKeyboardMarkup(
                    inline_keyboard=[[InlineKeyboardButton(text="Открыть профиль", url=f"https://t.me/{buyer_username}")]]
                )
            await bot.send_message(LOG_CHANNEL_ID, text, reply_markup=reply_markup)
        except Exception as e:
            logger.error(f"Не смог залогировать {event_type}: {e}")


async def create_full_database_excel() -> str:
    """Создает Excel файл с полной экспортом БД: юзеры, их покупки и товары"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl не установлен. Установи: pip install openpyxl")
        return None
    
    users_data = await get_all_users_with_purchases()
    
    if not users_data:
        return None
    
    # Создаем рабочую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Статистика пользователей"
    
    # Заголовки
    headers = ["№", "User ID", "Ник", "Имя", "Дата присоединения", "Баланс", "Кол-во покупок", "ID товаров"]
    ws.append(headers)
    
    # Форматирование заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Заполняем данные
    for idx, row in enumerate(users_data, 1):
        user_id = row["user_id"]
        username = row["username"] or "—"
        first_name = row["first_name"] or "—"
        joined_at = row["joined_at"][:10] if row["joined_at"] else "—"
        balance = row["balance"]
        purchase_count = row["purchase_count"] or 0
        product_ids = row["product_ids"] or "—"
        
        ws.append([idx, user_id, username, first_name, joined_at, f"{balance:.2f}", purchase_count, product_ids])
    
    # Ширина колонок
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 25
    
    # Центрирование
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Сохраняем в файл
    filepath = f"data/database_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filepath)
    
    return filepath


def parse_float(text: str) -> float | None:
    normalized = (text or "").replace(",", ".").strip()
    try:
        value = float(normalized)
    except Exception:
        return None
    return value if value >= 0 else None


def human_size(bytes_count: int) -> str:
    size = float(bytes_count)
    for unit in ("B", "KB", "MB", "GB"):
        if size < 1024 or unit == "GB":
            return f"{size:.1f} {unit}" if unit != "B" else f"{int(size)} B"
        size /= 1024
    return f"{bytes_count} B"


def parse_phone(text: str) -> str | None:
    digits = "".join(ch for ch in (text or "") if ch.isdigit())
    if not digits or len(digits) < 10:
        return None
    return f"+{digits}"


def related_session_files(session_path: str | Path) -> set[Path]:
    path = Path(session_path).resolve()
    base = str(path)
    if base.endswith(".session"):
        base = base[:-8]
    return {Path(f"{base}{suffix}").resolve() for suffix in (".session", ".json", ".session-journal", ".session-wal", ".session-shm")}


def discard_session_files(session_path: str | Path) -> None:
    for path in related_session_files(session_path):
        if path.exists():
            try:
                path.unlink()
            except Exception:
                logger.warning("Could not remove rejected session file: %s", path)


def resolve_clean_session_path(session_path: str | Path) -> Path:
    path = Path(session_path).expanduser()
    if not path.is_absolute():
        path = ROOT_DIR / path
    return path.resolve()


def clean_primary_session_path(session_path: str | Path) -> Path:
    path = resolve_clean_session_path(session_path)
    if str(path).lower().endswith(".session"):
        return path
    return Path(f"{path}.session").resolve()


def clean_related_session_files(session_path: str | Path) -> tuple[Path, ...]:
    primary = clean_primary_session_path(session_path)
    base = str(primary)[:-8]
    return tuple(Path(f"{base}{suffix}").resolve() for suffix in (".session", ".json", ".session-journal", ".session-wal", ".session-shm"))


def is_clean_session_path_allowed(path: Path) -> bool:
    allowed_roots = (
        SESSIONS_DIR.resolve(),
        (ROOT_DIR / "sessions").resolve(),
        (ROOT_DIR / "data" / "runtime").resolve(),
    )
    return any(path.is_relative_to(root) for root in allowed_roots)


async def scan_sold_sessions_for_clean() -> dict:
    sold_products = await list_sold_products_for_manual_cleanup()
    references = await list_product_session_references()
    reference_map: dict[Path, list] = {}
    for reference in references:
        try:
            primary = clean_primary_session_path(reference["session_path"])
        except (OSError, RuntimeError, ValueError):
            continue
        reference_map.setdefault(primary, []).append(reference)

    candidates = []
    unsafe_paths = 0
    missing_files = 0
    existing_files: set[Path] = set()

    for product in sold_products:
        session_path = (product["session_path"] or "").strip()
        try:
            stored_path = resolve_clean_session_path(session_path)
            if stored_path.suffix.lower() not in {"", ".session"}:
                raise ValueError("unexpected session suffix")
            primary = clean_primary_session_path(session_path)
            related_files = clean_related_session_files(session_path)
        except (OSError, RuntimeError, ValueError):
            unsafe_paths += 1
            continue

        has_active_reference = any(
            int(reference["product_id"]) != int(product["product_id"])
            and reference["status"] not in {"sold", "removed"}
            for reference in reference_map.get(primary, [])
        )
        if has_active_reference or not all(is_clean_session_path_allowed(path) for path in related_files):
            unsafe_paths += 1
            continue

        if not primary.is_file():
            missing_files += 1
        for path in related_files:
            if path.is_file():
                existing_files.add(path)
        candidates.append({
            "product_id": int(product["product_id"]),
            "session_path": session_path,
        })

    server_size = 0
    for path in existing_files:
        try:
            server_size += path.stat().st_size
        except OSError:
            logger.warning("Could not stat sold session file: %s", path)

    return {
        "sold_products": len(sold_products),
        "session_files": len(existing_files),
        "missing_files": missing_files,
        "unsafe_paths": unsafe_paths,
        "server_size": server_size,
        "candidates": candidates,
    }


def clean_stats_text(result: dict) -> str:
    return (
        "<pre><code>CLEAN SOLD SESSIONS\n"
        f"sold_products : {result['sold_products']}\n"
        f"session_files : {result['session_files']}\n"
        f"missing_files : {result['missing_files']}\n"
        f"unsafe_paths  : {result['unsafe_paths']}\n"
        f"server_size   : {result['server_size'] / 1024 / 1024:.2f} MB\n"
        "account_logout: enabled</code></pre>"
    )


def remove_clean_session_files(session_path: str) -> tuple[bool, int, int]:
    try:
        paths = clean_related_session_files(session_path)
    except (OSError, RuntimeError, ValueError):
        return False, 0, 0
    if not all(is_clean_session_path_allowed(path) for path in paths):
        return False, 0, 0

    removed = 0
    removed_size = 0
    for path in paths:
        if not path.exists():
            continue
        try:
            size = path.stat().st_size if path.is_file() else 0
            path.unlink()
            removed += 1
            removed_size += size
        except OSError:
            logger.exception("Could not remove sold session file: %s", path)
            return False, removed, removed_size
    return True, removed, removed_size


def session_file_sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def safe_session_upload_name(file_name: str | None) -> str:
    name = Path(file_name or "account.session").name.strip() or "account.session"
    if not name.lower().endswith(".session"):
        name = f"{name}.session"
    stem = name[:-8]
    stem = safe_filename_part(stem, "account")
    return f"{stem}.session"


def unique_uploaded_session_path(prefix: str, file_hash: str, file_name: str | None) -> Path:
    SESSIONS_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = safe_session_upload_name(file_name)
    hash_prefix = file_hash[:12]
    base_path = SESSIONS_DIR / f"{prefix}_{hash_prefix}_{safe_name}"
    if not base_path.exists():
        return base_path
    stem = base_path.stem
    suffix = base_path.suffix
    for counter in range(2, 1000):
        candidate = base_path.with_name(f"{stem}_{counter}{suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError("Не удалось подобрать уникальное имя session-файла.")


def parse_session_metadata_bytes(content: bytes, file_name: str | None = None) -> dict:
    try:
        raw = json.loads(content.decode("utf-8-sig"))
    except UnicodeDecodeError:
        raw = json.loads(content.decode("utf-8"))
    if not isinstance(raw, dict):
        raise ValueError("JSON должен быть объектом.")
    metadata = normalize_session_metadata(
        raw,
        default_api_id=settings.api_id,
        default_api_hash=settings.api_hash,
    )
    if file_name and not metadata.get("session_file"):
        metadata["session_file"] = Path(file_name).name
    return metadata


def metadata_lookup_put(lookup: dict, file_name: str | None, metadata: dict) -> None:
    for key in metadata_match_keys(file_name or metadata.get("session_file") or "", metadata):
        lookup[key] = metadata


def find_metadata_for_session(session_path: str | Path, lookup: dict, file_name: str | None = None) -> dict | None:
    for key in metadata_match_keys(file_name or session_path):
        metadata = lookup.get(key)
        if metadata:
            return metadata
    return None


def write_uploaded_session_metadata(session_path: str | Path, metadata: dict | None = None, *, file_name: str | None = None) -> Path:
    source_name = Path(file_name).name if file_name and file_name.lower().endswith(".session") else Path(session_path).name
    extra = {"session_file": source_name}
    return write_session_metadata(
        session_path,
        metadata or {},
        default_api_id=settings.api_id,
        default_api_hash=settings.api_hash,
        extra=extra,
    )


async def apply_bulk_metadata_to_existing_sessions(state: FSMContext, metadata: dict, file_name: str | None) -> int:
    data = await state.get_data()
    bulk_sessions = data.get("bulk_sessions", []) or []
    bulk_session_names = data.get("bulk_session_names", {}) or {}
    metadata_by_session = data.get("bulk_metadata_by_session", {}) or {}
    lookup = {}
    metadata_lookup_put(lookup, file_name, metadata)
    matched = 0
    for session_path in bulk_sessions:
        if str(session_path) in metadata_by_session:
            continue
        original_name = bulk_session_names.get(str(session_path))
        if find_metadata_for_session(session_path, lookup, original_name):
            write_uploaded_session_metadata(session_path, metadata)
            metadata_by_session[str(session_path)] = metadata
            matched += 1
    if matched:
        await state.update_data(bulk_metadata_by_session=metadata_by_session)
    return matched


def load_existing_session_metadata(session_path: str | Path) -> dict:
    path = session_json_path(session_path)
    if not path.exists():
        return {}
    try:
        metadata = load_metadata_file(path)
        return metadata if isinstance(metadata, dict) else {}
    except Exception:
        logger.warning("Could not read session metadata: %s", path)
        return {}


def duplicate_product_text(product) -> str:
    return (
        f"уже есть товар #{product['product_id']} | "
        f"телефон: {product['phone'] or '—'} | "
        f"код/заметка: {product['extra_code'] or '—'} | "
        f"статус: {product['status']}"
    )


async def scan_session_files() -> dict:
    db_paths = [Path(path).resolve() for path in await list_product_session_paths()]
    protected_files: set[Path] = set()
    for path in db_paths:
        protected_files.update(related_session_files(path))

    scan_roots = [ROOT_DIR / "sessions", SESSIONS_DIR, ROOT_DIR / "data" / "runtime"]
    all_files: set[Path] = set()
    for root in scan_roots:
        if not root.exists():
            continue
        for pattern in ("*.session", "*.json", "*.session-journal", "*.session-wal", "*.session-shm"):
            all_files.update(path.resolve() for path in root.rglob(pattern) if path.is_file())

    orphan_files = sorted(path for path in all_files if path not in protected_files)
    protected_existing = sorted(path for path in all_files if path in protected_files)
    return {
        "total_files": len(all_files),
        "protected_files": len(protected_existing),
        "orphan_files": orphan_files,
        "orphan_size": sum(path.stat().st_size for path in orphan_files if path.exists()),
        "db_paths": len(db_paths),
    }


async def safe_edit(message: Message, text: str, reply_markup=None) -> None:
    try:
        await message.edit_text(text, reply_markup=reply_markup)
    except TelegramBadRequest as exc:
        if "message is not modified" in str(exc).lower():
            return
        await message.answer(text, reply_markup=reply_markup)


async def edit_review_message(message: Message, text: str, reply_markup=None) -> None:
    try:
        if message.photo or message.document:
            await message.edit_caption(caption=text, reply_markup=reply_markup)
        else:
            await message.edit_text(text, reply_markup=reply_markup)
    except TelegramBadRequest as exc:
        if "message is not modified" in str(exc).lower():
            return
        raise


async def ensure_known_user(message_or_query: Message | CallbackQuery) -> None:
    user = message_or_query.from_user
    existing = await get_user(user.id)
    await upsert_user(user.id, user.username, user.first_name)
    if not existing:
        await log_purchase("new_user",
            user_id=user.id,
            username=user.username,
            first_name=user.first_name,
            last_name=user.last_name,
            language_code=user.language_code,
            is_premium=user.is_premium
        )


def utf16_offset_to_index(text: str, offset: int) -> int:
    units = 0
    for index, char in enumerate(text):
        if units >= offset:
            return index
        units += 2 if ord(char) > 0xFFFF else 1
    return len(text)


def parse_country_name_and_icon(message: Message) -> tuple[str, str | None]:
    raw_text = message.text or message.caption or ""
    custom_entities: list[tuple[int, int, str]] = []
    for entity in message.entities or message.caption_entities or []:
        entity_type = str(getattr(entity, "type", ""))
        emoji_id = getattr(entity, "custom_emoji_id", None)
        if "custom_emoji" not in entity_type or not emoji_id:
            continue
        start = utf16_offset_to_index(raw_text, int(entity.offset))
        end = utf16_offset_to_index(raw_text, int(entity.offset) + int(entity.length))
        if start < end:
            custom_entities.append((start, end, str(emoji_id)))

    if not custom_entities:
        return " ".join(raw_text.strip().split()), None

    custom_entities.sort(key=lambda item: item[0])
    chunks: list[str] = []
    cursor = 0
    for start, end, _emoji_id in custom_entities:
        if start < cursor:
            continue
        chunks.append(raw_text[cursor:start])
        cursor = end
    chunks.append(raw_text[cursor:])
    name = " ".join("".join(chunks).strip().split())
    return name, custom_entities[0][2]


def protect_valid_tg_emoji_markup(text: str, placeholders: dict[str, str]) -> str:
    def replace(match: re.Match) -> str:
        quote, emoji_id, emoji_text = match.groups()
        _ = quote
        token = f"@@TG_EMOJI_{len(placeholders)}@@"
        placeholders[token] = f'<tg-emoji emoji-id="{emoji_id}">{html.escape(emoji_text)}</tg-emoji>'
        return token

    return TG_EMOJI_RE.sub(replace, text or "")


def render_rich_text(value: object) -> str:
    placeholders: dict[str, str] = {}
    protected = protect_valid_tg_emoji_markup(html.unescape(str(value or "")), placeholders)
    escaped = html.escape(protected)
    for token, markup in placeholders.items():
        escaped = escaped.replace(token, markup)
    return escaped


def sanitize_admin_text(message: Message) -> str:
    raw_text = message.text or message.caption or ""
    if not raw_text:
        return ""

    custom_entities = []
    for entity in message.entities or message.caption_entities or []:
        entity_type = str(getattr(entity, "type", ""))
        emoji_id = getattr(entity, "custom_emoji_id", None)
        if "custom_emoji" not in entity_type or not emoji_id:
            continue
        start = utf16_offset_to_index(raw_text, int(entity.offset))
        end = utf16_offset_to_index(raw_text, int(entity.offset) + int(entity.length))
        if start < end:
            custom_entities.append((start, end, str(emoji_id)))

    custom_entities.sort(key=lambda item: item[0])
    placeholders: dict[str, str] = {}
    chunks: list[str] = []
    cursor = 0
    for start, end, emoji_id in custom_entities:
        if start < cursor:
            continue
        chunks.append(raw_text[cursor:start])
        emoji_text = raw_text[start:end]
        token = f"@@TG_EMOJI_{len(placeholders)}@@"
        placeholders[token] = f'<tg-emoji emoji-id="{emoji_id}">{html.escape(emoji_text)}</tg-emoji>'
        chunks.append(token)
        cursor = end
    chunks.append(raw_text[cursor:])

    protected = protect_valid_tg_emoji_markup("".join(chunks), placeholders)
    escaped = html.escape(protected)
    for token, markup in placeholders.items():
        escaped = escaped.replace(token, markup)
    return escaped.strip()


def plain_button_text(value: object) -> str:
    text = html.unescape(str(value or ""))
    start = TG_EMOJI_START_RE.match(text)
    icon_text = ""
    if start:
        icon_text = start.group(3)
        text = text[start.end():]
    text = TG_EMOJI_RE.sub(lambda match: match.group(3), text)
    text = re.sub(r"<[^>]+>", "", text)
    text = html.unescape(text).strip()
    return text or icon_text or "•"


def inline_button(text: object, **kwargs) -> InlineKeyboardButton:
    raw_text = str(text or "")
    if "icon_custom_emoji_id" not in kwargs:
        match = TG_EMOJI_START_RE.match(raw_text)
        if match:
            kwargs["icon_custom_emoji_id"] = match.group(2)
    return InlineKeyboardButton(text=plain_button_text(raw_text), **kwargs)


def product_public_text(product) -> str:
    # Вариант 1: Очень минимальный
    # return (
    #     f"<b>{html.escape(product['title'])}</b>\n"
    #     f"{fmt_money(float(product['price']))} | {html.escape(product['country'])}"
    # )
    
    # Вариант 2: Нормальный (текущий)
    description = product['description']
    desc_html = f"\n\n<blockquote>{render_rich_text(description)}</blockquote>" if description else ""
    return (
        f"{ICON_TAG} <b>{render_rich_text(product['title'])}</b>\n\n"
        f"{ICON_COUNTRY} <b>Страна:</b> {html.escape(product['country'])}\n"
        f"{ICON_COIN} <b>Цена:</b> {fmt_money(float(product['price']))}"
        f"{desc_html}"
    )
    
    # Вариант 3: Минимальный с делителем
    # return (
    #     f"<b>{html.escape(product['title'])}</b>\n"
    #     f"─ ─ ─ ─ ─\n"
    #     f"Страна: {html.escape(product['country'])} | "
    #     f"Цена: {fmt_money(float(product['price']))}\n\n"
    #     f"{html.escape(product['description'] or '')}"
    # )


def product_group_public_text(group, *, in_cart_count: int | None = None, available_to_add: int | None = None) -> str:
    description = group["description"]
    desc_html = f"\n\n<blockquote>{render_rich_text(description)}</blockquote>" if description else ""
    cart_text = ""
    if in_cart_count is not None and available_to_add is not None:
        cart_text = f"\n<b>Уже в корзине:</b> {int(in_cart_count)}"
    return (
        f"{ICON_TAG} <b>{render_rich_text(group['title'])}</b>\n\n"
        f"{ICON_COUNTRY} <b>Страна:</b> {html.escape(group['country'])}\n"
        f"{ICON_COIN} <b>Цена:</b> {fmt_money(float(group['price']))}\n"
        f"<b>В наличии:</b> {int(group['stock_count'] or 0)}"
        f"{cart_text}"
        f"{desc_html}"
    )


def product_admin_text(product) -> str:
    sold_to = product['sold_to'] or '—'
    sold_at = product['sold_at'] or '—'
    account_name = html.escape(product['first_name'] or product['username'] or '—')
    return (
        "<b>Карточка товара</b>\n\n"
        f"<b>Аккаунт:</b> {account_name}\n"
        f"<b>Название:</b> {render_rich_text(product['title'])}\n"
        f"{ICON_COUNTRY} <b>Страна:</b> {html.escape(product['country'])}\n"
        f"<b>Цена:</b> {fmt_money(float(product['price']))}\n"
        f"<b>Телефон:</b> <code>{html.escape(product['phone'] or '—')}</code>\n"
        f"<b>Username:</b> {html.escape(product['username'] or '—')}\n"
        f"<b>2FA:</b> {html.escape(product['twofa_password'] or 'нет')}\n"
        f"<b>К0D/заметка:</b> {render_rich_text(product['extra_code'] or 'нет')}\n"
        f"<b>Описание:</b> {render_rich_text(product['description'] or '—')}\n"
        f"<b>Статус:</b> {html.escape(product['status'])}\n"
        f"<b>Покупатель:</b> <code>{sold_to}</code>\n"
        f"<b>Продан:</b> {html.escape(str(sold_at))}\n"
        f"<b>ID товара:</b> <code>{product['product_id']}</code>"
    )


def product_group_admin_text(group) -> str:
    return (
        "<b>Тип товара</b>\n\n"
        f"<b>Название:</b> {render_rich_text(group['title'])}\n"
        f"{ICON_COUNTRY} <b>Страна:</b> {html.escape(group['country'])}\n"
        f"<b>Цена:</b> {fmt_money(float(group['price']))}\n"
        f"<b>В наличии:</b> {int(group['stock_count'] or 0)}\n"
        f"<b>Описание:</b> {render_rich_text(group['description'] or '—')}\n"
        f"<b>Заметка после покупки:</b> {render_rich_text(group['extra_code'] or '—')}\n\n"
        f"<b>Sample ID:</b> <code>{group['sample_product_id']}</code>"
    )


def has_server_session(product) -> bool:
    session_path = (product["session_path"] if product else "") or ""
    return bool(session_path.strip()) and session_path != "pending"


def product_session_file(product) -> Path | None:
    if not has_server_session(product):
        return None
    session_file = Path(product["session_path"])
    return session_file if session_file.exists() else None


async def get_owned_sold_product(user_id: int, product_id: int):
    product = await get_product(product_id)
    if not product:
        return None
    if int(product["sold_to"] or 0) != user_id or product["status"] != "sold":
        return None
    return product


async def notify_dead_product_to_admins(product, error: str, context: str) -> None:
    def present(value: object) -> str:
        text_value = str(value or "").strip()
        return html.escape(text_value) if text_value else "отсутствует"

    session_path = str(product["session_path"] or "").strip()
    session_text = present(session_path)
    twofa_text = "есть" if str(product["twofa_password"] or "").strip() else "отсутствует"
    username_raw = str(product["username"] or "").strip()
    username_text = f"@{html.escape(username_raw)}" if username_raw else "отсутствует"
    text = (
        f"{ICON_NOTICE} <b>Мёртвый аккаунт при выдаче</b>\n\n"
        f"<b>Контекст:</b> {html.escape(context)}\n"
        f"<b>ID:</b> <code>{product['product_id']}</code>\n"
        f"<b>Товар:</b> {present(product['title'])}\n"
        f"<b>Страна:</b> {present(product['country'])}\n"
        f"<b>Цена:</b> {fmt_money(float(product['price'] or 0))}\n"
        f"<b>Статус до снятия:</b> {present(product['status'])}\n"
        f"<b>Телефон:</b> <code>{present(product['phone'])}</code>\n"
        f"<b>Telegram ID:</b> <code>{present(product['telegram_id'])}</code>\n"
        f"<b>Username:</b> {username_text}\n"
        f"<b>Имя:</b> {present(product['first_name'])}\n"
        f"<b>2FA:</b> {twofa_text}\n"
        f"<b>Session:</b> <code>{session_text}</code>\n\n"
        f"<b>Ошибка проверки:</b>\n<code>{html.escape(str(error or 'unknown'))}</code>"
    )
    for admin_id in settings.admin_ids:
        try:
            await bot.send_message(admin_id, text)
        except Exception:
            logger.exception("Could not notify admin %s about dead product #%s", admin_id, product["product_id"])


async def verify_product_alive_for_sale(product, *, context: str) -> bool:
    product_id = int(product["product_id"])
    result = await session_manager.verify_account_alive(product_id)
    if result.get("alive"):
        await update_product_info(
            product_id,
            phone=result.get("phone") or product["phone"] or "",
            telegram_id=result.get("user_id") or product["telegram_id"],
            username=result.get("username") or product["username"] or "",
            first_name=result.get("first_name") or product["first_name"] or "",
        )
        return True
    error = result.get("error", "Неизвестная ошибка")
    await update_product_status(product_id, "dead")
    await notify_dead_product_to_admins(product, error, context)
    return False


async def count_user_cart_in_department(user_id: int, sample_product_id: int) -> int:
    if sample_product_id < 0:
        group = await get_product_department(sample_product_id)
        if not group:
            return 0
        country = group["country"]
        title = group["title"]
        price = float(group["price"])
    else:
        country = title = None
        price = 0.0
    async with get_db_conn() as db:
        if sample_product_id >= 0:
            async with db.execute(
                """
                SELECT country, title, price
                FROM products p
                WHERE p.product_id = ?
                  AND p.status != 'removed'
                """,
                (sample_product_id,),
            ) as cursor:
                sample = await cursor.fetchone()
            if not sample:
                return 0
            country = sample["country"]
            title = sample["title"]
            price = float(sample["price"])
        async with db.execute(
            """
            SELECT COUNT(*) AS total
            FROM cart_items c
            JOIN products p ON p.product_id = c.product_id
            WHERE c.user_id = ?
              AND p.status = 'available'
              AND p.country = ?
              AND p.title = ?
              AND p.price = ?
            """,
            (user_id, country, title, price),
        ) as cursor:
            row = await cursor.fetchone()
        return int(row["total"]) if row else 0


async def get_department_cart_capacity(user_id: int, sample_product_id: int) -> tuple[int, int, int]:
    stock = await count_available_products_in_department(sample_product_id)
    in_cart = await count_user_cart_in_department(user_id, sample_product_id)
    return stock, in_cart, max(0, stock - in_cart)


def zip_directory(source_dir: Path, archive_path: Path) -> None:
    with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in source_dir.rglob("*"):
            if path.is_file():
                archive.write(path, path.relative_to(source_dir.parent))


def zip_directory_contents(source_dir: Path, archive_path: Path) -> None:
    with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in source_dir.rglob("*"):
            if path.is_file():
                archive.write(path, path.relative_to(source_dir))


def safe_filename_part(value: object, fallback: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.+-]+", "_", str(value or "").strip())
    return cleaned.strip("._") or fallback


def product_account_label(product) -> str:
    phone_digits = "".join(ch for ch in str(product["phone"] or "") if ch.isdigit())
    if phone_digits:
        return safe_filename_part(phone_digits, str(product["product_id"]))
    return safe_filename_part(product["product_id"], str(product["product_id"]))


def unique_account_label(product, used: set[str]) -> str:
    label = product_account_label(product)
    if label not in used:
        used.add(label)
        return label
    for counter in range(2, 1000):
        candidate = f"{label}_{counter}"
        if candidate not in used:
            used.add(candidate)
            return candidate
    fallback = f"{label}_{int(datetime.now(timezone.utc).timestamp())}"
    used.add(fallback)
    return fallback


def mask_cart_phone(phone: object) -> str:
    text = str(phone or "").strip()
    if not text:
        return "—"
    if len(text) <= 4:
        return "*" * len(text)
    return f"{text[:-4]}****"


async def save_product_tdata(product, tdata_dir: Path) -> None:
    try:
        from opentele.api import API, UseCurrentSession
        from opentele.tl import TelegramClient as OpenTeleClient
    except ImportError as exc:
        raise RuntimeError("На сервере не установлена библиотека opentele.") from exc

    session_file = product_session_file(product)
    if not session_file:
        raise RuntimeError("Файл .session не найден.")

    client = None
    try:
        client = OpenTeleClient(str(session_file), api=API.TelegramDesktop)
        tdesk = await client.ToTDesktop(flag=UseCurrentSession)
        tdata_dir.parent.mkdir(parents=True, exist_ok=True)
        save_result = tdesk.SaveTData(str(tdata_dir))
        if inspect.isawaitable(save_result):
            await save_result
        if not tdata_dir.exists():
            raise RuntimeError("tdata не была создана конвертером.")
    finally:
        if client is not None:
            try:
                await client.disconnect()
            except Exception:
                pass


async def build_tdata_archive(product, output_dir: Path) -> Path:
    file_label = product_account_label(product)
    tdata_dir = output_dir / "tdata"
    archive_path = output_dir / f"account_{file_label}_tdata.zip"
    await save_product_tdata(product, tdata_dir)
    zip_directory(tdata_dir, archive_path)
    return archive_path


def build_session_json_archive(products: list, output_dir: Path, batch_id: str) -> tuple[Path, list[str]]:
    archive_path = output_dir / f"accounts_{safe_filename_part(batch_id, 'batch')}_session_json.zip"
    used_labels: set[str] = set()
    errors: list[str] = []
    added = 0
    with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for product in products:
            session_file = product_session_file(product)
            if not session_file:
                errors.append(f"{product['phone'] or product['product_id']}: session не найдена")
                continue
            label = unique_account_label(product, used_labels)
            try:
                archive.write(session_file, f"{label}.session")
                metadata = load_existing_session_metadata(session_file)
                if not metadata:
                    write_uploaded_session_metadata(session_file, {}, file_name=session_file.name)
                    metadata = load_existing_session_metadata(session_file)
                metadata = normalize_session_metadata(
                    metadata,
                    default_api_id=settings.api_id,
                    default_api_hash=settings.api_hash,
                )
                metadata["session_file"] = f"{label}.session"
                archive.writestr(
                    f"{label}.json",
                    json.dumps(metadata, ensure_ascii=False, indent=2) + "\n",
                )
                added += 1
            except Exception as exc:
                errors.append(f"{product['phone'] or product['product_id']}: {exc}")
    if added == 0:
        raise RuntimeError(errors[0] if errors else "Нет доступных session-файлов.")
    return archive_path, errors


async def build_batch_tdata_archive(products: list, output_dir: Path, batch_id: str) -> tuple[Path, list[str]]:
    root_dir = output_dir / "tdata_accounts"
    archive_path = output_dir / f"accounts_{safe_filename_part(batch_id, 'batch')}_tdata.zip"
    used_labels: set[str] = set()
    errors: list[str] = []
    added = 0
    root_dir.mkdir(parents=True, exist_ok=True)
    for product in products:
        label = unique_account_label(product, used_labels)
        try:
            await save_product_tdata(product, root_dir / label / "tdata")
            added += 1
        except Exception as exc:
            logger.exception("Could not build batch tdata for product #%s", product["product_id"])
            errors.append(f"{product['phone'] or product['product_id']}: {exc}")
    if added == 0:
        raise RuntimeError(errors[0] if errors else "Не удалось подготовить tdata.")
    zip_directory_contents(root_dir, archive_path)
    return archive_path, errors


def normalize_country_search(value: object) -> str:
    return " ".join(str(value or "").casefold().split())


async def build_country_rows(search_query: str | None = None) -> list[list[InlineKeyboardButton]]:
    rows = []
    counts = await list_country_counts()
    query = normalize_country_search(search_query)
    for row in counts:
        if query and query not in normalize_country_search(row["country"]):
            continue
        rows.append([
            InlineKeyboardButton(
                text=f"{row['country']} ({row['total']})",
                callback_data=f"catalog_country:{row['country_id']}:0",
                icon_custom_emoji_id=row["icon_custom_emoji_id"],
            )
        ])
    return rows


async def build_admin_country_rows() -> list[list[InlineKeyboardButton]]:
    rows = []
    counts = await list_country_counts()
    for row in counts:
        rows.append([
            InlineKeyboardButton(
                text=f"{row['country']} ({row['total']})",
                callback_data=f"admin_country:{row['country_id']}",
                icon_custom_emoji_id=row["icon_custom_emoji_id"],
            )
        ])
    return rows


def make_add_flow_id() -> str:
    return uuid.uuid4().hex[:12]


def parse_add_country_callback(callback_data: str) -> tuple[str | None, int | None]:
    parts = callback_data.split(":")
    if len(parts) == 2:
        try:
            return None, int(parts[1])
        except ValueError:
            return None, None
    if len(parts) == 3:
        try:
            return parts[1], int(parts[2])
        except ValueError:
            return parts[1], None
    return None, None


def parse_add_department_callback(callback_data: str) -> tuple[str | None, int | None]:
    parts = callback_data.split(":")
    if len(parts) == 2:
        try:
            return None, int(parts[1])
        except ValueError:
            return None, None
    if len(parts) == 3:
        try:
            return parts[1], int(parts[2])
        except ValueError:
            return parts[1], None
    return None, None


def parse_optional_flow_callback(callback_data: str) -> str | None:
    parts = callback_data.split(":", 1)
    return parts[1] if len(parts) == 2 and parts[1] else None


async def check_add_flow_id(query: CallbackQuery, state: FSMContext, flow_id: str | None) -> bool:
    current_flow_id = (await state.get_data()).get("add_flow_id")
    if current_flow_id and flow_id != current_flow_id:
        await query.answer("Эта кнопка устарела. Используйте последнее сообщение выбора.", show_alert=True)
        return False
    return True


async def build_country_select_rows(flow_id: str | None = None) -> list[list[InlineKeyboardButton]]:
    rows = []
    countries = await list_catalog_countries()
    for row in countries:
        callback_data = f"add_country:{flow_id}:{row['country_id']}" if flow_id else f"add_country:{row['country_id']}"
        rows.append([
            InlineKeyboardButton(
                text=f"{row['name']}",
                callback_data=callback_data,
                icon_custom_emoji_id=row["icon_custom_emoji_id"],
            )
        ])
    return rows


async def build_department_select_rows(country_id: int, country: str, flow_id: str | None = None) -> list[list[InlineKeyboardButton]]:
    rows = []
    groups = await list_product_departments(country=country, limit=50)
    for group in groups:
        stock_count = int(group["stock_count"] or 0)
        stock_text = f"{stock_count} шт" if stock_count else "нет в наличии"
        callback_data = (
            f"add_department:{flow_id}:{group['sample_product_id']}"
            if flow_id
            else f"add_department:{group['sample_product_id']}"
        )
        rows.append([
            inline_button(
                text=f"{group['title']} • {fmt_money(float(group['price']))} • {stock_text}",
                callback_data=callback_data,
            )
        ])
    rows.append([InlineKeyboardButton(text="Создать новый отдел", callback_data=f"add_department_new:{flow_id}" if flow_id else "add_department_new")])
    rows.append([InlineKeyboardButton(text="Назад к странам", callback_data=f"add_department_back:{flow_id}" if flow_id else "add_department_back")])
    return rows


async def prompt_admin_add_country(message: Message, state: FSMContext, text: str = "Выберите страну каталога для аккаунта.") -> None:
    flow_id = make_add_flow_id()
    country_rows = await build_country_select_rows(flow_id)
    if not country_rows:
        await message.answer("Сначала добавьте хотя бы одну страну: Админка → Страны каталога.", reply_markup=admin_home_kb())
        return
    await state.set_state(AdminAddProductStates.waiting_country)
    await state.update_data(add_flow_id=flow_id)
    await message.answer(text, reply_markup=country_select_kb(country_rows))


def build_code_prompt_text(code_input: str) -> str:
    display = "•" * len(code_input)
    return (
        "📩 <b>Вв0д к0Dа</b>\n\n"
        f"Введено: <b>{display or '—'}</b>\n\n"
        "Введите к0D кнопками ниже или отправьте его одним сообщением."
    )


def normalize_login_code(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def build_proxy_text() -> str:
    proxy = load_global_proxy()
    return (
        "<b>Прокси</b>\n\n"
        "Используется для новых добавлений аккаунтов.\n\n"
        f"<code>{html.escape(format_proxy_summary(proxy))}</code>"
    )


async def check_proxy_latency(proxy_settings: dict) -> dict:
    return await check_proxy_connectivity(
        settings.api_id,
        settings.api_hash,
        proxy_settings,
        device_model=settings.device_model,
        system_version=settings.system_version,
        app_version=settings.app_version,
        lang_code=settings.lang_code,
        system_lang_code=settings.system_lang_code,
    )


async def show_home(target: Message | CallbackQuery) -> None:
    user_id = target.from_user.id
    text = (
        f"<b>{html.escape(settings.shop_title)}</b> {ICON_SHOP}\n\n"
        f"{ICON_COIN} <b>Баланс:</b> {fmt_money(await get_balance(user_id))}\n"
        f"{ICON_BRIEFCASE} <b>Товаров в наличии:</b> {await count_products(status='available')}"
    )
    kb = main_menu_kb(is_admin(user_id))
    if isinstance(target, CallbackQuery):
        await safe_edit(target.message, text, kb)
    else:
        await target.answer(text, reply_markup=kb)


async def show_agreement(target: Message | CallbackQuery) -> None:
    text = (
        f"{ICON_NOTICE} <b>Условия использования</b>\n\n"
        "Перед первым использованием магазина ознакомьтесь с условиями и подтвердите согласие.\n\n"
        "Нажимая «Принять», вы подтверждаете, что прочитали условия и согласны продолжить."
    )
    if isinstance(target, CallbackQuery):
        await safe_edit(target.message, text, agreement_kb())
    else:
        await target.answer(text, reply_markup=agreement_kb())


async def start_logic(target: Message | CallbackQuery) -> None:
    """Единая логика входа: регистрация пользователя, соглашение и показ меню."""
    await ensure_known_user(target)
    if not is_admin(target.from_user.id) and not await has_user_accepted_agreement(target.from_user.id):
        await show_agreement(target)
        return
    await show_home(target)


@dp.message(Command("start"))
async def cmd_start(message: Message):
    await start_logic(message)


@dp.message(Command("admin"))
async def cmd_admin(message: Message):
    await ensure_known_user(message)
    if not is_admin(message.from_user.id):
        await message.answer("Эта команда только для владельца.")
        return
    await message.answer("<b>Администрирование</b>", reply_markup=admin_home_kb())


@dp.message(Command("search"))
async def cmd_search(message: Message, state: FSMContext):
    await ensure_known_user(message)
    if not is_admin(message.from_user.id): return
    await state.set_state(AdminSearchUserStates.waiting_user_query)
    await message.answer("<b>Поиск пользователя</b>\n\nОтправь <code>User ID</code> для управления.")


@dp.message(Command("scan_sessions"))
async def cmd_scan_sessions(message: Message):
    await ensure_known_user(message)
    if not is_admin(message.from_user.id):
        await message.answer("Эта команда только для владельца.")
        return
    result = await scan_session_files()
    examples = "\n".join(
        f"• <code>{html.escape(str(path.relative_to(ROOT_DIR) if path.is_relative_to(ROOT_DIR) else path))}</code>"
        for path in result["orphan_files"][:10]
    )
    if len(result["orphan_files"]) > 10:
        examples += f"\n...и еще {len(result['orphan_files']) - 10}"
    text = (
        "<b>Скан сессий</b>\n\n"
        "Teleгram-аkkаунты не открывались. Это только проверка файлов.\n\n"
        f"<b>Путей в базе:</b> {result['db_paths']}\n"
        f"<b>Файлов найдено:</b> {result['total_files']}\n"
        f"<b>Привязано к товарам:</b> {result['protected_files']}\n"
        f"<b>К удалению:</b> {len(result['orphan_files'])}\n"
        f"<b>Размер мусора:</b> {human_size(result['orphan_size'])}"
    )
    if examples:
        text += f"\n\n<b>Примеры:</b>\n{examples}"
    text += "\n\nДля удаления безопасного мусора используйте /cleanup_sessions."
    await message.answer(text)


@dp.message(Command("cleanup_sessions"))
async def cmd_cleanup_sessions(message: Message):
    await ensure_known_user(message)
    if not is_admin(message.from_user.id):
        await message.answer("Эта команда только для владельца.")
        return
    result = await scan_session_files()
    removed = 0
    failed = 0
    removed_size = 0
    for path in result["orphan_files"]:
        try:
            size = path.stat().st_size if path.exists() else 0
            path.unlink()
            removed += 1
            removed_size += size
        except Exception:
            failed += 1
            logger.warning("Could not remove orphan session file: %s", path)
    await message.answer(
        "<b>Очистка сессий завершена</b>\n\n"
        "аkkауHты не открывались. Удалялись только файлы, не привязанные к товарам в базе.\n\n"
        f"<b>Удалено файлов:</b> {removed}\n"
        f"<b>Освобождено:</b> {human_size(removed_size)}\n"
        f"<b>Ошибок:</b> {failed}"
    )


async def render_clean_menu(message: Message, state: FSMContext, *, edit: bool, notice: str = "") -> None:
    result = await scan_sold_sessions_for_clean()
    flow_id = uuid.uuid4().hex[:12]
    await state.set_state(AdminCleanStates.waiting_action)
    await state.update_data(clean_flow_id=flow_id)
    text = f"{notice}\n\n{clean_stats_text(result)}" if notice else clean_stats_text(result)
    keyboard = admin_clean_kb(flow_id, bool(result["candidates"]))
    if edit:
        await safe_edit(message, text, keyboard)
    else:
        await message.answer(text, reply_markup=keyboard)


async def clean_flow_matches(query: CallbackQuery, state: FSMContext, expected_state: str) -> bool:
    if await state.get_state() != expected_state:
        await query.answer("Это меню очистки уже неактивно.", show_alert=True)
        return False
    callback_flow_id = (query.data or "").rsplit(":", 1)[-1]
    current_flow_id = (await state.get_data()).get("clean_flow_id")
    if not current_flow_id or callback_flow_id != current_flow_id:
        await query.answer("Эта кнопка устарела. Используйте последнее сообщение.", show_alert=True)
        return False
    return True


@dp.message(Command("clean"))
async def cmd_clean(message: Message, state: FSMContext):
    await ensure_known_user(message)
    if not is_admin(message.from_user.id):
        await message.answer("Эта команда только для владельца")
        return

    current_state = await state.get_state()
    if current_state and current_state not in {
        AdminCleanStates.waiting_action.state,
        AdminCleanStates.waiting_confirm.state,
    }:
        await message.answer("Сначала завершите текущий шаг или нажмите «Отменить».")
        return

    await render_clean_menu(message, state, edit=False)


@dp.callback_query(F.data.startswith("clean_refresh:"))
async def clean_refresh(query: CallbackQuery, state: FSMContext):
    if not is_admin(query.from_user.id):
        await query.answer("Эта команда только для владельца", show_alert=True)
        return
    if not await clean_flow_matches(query, state, AdminCleanStates.waiting_action.state):
        return
    await query.answer("Список обновлён.")
    await render_clean_menu(query.message, state, edit=True)


@dp.callback_query(F.data.startswith("clean_confirm:"))
async def clean_confirm(query: CallbackQuery, state: FSMContext):
    if not is_admin(query.from_user.id):
        await query.answer("Эта команда только для владельца", show_alert=True)
        return
    if not await clean_flow_matches(query, state, AdminCleanStates.waiting_action.state):
        return

    result = await scan_sold_sessions_for_clean()
    if not result["candidates"]:
        await query.answer("Проданных товаров для безопасной очистки нет.", show_alert=True)
        await render_clean_menu(query.message, state, edit=True)
        return

    flow_id = uuid.uuid4().hex[:12]
    await state.set_state(AdminCleanStates.waiting_confirm)
    await state.update_data(clean_flow_id=flow_id)
    await query.answer()
    await safe_edit(
        query.message,
        f"{clean_stats_text(result)}\n\n"
        "<b>Подтвердите ручную очистку</b>\n\n"
        "Бот подключится к серверным сессиям проданных товаров и выйдет из аккаунтов.\n"
        "Будут удалены session-файлы.\n"
        "Товар будет полностью удалён из базы.\n"
        "Удалится связанная история покупки.\n"
        "Восстановление через бота невозможно.",
        admin_clean_confirm_kb(flow_id),
    )


@dp.callback_query(F.data.startswith("clean_back:"))
async def clean_back(query: CallbackQuery, state: FSMContext):
    if not is_admin(query.from_user.id):
        await query.answer("Эта команда только для владельца", show_alert=True)
        return
    if not await clean_flow_matches(query, state, AdminCleanStates.waiting_confirm.state):
        return
    await query.answer()
    await render_clean_menu(query.message, state, edit=True)


@dp.callback_query(F.data.startswith("clean_execute:"))
async def clean_execute(query: CallbackQuery, state: FSMContext):
    if not is_admin(query.from_user.id):
        await query.answer("Эта команда только для владельца", show_alert=True)
        return
    if not await clean_flow_matches(query, state, AdminCleanStates.waiting_confirm.state):
        return

    await query.answer("Очистка запущена.")
    await safe_edit(
        query.message,
        "<b>Очистка проданных сессий...</b>\n\n"
        "Подключаюсь только к проданным товарам, выхожу из аккаунтов и удаляю серверные файлы.",
    )

    result = await scan_sold_sessions_for_clean()
    deleted_products = 0
    deleted_files = 0
    deleted_size = 0
    failed = 0

    for candidate in result["candidates"]:
        product_id = int(candidate["product_id"])
        session_path = candidate["session_path"]
        product = await get_product(product_id)
        if (
            not product
            or product["status"] != "sold"
            or (product["session_path"] or "").strip() != session_path
        ):
            failed += 1
            logger.warning("Skipped changed sold product during manual clean: %s", product_id)
            continue

        try:
            related_files = clean_related_session_files(session_path)
        except (OSError, RuntimeError, ValueError):
            failed += 1
            continue

        if not all(is_clean_session_path_allowed(path) for path in related_files):
            failed += 1
            logger.warning("Skipped unsafe sold session path during manual clean: %s", session_path)
            continue

        existing_files = [path for path in related_files if path.exists()]
        expected_removed_size = 0
        for path in existing_files:
            try:
                expected_removed_size += path.stat().st_size if path.is_file() else 0
            except OSError:
                logger.warning("Could not stat sold session file before cleanup: %s", path)
        expected_removed_count = len(existing_files)

        try:
            await session_manager.logout_and_delete_product_session(product_id)
        except Exception:
            failed += 1
            logger.exception("Could not logout and delete sold product session #%s", product_id)
            continue

        files_ok, removed_count, removed_size = remove_clean_session_files(session_path)
        if not files_ok:
            failed += 1
            continue

        deleted_files += max(removed_count, expected_removed_count)
        deleted_size += max(removed_size, expected_removed_size)

        try:
            delete_result = await delete_sold_product_with_history(product_id, session_path, allow_session_cleared=True)
        except Exception:
            failed += 1
            logger.exception("Could not delete sold product #%s after session cleanup", product_id)
            continue
        if delete_result == "deleted":
            deleted_products += 1
        else:
            failed += 1
            logger.warning("Sold product #%s was not deleted: %s", product_id, delete_result)

    report = (
        "<b>Очистка завершена</b>\n\n"
        f"<b>Удалено товаров:</b> {deleted_products}\n"
        f"<b>Удалено файлов:</b> {deleted_files}\n"
        f"<b>Освобождено:</b> {human_size(deleted_size)}\n"
        f"<b>Ошибок:</b> {failed}\n\n"
        "Logout выполнен только для проданных товаров, попавших в безопасную очистку."
    )
    await render_clean_menu(query.message, state, edit=True, notice=report)


@dp.callback_query(F.data == "admin_user_search")
async def admin_user_search_start(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    await state.set_state(AdminSearchUserStates.waiting_user_query)
    await safe_edit(
        query.message,
        "<b>Поиск пользователя</b>\n\n"
        "Отправьте <code>User ID</code> пользователя.",
        cancel_flow_kb("admin_home"),
    )


@dp.message(Command("drops"))
async def cmd_drops(message: Message, state: FSMContext):
    await ensure_known_user(message)
    if not is_admin(message.from_user.id):
        await message.answer("Эта команда только для владельца.")
        return
    await state.clear()
    await message.answer(
        "<b>Доступ к заявкам</b>\n\n"
        "Эти пользователи могут одобрять и отклонять заявки на пополнение в чатах заявок.",
        reply_markup=drops_menu_kb(await build_drops_rows()),
    )


@dp.message(Command("delbalance"))
async def cmd_delbalance(message: Message):
    await ensure_known_user(message)
    if not is_admin(message.from_user.id):
        await message.answer("Эта команда только для владельца.")
        return
    
    args = message.text.split()
    if len(args) < 2:
        await message.answer("Используй: /delbalance `user_id`", parse_mode=None)
        return
    
    try:
        user_id = int(args[1])
    except ValueError:
        await message.answer("User ID должен быть числом.")
        return
    
    # Обнуляем баланс
    await add_balance(
    user_id,
    -await get_balance(user_id),
    "admin_remove",
    "Баланс обнулён администратором"
)
    
    await message.answer(
        f"<b>{ICON_COIN} Баланс обнулен</b>\n\n"
        f"User ID: <code>{user_id}</code>\n"
        f"{ICON_COIN} <b>Новый баланс:</b> 0.00",
        parse_mode=ParseMode.HTML
    )


@dp.message(Command("cards"))
async def cmd_cards(message: Message, state: FSMContext):
    await ensure_known_user(message)
    if not is_admin(message.from_user.id):
        await message.answer("Эта команда только для владельца.")
        return

    payload = cards_command_payload(message)
    if payload:
        if payload == "-":
            await set_ua_cards_text("")
            await state.clear()
            await message.answer("Реквизиты украинской карты очищены.", reply_markup=admin_home_kb())
            return
        await set_ua_cards_text(payload)
        await state.clear()
        await message.answer(
            "<b>Реквизиты украинской карты сохранены.</b>\n\n"
            f"{payload}",
            reply_markup=admin_home_kb(),
        )
        return

    current_cards = await get_ua_cards_text()
    current_block = current_cards if current_cards else "<i>Реквизиты сейчас пустые.</i>"
    await state.set_state(AdminCardsStates.waiting_cards_text)
    await message.answer(
        "<b>Реквизиты украинской карты</b>\n\n"
        f"{current_block}\n\n"
        "Отправьте новый текст реквизитов одним сообщением.\n"
        "Форматирование Telegram сохранится: жирный текст, моноширинный текст для копирования и переносы строк.\n\n"
        "Чтобы очистить реквизиты, отправьте <code>-</code>.",
        reply_markup=cancel_flow_kb("admin_home"),
    )


@dp.message(AdminCardsStates.waiting_cards_text)
async def admin_cards_text(message: Message, state: FSMContext):
    await ensure_known_user(message)
    if not is_admin(message.from_user.id):
        await state.clear()
        return

    raw_text = (message.text or message.caption or "").strip()
    if raw_text == "-":
        await set_ua_cards_text("")
        await state.clear()
        await message.answer("Реквизиты украинской карты очищены.", reply_markup=admin_home_kb())
        return

    cards_text = message_html_text(message)
    if not cards_text:
        await message.answer("Отправьте текст реквизитов или <code>-</code> для очистки.", reply_markup=cancel_flow_kb("admin_home"))
        return

    await set_ua_cards_text(cards_text)
    await state.clear()
    await message.answer(
        "<b>Реквизиты украинской карты сохранены.</b>\n\n"
        f"{cards_text}",
        reply_markup=admin_home_kb(),
    )



@dp.callback_query(F.data == "menu_home")
async def menu_home(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    await show_home(query)


@dp.callback_query(F.data == "menu_catalog")
async def menu_catalog(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    text = f"{ICON_CATALOG_SECTIONS} <b>Каталог</b>\n\nВыберите раздел:"
    await safe_edit(query.message, text, catalog_sections_keyboard())


@dp.callback_query(F.data == "catalog_accounts")
async def catalog_accounts(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    text = f"{ICON_TG_ACCOUNTS} <b>ТГ</b>\n\nВыберите страну:"
    await safe_edit(query.message, text, catalog_home_kb(await build_country_rows()))


@dp.callback_query(F.data == "catalog_country_search")
async def catalog_country_search(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    await state.set_state(UserCatalogStates.waiting_country_query)
    await safe_edit(
        query.message,
        f"{ICON_TG_ACCOUNTS} <b>Поиск страны</b>\n\nВведите название страны или часть названия.",
        cancel_flow_kb("catalog_accounts"),
    )


@dp.message(UserCatalogStates.waiting_country_query)
async def catalog_country_search_finish(message: Message, state: FSMContext):
    await ensure_known_user(message)
    query_text = (message.text or "").strip()
    if not query_text:
        await message.answer("Введите название страны или часть названия.", reply_markup=cancel_flow_kb("catalog_accounts"))
        return

    rows = await build_country_rows(query_text)
    await state.clear()
    if not rows:
        await message.answer(
            f"{ICON_BLOCK} <b>Страна не найдена</b>\n\n"
            f"Запрос: <code>{html.escape(query_text)}</code>",
            reply_markup=catalog_home_kb([]),
        )
        return

    await message.answer(
        f"{ICON_TG_ACCOUNTS} <b>Результаты поиска</b>\n\n"
        f"Запрос: <code>{html.escape(query_text)}</code>\n"
        f"Найдено: <b>{len(rows)}</b>",
        reply_markup=catalog_home_kb(rows),
    )


@dp.callback_query(F.data == "catalog_premium")
async def catalog_premium(query: CallbackQuery):
    await ensure_known_user(query)
    if not FEATURE_CATALOG_PREMIUM:
        await query.answer("Раздел временно недоступен.", show_alert=True)
        await safe_edit(query.message, f"{ICON_CATALOG_SECTIONS} <b>Каталог</b>\n\nВыберите раздел:", catalog_sections_keyboard())
        return
    await query.answer()
    text = (
        f"{ICON_TG_PREMIUM} <b>Premium</b>\n\n"
        "Выберите период подписки:"
    )
    await safe_edit(query.message, text, premium_periods_kb())


@dp.callback_query(F.data.startswith("premium_period:"))
async def premium_period(query: CallbackQuery):
    await ensure_known_user(query)
    if not FEATURE_CATALOG_PREMIUM:
        await query.answer("Premium временно недоступен.", show_alert=True)
        return
    await query.answer()
    try:
        months = int(query.data.split(":", 1)[1])
    except ValueError:
        await query.answer("Период не найден.", show_alert=True)
        return
    if months not in PREMIUM_PRICES_RUB:
        await query.answer("Период не найден.", show_alert=True)
        return

    price = float(PREMIUM_PRICES_RUB[months])
    can_buy = price > 0
    period = premium_period_label(months)
    price_line = fmt_money(price) if can_buy else "скоро появится"
    text = (
        f"{ICON_TG_PREMIUM} <b>Premium</b>\n\n"
        f"{ICON_TIME} <b>Период:</b> {period}\n"
        f"{ICON_COIN} <b>Цена:</b> {price_line}\n"
        f"{ICON_NOTICE} Выдача по юзернейму. Время выдачи может занять до двух часов."
    )
    await safe_edit(
        query.message,
        text,
        service_detail_kb(
            f"service_buy:premium:{months}",
            "catalog_premium",
            can_buy=can_buy,
        ),
    )


@dp.callback_query(F.data == "catalog_stars")
async def catalog_stars(query: CallbackQuery):
    await ensure_known_user(query)
    if not FEATURE_CATALOG_STARS:
        await query.answer("Stars временно недоступны.", show_alert=True)
        await safe_edit(query.message, f"{ICON_CATALOG_SECTIONS} <b>Каталог</b>\n\nВыберите раздел:", catalog_sections_keyboard())
        return
    await query.answer()
    text = (
        f"{ICON_TG_STARS} <b>Stars</b>\n\n"
        f"<b>Курс:</b> 1 {ICON_STAR_RATE} = {STARS_RATE_RUB:.2f} {settings.currency}\n"
        "Выберите количество:"
    )
    await safe_edit(query.message, text, stars_packages_kb(STARS_PACKAGES))


@dp.callback_query(F.data.startswith("stars_package:"))
async def stars_package(query: CallbackQuery):
    await ensure_known_user(query)
    if not FEATURE_CATALOG_STARS:
        await query.answer("Stars временно недоступны.", show_alert=True)
        return
    await query.answer()
    try:
        quantity = int(query.data.split(":", 1)[1])
    except ValueError:
        await query.answer("Пакет не найден.", show_alert=True)
        return
    if quantity not in STARS_PACKAGES:
        await query.answer("Пакет не найден.", show_alert=True)
        return

    amount = quantity * STARS_RATE_RUB
    balance = await get_balance(query.from_user.id)
    text = (
        f"{ICON_TG_STARS} <b>TStars</b>\n\n"
        f"<b>Количество:</b> {quantity}\n"
        f"<b>Курс:</b> 1 {ICON_STAR_RATE} = {STARS_RATE_RUB:.2f} {settings.currency}\n"
        f"<b>К оплате:</b> {fmt_money(amount)}\n"
        f"{ICON_COIN} <b>Ваш баланс:</b> {fmt_money(balance)}\n\n"
        f"{ICON_NOTICE} Выдача по юзернейму. Время выдачи может занять до 30 минут."
    )
    await safe_edit(
        query.message,
        text,
        service_detail_kb(f"service_buy:stars:{quantity}", "catalog_stars"),
    )


@dp.callback_query(F.data.startswith("service_buy:"))
async def service_buy(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    parts = query.data.split(":")
    if len(parts) != 3:
        await query.answer("Услуга не найдена.", show_alert=True)
        return

    service_type, package_raw = parts[1], parts[2]
    if service_type == "premium" and not FEATURE_CATALOG_PREMIUM:
        await query.answer("Premium временно недоступен.", show_alert=True)
        return
    if service_type == "stars" and not FEATURE_CATALOG_STARS:
        await query.answer("Stars временно недоступны.", show_alert=True)
        return
    if service_type == "premium":
        try:
            months = int(package_raw)
        except ValueError:
            await query.answer("Период не найден.", show_alert=True)
            return
        amount = float(PREMIUM_PRICES_RUB.get(months, 0.0))
        if amount <= 0:
            await query.answer("Цена пока не настроена.", show_alert=True)
            return
        service_label = premium_service_label(months)
        quantity = months
        back_callback = f"premium_period:{months}"
        prompt = (
            f"{ICON_TG_PREMIUM} <b>{html.escape(service_label)}</b>\n\n"
            "Отправьте username получателя, на который нужно оформить Premium.\n"
            "Пример: <code>@username</code>"
        )
    elif service_type == "stars":
        try:
            quantity = int(package_raw)
        except ValueError:
            await query.answer("Пакет не найден.", show_alert=True)
            return
        if quantity not in STARS_PACKAGES:
            await query.answer("Пакет не найден.", show_alert=True)
            return
        amount = quantity * STARS_RATE_RUB
        service_label = f"Stars: {quantity}"
        back_callback = f"stars_package:{quantity}"
        prompt = (
            f"{ICON_TG_STARS} <b>Stars</b>\n\n"
            "Отправьте username получателя.\n"
            "Пример: <code>@username</code>"
        )
    else:
        await query.answer("Услуга не найдена.", show_alert=True)
        return

    balance = await get_balance(query.from_user.id)
    if balance + MONEY_EPSILON < amount:
        await query.answer("Недостаточно средств на балансе.", show_alert=True)
        await safe_edit(
            query.message,
            f"{ICON_COIN} <b>Недостаточно средств</b>\n\n"
            f"<b>К оплате:</b> {fmt_money(amount)}\n"
            f"<b>Ваш баланс:</b> {fmt_money(balance)}",
            back_to_main_kb(is_admin(query.from_user.id)),
        )
        return

    await state.update_data(
        service_type=service_type,
        service_label=service_label,
        service_quantity=quantity,
        service_amount=amount,
        service_back_callback=back_callback,
    )
    await state.set_state(ServiceOrderStates.waiting_recipient)
    await safe_edit(query.message, prompt, service_recipient_cancel_kb(back_callback))


@dp.callback_query(F.data == "service_order_cancel")
async def service_order_cancel(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer("Отменено.")
    await state.clear()
    text = f"{ICON_CATALOG_SECTIONS} <b>Каталог</b>\n\nВыберите раздел:"
    await safe_edit(query.message, text, catalog_sections_keyboard())


@dp.message(ServiceOrderStates.waiting_recipient)
async def service_order_recipient(message: Message, state: FSMContext):
    await ensure_known_user(message)
    data = await state.get_data()
    recipient = normalize_recipient_username(message.text or "")
    if not recipient:
        await message.answer("Отправьте корректный username. Пример: <code>@username</code>")
        return

    service_type = str(data.get("service_type") or "")
    service_label = str(data.get("service_label") or "")
    quantity = int(data.get("service_quantity") or 0)
    amount = float(data.get("service_amount") or 0)
    if service_type not in {"premium", "stars"} or not service_label or quantity <= 0 or amount <= 0:
        await state.clear()
        await message.answer("Заявка устарела. Откройте каталог и попробуйте еще раз.", reply_markup=back_to_main_kb(is_admin(message.from_user.id)))
        return

    ok, new_balance, order_id = await create_service_order_with_charge(
        user_id=message.from_user.id,
        username=message.from_user.username,
        first_name=message.from_user.first_name,
        service_type=service_type,
        service_label=service_label,
        recipient=recipient,
        quantity=quantity,
        amount=amount,
    )
    if not ok or order_id is None:
        await message.answer(
            f"{ICON_COIN} <b>Недостаточно средств</b>\n\n"
            f"<b>К оплате:</b> {fmt_money(amount)}\n"
            f"<b>Ваш баланс:</b> {fmt_money(new_balance)}",
            reply_markup=back_to_main_kb(is_admin(message.from_user.id)),
        )
        return

    try:
        await send_service_order_log(
            order_id=order_id,
            user=message.from_user,
            service_label=service_label,
            recipient=recipient,
            amount=amount,
            new_balance=new_balance,
        )
    except Exception as exc:
        logger.exception("Не удалось отправить заявку услуги в лог-канал: %s", exc)

    await state.clear()
    delivery_note = (
        f"{ICON_NOTICE} Выдача по юзернейму. Время выдачи может занять до 30 минут."
        if service_type == "stars"
        else f"{ICON_NOTICE} Выдача по юзернейму. Время выдачи может занять до двух часов."
    )
    text = (
        f"{ICON_SUCCESS} <b>Заявка принята</b>\n\n"
        f"<b>Услуга:</b> {html.escape(service_label)}\n"
        f"<b>Получатель:</b> <code>{html.escape(recipient)}</code>\n\n"
        f"{delivery_note}\n\n"
        f"<b>Текущий баланс:</b> {fmt_money(new_balance)}"
    )
    await message.answer(text, reply_markup=support_kb())


async def can_review_service_order(query: CallbackQuery) -> bool:
    return is_admin(query.from_user.id)


@dp.callback_query(F.data.startswith("service_done:"))
async def service_done_start(query: CallbackQuery):
    if not await can_review_service_order(query):
        await query.answer("Нет доступа.", show_alert=True)
        return
    order_id = int(query.data.split(":", 1)[1])
    order = await get_service_order(order_id)
    if not order:
        await query.answer("Заявка не найдена.", show_alert=True)
        return
    if order["status"] != "pending":
        await query.answer(f"Заявка уже обработана: {order['status']}", show_alert=True)
        return
    await query.answer()
    text = service_order_caption_from_row(order) + "\n\n<b>Подтвердить выдачу?</b>"
    await edit_review_message(query.message, text, service_order_confirm_kb("done", order_id, order["username"]))


@dp.callback_query(F.data.startswith("service_reject:"))
async def service_reject_start(query: CallbackQuery):
    if not await can_review_service_order(query):
        await query.answer("Нет доступа.", show_alert=True)
        return
    order_id = int(query.data.split(":", 1)[1])
    order = await get_service_order(order_id)
    if not order:
        await query.answer("Заявка не найдена.", show_alert=True)
        return
    if order["status"] != "pending":
        await query.answer(f"Заявка уже обработана: {order['status']}", show_alert=True)
        return
    await query.answer()
    text = service_order_caption_from_row(order) + "\n\n<b>Отказать по заявке и вернуть средства?</b>"
    await edit_review_message(query.message, text, service_order_confirm_kb("reject", order_id, order["username"]))


@dp.callback_query(F.data.startswith("service_review_cancel:"))
async def service_review_cancel(query: CallbackQuery):
    if not await can_review_service_order(query):
        await query.answer("Нет доступа.", show_alert=True)
        return
    order_id = int(query.data.split(":", 1)[1])
    order = await get_service_order(order_id)
    if not order:
        await query.answer("Заявка не найдена.", show_alert=True)
        return
    await query.answer("Действие отменено.")
    await edit_review_message(
        query.message,
        service_order_caption_from_row(order),
        service_order_review_kb(order_id, order["username"]),
    )


@dp.callback_query(F.data.startswith("service_confirm:"))
async def service_confirm_action(query: CallbackQuery):
    if not await can_review_service_order(query):
        await query.answer("Нет доступа.", show_alert=True)
        return
    _, action, order_id_raw = query.data.split(":", 2)
    order_id = int(order_id_raw)
    order = await get_service_order(order_id)
    if not order:
        await query.answer("Заявка не найдена.", show_alert=True)
        return

    base_text = service_order_caption_from_row(order)
    user_id = int(order["user_id"])
    service_label = order["service_label"]
    recipient = order["recipient"]
    if action == "done":
        ok, status, amount, _ = await approve_service_order(order_id, query.from_user.id)
        if not ok:
            await query.answer(f"Заявка уже обработана: {status}", show_alert=True)
            await edit_review_message(query.message, base_text + f"\n\nСтатус: <b>{html.escape(status)}</b>")
            return
        await query.answer("Заявка отмечена как выданная.", show_alert=True)
        await edit_review_message(query.message, base_text + f"\n\n<b>Выдано</b>\nАдмин: <code>{query.from_user.id}</code>")
        await log_purchase("admin_action", action=f"Выдана заявка #{order_id}: {service_label}", admin_id=query.from_user.id)
        if order["service_type"] == "stars":
            title = "Звёзды выданы"
            body = "Звёзды уже отправлены на указанный аккаунт."
        else:
            title = "Premium выдан"
            body = "Подписка оформлена на указанный аккаунт."
        try:
            await bot.send_message(
                user_id,
                f"{ICON_PARTY} <b>{title}</b>\n\n"
                f"<b>Услуга:</b> {html.escape(service_label)}\n"
                f"<b>Получатель:</b> <code>{html.escape(recipient)}</code>\n\n"
                f"{body}\n"
                f"Спасибо за покупку. {ICON_HEART}",
                reply_markup=purchase_success_kb(is_admin(user_id)),
            )
        except Exception:
            logger.exception("Could not notify user %s about delivered service order #%s", user_id, order_id)
        return

    if action == "reject":
        ok, status, amount, _, new_balance = await reject_service_order(order_id, query.from_user.id)
        if not ok:
            await query.answer(f"Заявка уже обработана: {status}", show_alert=True)
            await edit_review_message(query.message, base_text + f"\n\nСтатус: <b>{html.escape(status)}</b>")
            return
        await query.answer("Заявка отклонена, средства возвращены.", show_alert=True)
        await edit_review_message(query.message, base_text + f"\n\n<b>Отказано</b>\nАдмин: <code>{query.from_user.id}</code>\nВозврат: <b>{fmt_money(amount or 0)}</b>")
        await log_purchase("admin_action", action=f"Отклонена заявка #{order_id}: {service_label}, возврат {fmt_money(amount or 0)}", admin_id=query.from_user.id)
        try:
            await bot.send_message(
                user_id,
                f"{ICON_BLOCK} <b>Заявка отклонена</b>\n\n"
                f"<b>Услуга:</b> {html.escape(service_label)}\n"
                f"<b>Возврат:</b> {fmt_money(amount or 0)}\n"
                f"<b>Текущий баланс:</b> {fmt_money(new_balance or 0)}\n\n"
                "Если вы считаете, что это ошибка, напишите в поддержку.",
                reply_markup=support_kb(),
            )
        except Exception:
            logger.exception("Could not notify user %s about rejected service order #%s", user_id, order_id)
        return

    await query.answer("Неизвестное действие.", show_alert=True)


@dp.callback_query(F.data == "menu_balance")
async def menu_balance(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    text = (
        f"{ICON_COIN} <b>Баланс:</b> {fmt_money(await get_balance(query.from_user.id))}"
    )
    await safe_edit(query.message, text, back_to_main_kb(is_admin(query.from_user.id)))


async def call_crypto_pay(method: str, params: dict = None) -> dict:
    """Вспомогательная функция для запросов к Crypto Pay API."""
    if not settings.cryptopay_token:
        logger.error("CRYPTOPAY_TOKEN не настроен!")
        return {"ok": False, "error": "Token missing"}
    
    url = f"https://pay.crypt.bot/api/{method}"
    headers = {"Crypto-Pay-API-Token": settings.cryptopay_token}
    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(url, json=params, headers=headers) as resp:
                try:
                    payload = await resp.json()
                except Exception as exc:
                    body = await resp.text()
                    logger.exception("CryptoPay returned non-json response for %s: status=%s body=%s", method, resp.status, body[:500])
                    return {"ok": False, "error": f"Invalid JSON response: {exc}"}
                if resp.status >= 400:
                    logger.error("CryptoPay HTTP error for %s: status=%s payload=%s", method, resp.status, payload)
                return payload
    except Exception as exc:
        logger.exception("CryptoPay request failed for %s", method)
        return {"ok": False, "error": str(exc)}


def crypto_invoice_amount_matches(invoice: dict, expected_amount: float) -> bool:
    api_amount_raw = invoice.get("amount")
    if api_amount_raw is None:
        return True
    api_amount = parse_float(str(api_amount_raw))
    return api_amount is not None and abs(api_amount - expected_amount) <= MONEY_EPSILON


def crypto_invoice_fiat_matches(invoice: dict, expected_fiat: str) -> bool:
    api_fiat = (invoice.get("fiat") or "").strip().upper()
    if not api_fiat:
        return True
    return api_fiat == (expected_fiat or "").strip().upper()


@dp.callback_query(F.data == "user_topup_start")
async def user_topup_start(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    await state.clear()
    text = (
        f"<b>{ICON_WALLET} Пополнение баланса</b>\n\n"
        f"{ICON_COIN} <b>Текущий баланс:</b> {fmt_money(await get_balance(query.from_user.id))}\n\n"
        "Выберите метод пополнения:"
    )
    await safe_edit(query.message, text, topup_methods_keyboard())


@dp.callback_query(F.data == "user_topup_methods")
async def user_topup_methods(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    await state.clear()
    text = (
        f"<b>{ICON_WALLET} Пополнение баланса</b>\n\n"
        f"{ICON_COIN} <b>Текущий баланс:</b> {fmt_money(await get_balance(query.from_user.id))}\n\n"
        "Выберите метод пополнения:"
    )
    await safe_edit(query.message, text, topup_methods_keyboard())


@dp.callback_query(F.data == "topup_other")
async def topup_other(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    if not FEATURE_TOPUP_OTHER:
        await query.answer("Этот способ пополнения временно недоступен.", show_alert=True)
        return
    await query.answer()
    await state.clear()
    text = (
        "<b>Другие способы пополнения</b>\n\n"
        "Доступны карты Беларуси, Казахстана и Узбекистана.\n"
        "Также возможно пополнение криптовалютой по адресу.\n\n"
        "Напишите в поддержку — оператор подскажет актуальные реквизиты и поможет с оплатой."
    )
    await safe_edit(query.message, text, topup_other_kb())


@dp.callback_query(F.data == "cancel_topup_receipt")
async def cancel_topup_receipt(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer("Пополнение отменено.")
    await state.clear()
    await show_home(query)


@dp.callback_query(F.data.startswith("topup_method:"))
async def user_topup_method(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    method = query.data.split(":", 1)[1]
    if not is_topup_method_enabled(method):
        await query.answer("Этот способ пополнения временно недоступен.", show_alert=True)
        await safe_edit(
            query.message,
            f"<b>{ICON_WALLET} Пополнение баланса</b>\n\n"
            f"{ICON_COIN} <b>Текущий баланс:</b> {fmt_money(await get_balance(query.from_user.id))}\n\n"
            "Выберите метод пополнения:",
            topup_methods_keyboard(),
        )
        return
    await state.update_data(topup_method=method)
    await state.set_state(UserTopUpStates.waiting_amount)
    currency = settings.currency
    text = (
        f"{topup_method_icon(method)} <b>{html.escape(topup_method_title(method))}</b>\n\n"
        f"Введите любую сумму пополнения в {html.escape(currency)}."
    )
    await safe_edit(query.message, text, cancel_flow_kb("user_topup_methods"))


@dp.message(UserTopUpStates.waiting_amount)
async def user_topup_amount(message: Message, state: FSMContext):
    await ensure_known_user(message)
    data = await state.get_data()
    method = data.get("topup_method")
    if method not in {"ru", "ua", "crypto"} or not is_topup_method_enabled(method):
        await state.clear()
        await message.answer("Метод пополнения не выбран. Откройте пополнение заново.", reply_markup=topup_methods_keyboard())
        return

    credit_amount = parse_float(message.text)
    if credit_amount is None or credit_amount < MIN_CRYPTO_TOPUP_AMOUNT:
        await message.answer(
            f"Введите корректную сумму. Минимум: {MIN_CRYPTO_TOPUP_AMOUNT:g} {html.escape(settings.cryptopay_fiat)}."
        )
        return

    if method in {"ru", "ua"}:
        quote = await build_topup_quote(method, credit_amount)
        await state.update_data(
            topup_payment_amount=quote["payment_amount"],
            topup_credit_amount=quote["credit_amount"],
            topup_currency=quote["payment_currency"],
            topup_rate=quote["rate"],
            topup_rate_source=quote["rate_source"],
        )
        await state.set_state(UserTopUpStates.waiting_receipt)
        await message.answer(await manual_topup_requisites(method, quote), reply_markup=topup_receipt_kb())
        return

    # Создаем инвойс через Crypto Pay
    res = await call_crypto_pay("createInvoice", {
        "amount": str(credit_amount),
        "fiat": settings.cryptopay_fiat,
        "currency_type": "fiat",
        "description": f"Пополнение баланса Stryx Shop (User ID: {message.from_user.id})"
    })

    if not res.get("ok"):
        logger.error(f"CryptoPay error: {res}")
        await message.answer("Не удалось создать счет. Попробуйте позже.")
        return

    invoice = res["result"]
    pay_url = invoice["pay_url"]
    invoice_id = str(invoice["invoice_id"])

    try:
        await record_crypto_invoice(
            invoice_id=invoice_id,
            user_id=message.from_user.id,
            amount=credit_amount,
            fiat=settings.cryptopay_fiat,
            pay_url=pay_url,
        )
    except Exception:
        logger.exception("Could not record CryptoPay invoice %s for user %s", invoice_id, message.from_user.id)
        await message.answer("Счет создан, но не удалось сохранить его в базе. Создайте счет заново через минуту.")
        return

    text = (
        f"{ICON_CRYPTO} <b>Счет Crypto Bot создан</b>\n\n"
        f"<b>Сумма:</b> {credit_amount:.2f} {html.escape(settings.cryptopay_fiat)}\n"
        "<b>Статус:</b> ожидает оплаты\n\n"
        "Откройте счет в Crypto Bot и оплатите удобной криптовалютой. "
        f"После оплаты нажмите проверку, {ICON_COIN} баланс обновится автоматически."
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Оплатить", url=pay_url, icon_custom_emoji_id=BTN_ICON_PAY)],
        [InlineKeyboardButton(text="Проверить оплату", callback_data=f"check_pay:{invoice_id}", icon_custom_emoji_id=BTN_ICON_CHECK)],
        [InlineKeyboardButton(text="Отменить", callback_data="menu_home", icon_custom_emoji_id=BTN_ICON_CANCEL)]
    ])
    await message.answer(text, reply_markup=kb)
    await state.clear()


@dp.message(UserTopUpStates.waiting_receipt)
async def user_topup_receipt(message: Message, state: FSMContext):
    await ensure_known_user(message)
    data = await state.get_data()
    method = data.get("topup_method")
    payment_amount = float(data.get("topup_payment_amount") or 0)
    credit_amount = float(data.get("topup_credit_amount") or 0)
    currency = data.get("topup_currency") or topup_currency(method)
    if method not in {"ru", "ua"} or payment_amount <= 0 or credit_amount <= 0:
        await state.clear()
        await message.answer("Заявка устарела. Создайте пополнение заново.", reply_markup=topup_methods_keyboard())
        return

    receipt_type = ""
    receipt_file_id = ""
    if message.photo:
        receipt_type = "photo"
        receipt_file_id = message.photo[-1].file_id
    elif message.document:
        document = message.document
        mime_type = (document.mime_type or "").lower()
        file_name = (document.file_name or "").lower()
        if mime_type not in {"application/pdf", "image/jpeg", "image/png", "image/webp"} and not file_name.endswith((".pdf", ".jpg", ".jpeg", ".png", ".webp")):
            await message.answer("Отправьте фото чека или PDF-файл.")
            return
        receipt_type = "document"
        receipt_file_id = document.file_id
    else:
        await message.answer("Отправьте фото чека или PDF-файл одним сообщением.")
        return

    request_id = await create_topup_request(
        user_id=message.from_user.id,
        username=message.from_user.username,
        first_name=message.from_user.first_name,
        method=method,
        amount=payment_amount,
        currency=currency,
        credit_amount=credit_amount,
        receipt_type=receipt_type,
        receipt_file_id=receipt_file_id,
    )
    caption = topup_request_caption(
        request_id,
        user=message.from_user,
        method=method,
        payment_amount=payment_amount,
        currency=currency,
        credit_amount=credit_amount,
    )
    try:
        if receipt_type == "photo":
            await bot.send_photo(topup_chat_id(method), receipt_file_id, caption=caption, reply_markup=topup_review_kb(request_id, message.from_user.id, message.from_user.username))
        else:
            await bot.send_document(topup_chat_id(method), receipt_file_id, caption=caption, reply_markup=topup_review_kb(request_id, message.from_user.id, message.from_user.username))
    except Exception as exc:
        logger.exception("Could not send topup request #%s", request_id)
        await message.answer(
            f"{ICON_BLOCK} <b>Не удалось передать чек оператору</b>\n\n"
            "Чек получен, но заявка не была отправлена на проверку. "
            "Напишите в поддержку, чтобы мы быстро разобрались.",
            reply_markup=support_kb(),
        )
        await state.clear()
        return

    await state.clear()
    await message.answer(
        "<b>Заявка принята</b>\n\n"
        "Чек отправлен на проверку. После подтверждения баланс обновится.",
        reply_markup=support_kb(),
    )

@dp.callback_query(F.data.startswith("check_pay:"))
async def check_pay_callback(query: CallbackQuery):
    if not FEATURE_TOPUP_CRYPTO:
        await query.answer("Crypto Bot временно недоступен.", show_alert=True)
        return
    parts = (query.data or "").split(":", 2)
    if len(parts) < 2 or not parts[1].strip():
        await query.answer("Некорректный счет.", show_alert=True)
        return

    invoice_id = parts[1].strip()
    stored_invoice = await get_crypto_invoice(invoice_id)
    if not stored_invoice:
        logger.warning("CryptoPay invoice check without local invoice: invoice=%s user=%s", invoice_id, query.from_user.id)
        await query.answer("Счет не найден. Создайте новое пополнение.", show_alert=True)
        return

    invoice_user_id = int(stored_invoice["user_id"])
    if invoice_user_id != query.from_user.id:
        logger.warning(
            "CryptoPay invoice ownership mismatch: invoice=%s owner=%s checker=%s",
            invoice_id,
            invoice_user_id,
            query.from_user.id,
        )
        await query.answer("Это не ваш счет.", show_alert=True)
        return

    amount = float(stored_invoice["amount"])
    fiat = (stored_invoice["fiat"] or settings.cryptopay_fiat).upper()
    
    res = await call_crypto_pay("getInvoices", {"invoice_ids": invoice_id})
    items = ((res.get("result") or {}).get("items") or []) if isinstance(res, dict) else []
    if not res.get("ok") or not items:
        logger.warning("CryptoPay invoice check failed: invoice=%s response=%s", invoice_id, res)
        await query.answer("Счет не найден.", show_alert=True)
        return

    invoice = items[0]
    if str(invoice.get("invoice_id") or invoice_id) != invoice_id:
        logger.error("CryptoPay returned different invoice: requested=%s payload=%s", invoice_id, invoice)
        await query.answer("Ошибка проверки счета. Напишите в поддержку.", show_alert=True)
        return

    status = invoice.get("status")

    if status == "paid":
        if not crypto_invoice_amount_matches(invoice, amount) or not crypto_invoice_fiat_matches(invoice, fiat):
            logger.error(
                "CryptoPay invoice amount mismatch: invoice=%s expected=%s %s payload=%s",
                invoice_id,
                amount,
                fiat,
                invoice,
            )
            await query.answer("Сумма счета не совпала. Напишите в поддержку.", show_alert=True)
            return

        await mark_crypto_invoice_status(invoice_id, "paid", str(invoice.get("paid_at") or ""))
        processed, _new_balance = await process_crypto_topup(invoice_id, invoice_user_id, amount)
        if not processed:
            await query.answer("Этот счет уже был зачислен.", show_alert=True)
            await show_home(query)
            return
        await log_purchase("crypto_topup", user_id=invoice_user_id, invoice_id=invoice_id, amount=amount)
        await query.answer("Оплата получена. Баланс пополнен.", show_alert=True)
        await show_home(query)
    elif status == "expired":
        await mark_crypto_invoice_status(invoice_id, "expired")
        await query.answer("Время оплаты счета истекло.", show_alert=True)
        await safe_edit(query.message, "<b>Счет просрочен.</b> Попробуйте создать новый.", back_to_main_kb(is_admin(query.from_user.id)))
    else:
        await mark_crypto_invoice_status(invoice_id, str(status or "created"))
        await query.answer("Оплата еще не поступила. Попробуйте через минуту.", show_alert=True)


@dp.callback_query(F.data.startswith("topup_approve:"))
async def topup_approve_start(query: CallbackQuery):
    if not await can_review_topup(query):
        await query.answer("Нет доступа.", show_alert=True)
        return
    request_id = int(query.data.split(":", 1)[1])
    request = await get_topup_request(request_id)
    if not request:
        await query.answer("Заявка не найдена.", show_alert=True)
        return
    if request["status"] != "pending":
        await query.answer(f"Заявка уже обработана: {request['status']}", show_alert=True)
        return
    await query.answer()
    text = topup_request_caption_from_row(request) + "\n\n<b>Вы уверены, что хотите одобрить эту заявку?</b>"
    await edit_review_message(query.message, text, topup_confirm_kb("approve", request_id, int(request["user_id"]), request["username"]))


@dp.callback_query(F.data.startswith("topup_reject:"))
async def topup_reject_start(query: CallbackQuery):
    if not await can_review_topup(query):
        await query.answer("Нет доступа.", show_alert=True)
        return
    request_id = int(query.data.split(":", 1)[1])
    request = await get_topup_request(request_id)
    if not request:
        await query.answer("Заявка не найдена.", show_alert=True)
        return
    if request["status"] != "pending":
        await query.answer(f"Заявка уже обработана: {request['status']}", show_alert=True)
        return
    await query.answer()
    text = topup_request_caption_from_row(request) + "\n\n<b>Вы уверены, что хотите отказать по этой заявке?</b>"
    await edit_review_message(query.message, text, topup_confirm_kb("reject", request_id, int(request["user_id"]), request["username"]))


@dp.callback_query(F.data.startswith("topup_cancel:"))
async def topup_cancel_review(query: CallbackQuery):
    if not await can_review_topup(query):
        await query.answer("Нет доступа.", show_alert=True)
        return
    request_id = int(query.data.split(":", 1)[1])
    request = await get_topup_request(request_id)
    if not request:
        await query.answer("Заявка не найдена.", show_alert=True)
        return
    await query.answer("Действие отменено.")
    await edit_review_message(
        query.message,
        topup_request_caption_from_row(request),
        topup_review_kb(request_id, int(request["user_id"]), request["username"]),
    )


@dp.callback_query(F.data.startswith("topup_confirm:"))
async def topup_confirm_action(query: CallbackQuery):
    if not await can_review_topup(query):
        await query.answer("Нет доступа.", show_alert=True)
        return
    _, action, request_id_raw = query.data.split(":", 2)
    request_id = int(request_id_raw)
    request = await get_topup_request(request_id)
    if not request:
        await query.answer("Заявка не найдена.", show_alert=True)
        return

    base_text = topup_request_caption_from_row(request)
    if action == "approve":
        ok, status, amount, user_id, currency = await approve_topup_request(request_id, query.from_user.id)
        if not ok:
            await query.answer(f"Заявка уже обработана: {status}", show_alert=True)
            await edit_review_message(query.message, base_text + f"\n\nСтатус: <b>{html.escape(status)}</b>")
            return
        await query.answer("Баланс зачислен.", show_alert=True)
        await edit_review_message(query.message, base_text + f"\n\n<b>Одобрено</b>\nАдмин: <code>{query.from_user.id}</code>")
        await log_purchase("admin_action", action=f"Одобрена заявка пополнения #{request_id} на {fmt_money(amount)}", admin_id=query.from_user.id)
        try:
            await bot.send_message(
                user_id,
                f"{ICON_SUCCESS} <b>Пополнение одобрено</b>\n\n"
                f"{ICON_COIN} Баланс пополнен на <b>{fmt_money(amount)}</b>.",
                reply_markup=back_to_main_kb(is_admin(user_id)),
            )
        except Exception:
            logger.exception("Could not notify user %s about approved topup #%s", user_id, request_id)
        return

    if action == "reject":
        ok, status, amount, user_id, currency = await reject_topup_request(request_id, query.from_user.id)
        if not ok:
            await query.answer(f"Заявка уже обработана: {status}", show_alert=True)
            await edit_review_message(query.message, base_text + f"\n\nСтатус: <b>{html.escape(status)}</b>")
            return
        await query.answer("Заявка отклонена.", show_alert=True)
        await edit_review_message(query.message, base_text + f"\n\n<b>Отклонено</b>\nАдмин: <code>{query.from_user.id}</code>")
        await log_purchase("admin_action", action=f"Отклонена заявка пополнения #{request_id}", admin_id=query.from_user.id)
        try:
            await bot.send_message(
                user_id,
                f"{ICON_BLOCK} <b>Пополнение отклонено</b>\n\n"
                "Заявка не прошла проверку. "
                "Если вы считаете, что это ошибка, напишите в поддержку.",
                reply_markup=support_kb(),
            )
        except Exception:
            logger.exception("Could not notify user %s about rejected topup #%s", user_id, request_id)
        return

    await query.answer("Неизвестное действие.", show_alert=True)


@dp.callback_query(F.data == "drops_menu")
async def drops_menu(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id):
        await query.answer("Нет доступа.", show_alert=True)
        return
    await query.answer()
    await state.clear()
    await safe_edit(
        query.message,
        "<b>Доступ к заявкам</b>\n\n"
        "Эти пользователи могут одобрять и отклонять заявки на пополнение в чатах заявок.",
        drops_menu_kb(await build_drops_rows()),
    )


@dp.callback_query(F.data == "drops_add")
async def drops_add_start(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id):
        await query.answer("Нет доступа.", show_alert=True)
        return
    await query.answer()
    await state.set_state(AdminDropsStates.waiting_user_id)
    await safe_edit(query.message, "<b>Добавить доступ</b>\n\nОтправьте ID пользователя.", cancel_flow_kb("drops_menu"))


@dp.message(AdminDropsStates.waiting_user_id)
async def drops_add_finish(message: Message, state: FSMContext):
    await ensure_known_user(message)
    if not is_admin(message.from_user.id):
        return
    text = (message.text or "").strip()
    if not text.isdigit():
        await message.answer("Отправьте числовой ID.")
        return
    user_id = int(text)
    added = await add_topup_reviewer(user_id, message.from_user.id)
    await state.clear()
    if added:
        await message.answer(f"Доступ добавлен: <code>{user_id}</code>", reply_markup=drops_menu_kb(await build_drops_rows()))
    else:
        await message.answer(f"У пользователя уже есть доступ: <code>{user_id}</code>", reply_markup=drops_menu_kb(await build_drops_rows()))


@dp.callback_query(F.data.startswith("drops_user:"))
async def drops_user_view(query: CallbackQuery):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id):
        await query.answer("Нет доступа.", show_alert=True)
        return
    user_id = int(query.data.split(":", 1)[1])
    await query.answer()
    user = await get_user(user_id)
    label = user["username"] if user and user["username"] else None
    text = (
        "<b>Доступ к заявкам</b>\n\n"
        f"User ID: <code>{user_id}</code>\n"
        f"Username: @{html.escape(label) if label else '—'}"
    )
    await safe_edit(query.message, text, drop_manage_kb(user_id))


@dp.callback_query(F.data.startswith("drops_remove:"))
async def drops_remove(query: CallbackQuery):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id):
        await query.answer("Нет доступа.", show_alert=True)
        return
    user_id = int(query.data.split(":", 1)[1])
    removed = await remove_topup_reviewer(user_id)
    await query.answer("Доступ удален." if removed else "Доступ не найден.", show_alert=True)
    await safe_edit(
        query.message,
        "<b>Доступ к заявкам</b>\n\n"
        "Эти пользователи могут одобрять и отклонять заявки на пополнение в чатах заявок.",
        drops_menu_kb(await build_drops_rows()),
    )


@dp.callback_query(F.data == "menu_help")
async def menu_help(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    text = f"{ICON_HELP} <b>Помощь</b>\n\nСвязь с поддержкой:"
    await safe_edit(query.message, text, help_menu_kb(is_admin(query.from_user.id)))


async def render_cart(query: CallbackQuery) -> None:
    items = await list_cart_items(query.from_user.id)
    if not items:
        await safe_edit(
            query.message,
            f"{ICON_FOLDER} <b>Корзина</b>\n\nПока пусто.",
            cart_kb(can_checkout=False),
        )
        return

    total = sum(float(item["price"]) for item in items if item["status"] == "available")
    unavailable_count = sum(1 for item in items if item["status"] != "available")
    lines = [
        f"{ICON_FOLDER} <b>Корзина</b>",
        "",
        f"Товаров: <b>{len(items)}</b>",
        f"К оплате: <b>{fmt_money(total)}</b>",
        f"{ICON_COIN} Баланс: <b>{fmt_money(await get_balance(query.from_user.id))}</b>",
    ]
    if unavailable_count:
        lines.append(f"\n{ICON_NOTICE} Недоступных товаров: <b>{unavailable_count}</b>. Их нужно удалить из корзины.")
    lines.append("")
    grouped_items: dict[tuple[str, str, float, str], dict[str, object]] = {}
    for item in items:
        status = str(item["status"])
        key = (
            str(item["title"] or "Товар"),
            str(item["country"] or "—"),
            float(item["price"]),
            status,
        )
        bucket = grouped_items.setdefault(
            key,
            {
                "title": key[0],
                "country": key[1],
                "price": key[2],
                "status": status,
                "count": 0,
                "total": 0.0,
            },
        )
        bucket["count"] = int(bucket["count"]) + 1
        if status == "available":
            bucket["total"] = float(bucket["total"]) + float(item["price"])

    for idx, group in enumerate(grouped_items.values(), 1):
        status_note = "" if group["status"] == "available" else " недоступен"
        lines.append(
            f"{idx}. <b>{render_rich_text(group['title'])}</b> • "
            f"{html.escape(str(group['country']))} • "
            f"{int(group['count'])} шт • "
            f"{fmt_money(float(group['total']))}{status_note}"
        )

    await safe_edit(
        query.message,
        "\n".join(lines),
        cart_kb(can_checkout=True),
    )


async def refresh_cart_unavailable_items(user_id: int) -> tuple[list[int], list[int]]:
    """Replace stale cart rows with currently available accounts from the same department."""
    replaced: list[int] = []
    removed: list[int] = []
    for item in await list_cart_items(user_id):
        if item["status"] == "available":
            continue
        old_product_id = int(item["product_id"])
        await remove_product_from_cart(user_id, old_product_id)
        ok, _reason, new_product_id = await add_product_group_to_cart(user_id, old_product_id)
        if ok and new_product_id is not None:
            replaced.append(int(new_product_id))
        else:
            removed.append(old_product_id)
    return replaced, removed


@dp.callback_query(F.data == "menu_cart")
async def menu_cart(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    await render_cart(query)


@dp.callback_query(F.data.startswith("cart_add:"))
async def cart_add(query: CallbackQuery):
    await ensure_known_user(query)
    try:
        product_id = int(query.data.split(":", 1)[1])
    except ValueError:
        await query.answer("Некорректный товар.", show_alert=True)
        return

    ok, reason = await add_product_to_cart(query.from_user.id, product_id)
    if ok:
        await query.answer("Добавлено в корзину.")
        return
    if reason == "already_in_cart":
        await query.answer("Этот товар уже в корзине.", show_alert=True)
        return
    if reason == "not_available":
        await query.answer("Товар уже недоступен.", show_alert=True)
        return
    await query.answer("Товар не найден.", show_alert=True)


@dp.callback_query(F.data.startswith("cart_add_group:"))
async def cart_add_group(query: CallbackQuery):
    await ensure_known_user(query)
    try:
        sample_product_id = int(query.data.split(":", 1)[1])
    except ValueError:
        await query.answer("Некорректный товар.", show_alert=True)
        return

    ok, reason, product_id = await add_product_group_to_cart(query.from_user.id, sample_product_id)
    if ok:
        await query.answer(f"Добавлено в корзину: #{product_id}.")
        return
    if reason == "already_in_cart":
        await query.answer("Все доступные аккаунты этого типа уже в корзине.", show_alert=True)
        return
    if reason == "not_available":
        await query.answer("Этот тип товара закончился.", show_alert=True)
        return
    await query.answer("Товар не найден.", show_alert=True)


async def add_product_group_quantity_to_cart(user_id: int, sample_product_id: int, quantity: int) -> tuple[int, str]:
    added = 0
    last_reason = "ok"
    for _ in range(max(0, quantity)):
        ok, reason, _product_id = await add_product_group_to_cart(user_id, sample_product_id)
        if not ok:
            last_reason = reason
            break
        added += 1
    return added, last_reason


@dp.callback_query(F.data.startswith("cart_add_group_qty:"))
async def cart_add_group_quantity(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    try:
        _, sample_id_raw, qty_raw = query.data.split(":", 2)
        sample_product_id = int(sample_id_raw)
        quantity = int(qty_raw)
    except ValueError:
        await query.answer("Некорректное количество.", show_alert=True)
        return
    if quantity <= 0:
        await query.answer("Количество должно быть больше 0.", show_alert=True)
        return
    stock_count, in_cart_count, available_to_add = await get_department_cart_capacity(query.from_user.id, sample_product_id)
    if stock_count <= 0:
        await query.answer("Все доступные аккаунты этого типа уже в корзине или закончились.", show_alert=True)
        return
    if available_to_add <= 0:
        await query.answer("Вы уже добавили всё текущее наличие этого типа в корзину.", show_alert=True)
        return
    if quantity > available_to_add:
        await query.answer(f"Нельзя добавить больше. Уже в корзине: {in_cart_count}, можно добавить ещё: {available_to_add}.", show_alert=True)
        return
    added, reason = await add_product_group_quantity_to_cart(query.from_user.id, sample_product_id, quantity)
    if added == quantity:
        await query.answer("Добавлено в корзину.")
        group = await get_product_department(sample_product_id)
        title = render_rich_text(group["title"] if group else "товар")
        await safe_edit(
            query.message,
            f"<b>Добавлено в корзину</b>\n\n"
            f"<b>{title}</b>\n"
            f"Количество: <b>{added}</b>\n\n"
            "Оплатить можно через корзину.",
            open_cart_kb(),
        )
        return
    if added > 0:
        await query.answer(f"Добавлено только {added}: больше свободных нет.", show_alert=True)
        group = await get_product_department(sample_product_id)
        title = render_rich_text(group["title"] if group else "товар")
        await safe_edit(
            query.message,
            f"<b>Добавлено в корзину</b>\n\n"
            f"<b>{title}</b>\n"
            f"Количество: <b>{added}</b>\n\n"
            "Оплатить можно через корзину.",
            open_cart_kb(),
        )
        return
    if reason == "already_in_cart":
        await query.answer("Все доступные аккаунты этого типа уже в корзине.", show_alert=True)
        return
    if reason == "not_available":
        await query.answer("Этот тип товара закончился.", show_alert=True)
        return
    await query.answer("Товар не найден.", show_alert=True)


@dp.callback_query(F.data.startswith("cart_add_group_manual:"))
async def cart_add_group_manual_start(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    try:
        sample_product_id = int(query.data.split(":", 1)[1])
    except ValueError:
        await query.answer("Некорректный товар.", show_alert=True)
        return
    group = await get_product_group(sample_product_id)
    if not group:
        await query.answer("Этот тип товара закончился.", show_alert=True)
        return
    stock_count, in_cart_count, available_to_add = await get_department_cart_capacity(query.from_user.id, sample_product_id)
    if stock_count <= 0:
        await query.answer("Все доступные аккаунты этого типа уже в корзине или закончились.", show_alert=True)
        return
    if available_to_add <= 0:
        await query.answer("Вы уже добавили всё текущее наличие этого типа в корзину.", show_alert=True)
        return
    await query.answer()
    await state.update_data(cart_group_sample_id=sample_product_id, cart_group_max_qty=available_to_add)
    await state.set_state(UserCartStates.waiting_group_quantity)
    await safe_edit(
        query.message,
        f"<b>Сколько добавить?</b>\n\n"
        f"<b>{render_rich_text(group['title'])}</b>\n"
        f"<b>В наличии:</b> {stock_count}\n"
        f"<b>Уже в корзине:</b> {in_cart_count}\n"
        f"<b>Можно добавить ещё:</b> {available_to_add}\n\n"
        "Отправьте количество числом.",
        cancel_flow_kb("menu_catalog"),
    )


@dp.message(UserCartStates.waiting_group_quantity)
async def cart_add_group_manual_finish(message: Message, state: FSMContext):
    text = (message.text or "").strip()
    if not text.isdigit():
        await message.answer("Отправьте количество числом.")
        return
    quantity = int(text)
    if quantity <= 0:
        await message.answer("Количество должно быть больше 0.")
        return
    data = await state.get_data()
    sample_product_id = int(data.get("cart_group_sample_id") or 0)
    max_qty = int(data.get("cart_group_max_qty") or 0)
    stock_count, in_cart_count, available_to_add = await get_department_cart_capacity(message.from_user.id, sample_product_id)
    if max_qty > 0:
        available_to_add = min(available_to_add, max_qty)
    if stock_count <= 0:
        await state.clear()
        await message.answer("Все доступные аккаунты этого типа уже в корзине или закончились.", reply_markup=back_to_main_kb(is_admin(message.from_user.id)))
        return
    if available_to_add <= 0:
        await state.clear()
        await message.answer("Вы уже добавили всё текущее наличие этого типа в корзину.", reply_markup=open_cart_kb())
        return
    if quantity > available_to_add:
        await message.answer(f"Нельзя добавить больше. Уже в корзине: <b>{in_cart_count}</b>, можно добавить ещё: <b>{available_to_add}</b>.")
        return
    await state.clear()
    added, reason = await add_product_group_quantity_to_cart(message.from_user.id, sample_product_id, quantity)
    if added == quantity:
        group = await get_product_department(sample_product_id)
        title = render_rich_text(group["title"] if group else "товар")
        await message.answer(
            f"<b>Добавлено в корзину</b>\n\n"
            f"<b>{title}</b>\n"
            f"Количество: <b>{added}</b>\n\n"
            "Оплатить можно через корзину.",
            reply_markup=open_cart_kb(),
        )
        return
    if added > 0:
        group = await get_product_department(sample_product_id)
        title = render_rich_text(group["title"] if group else "товар")
        await message.answer(
            f"<b>Добавлено в корзину</b>\n\n"
            f"<b>{title}</b>\n"
            f"Количество: <b>{added}</b>\n\n"
            "Оплатить можно через корзину.",
            reply_markup=open_cart_kb(),
        )
        return
    if reason == "already_in_cart":
        await message.answer("Все доступные аккаунты этого типа уже в корзине.", reply_markup=back_to_main_kb(is_admin(message.from_user.id)))
        return
    if reason == "not_available":
        await message.answer("Этот тип товара закончился.", reply_markup=back_to_main_kb(is_admin(message.from_user.id)))
        return
    await message.answer("Товар не найден.", reply_markup=back_to_main_kb(is_admin(message.from_user.id)))


@dp.callback_query(F.data.startswith("cart_remove:"))
async def cart_remove(query: CallbackQuery):
    await ensure_known_user(query)
    try:
        product_id = int(query.data.split(":", 1)[1])
    except ValueError:
        await query.answer("Некорректный товар.", show_alert=True)
        return
    removed = await remove_product_from_cart(query.from_user.id, product_id)
    await query.answer("Удалено." if removed else "Товара нет в корзине.")
    await render_cart(query)


@dp.callback_query(F.data == "cart_clear")
async def cart_clear_callback(query: CallbackQuery):
    await ensure_known_user(query)
    await clear_cart(query.from_user.id)
    await query.answer("Корзина очищена.")
    await render_cart(query)


@dp.callback_query(F.data == "cart_checkout")
async def cart_checkout(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer("Оформляю покупку...")
    replaced_unavailable, removed_unavailable = await refresh_cart_unavailable_items(query.from_user.id)
    cart_items = await list_cart_items(query.from_user.id)
    payable_items = [item for item in cart_items if item["status"] == "available"]
    payable_total = round(sum(float(item["price"]) for item in payable_items), 8)
    balance = await get_balance(query.from_user.id)
    if not payable_items:
        await render_cart(query)
        return
    if balance + MONEY_EPSILON < payable_total:
        await safe_edit(
            query.message,
            f"{ICON_BLOCK} <b>Недостаточно средств</b>\n\n"
            f"К оплате: <b>{fmt_money(payable_total)}</b>\n"
            f"Ваш баланс: <b>{fmt_money(balance)}</b>\n\n"
            "Аккаунты не проверялись и не готовились к выдаче.",
            back_to_main_kb(is_admin(query.from_user.id)),
        )
        return
    await safe_edit(
        query.message,
        f"{ICON_KEYBOARD} <b>Формируем заказ</b>\n\n"
        f"Проверяю аккаунты перед выдачей: <b>{len(payable_items)}</b> шт. Обычно это занимает несколько секунд.",
    )
    removed_dead = []
    removed_dead_total = 0.0
    for item in cart_items:
        if item["status"] != "available":
            continue
        if not await verify_product_alive_for_sale(item, context=f"Оплата корзины пользователем {query.from_user.id}"):
            removed_dead.append(int(item["product_id"]))
            removed_dead_total += float(item["price"])
            await remove_product_from_cart(query.from_user.id, int(item["product_id"]))
    unavailable_after_check: list[int] = []
    while True:
        result = await purchase_cart(query.from_user.id)
        if result.ok or result.reason != "not_available" or not result.unavailable_ids:
            break
        for product_id in result.unavailable_ids:
            await remove_product_from_cart(query.from_user.id, int(product_id))
            ok, _reason, new_product_id = await add_product_group_to_cart(query.from_user.id, int(product_id))
            if ok and new_product_id is not None:
                replaced_unavailable.append(int(new_product_id))
            else:
                unavailable_after_check.append(int(product_id))
    if not result.ok:
        if result.reason == "cart_empty":
            if removed_dead or unavailable_after_check:
                await safe_edit(
                    query.message,
                    f"{ICON_BLOCK} <b>Живых аккаунтов не осталось</b>\n\n"
                    f"Мёртвых позиций: <b>{len(removed_dead)}</b>\n"
                    f"Недоступных позиций: <b>{len(unavailable_after_check)}</b>\n"
                    f"К возврату/не списано: <b>{fmt_money(removed_dead_total)}</b>\n\n"
                    "Деньги за них не списались. По мёртвым аккаунтам админы уже получили уведомление.",
                    back_to_main_kb(is_admin(query.from_user.id)),
                )
            else:
                await render_cart(query)
            return
        if result.reason == "not_available":
            ids = ", ".join(str(pid) for pid in (result.unavailable_ids or []))
            await safe_edit(
                query.message,
                f"{ICON_BLOCK} <b>Не удалось оплатить корзину</b>\n\nНедоступные товары: <code>{html.escape(ids)}</code>\nУдалите их из корзины и попробуйте снова.",
                cart_kb(can_checkout=False),
            )
            return
        if result.reason == "insufficient_funds":
            await safe_edit(
                query.message,
                f"{ICON_BLOCK} <b>Недостаточно средств</b>\n\n"
                f"К оплате: <b>{fmt_money(result.total)}</b>\n"
                f"Ваш баланс: <b>{fmt_money(result.balance)}</b>",
                back_to_main_kb(is_admin(query.from_user.id)),
            )
            return
        await safe_edit(query.message, f"{ICON_BLOCK} <b>Покупка не удалась</b>\n\nПопробуйте открыть корзину и повторить оплату.", back_to_main_kb(is_admin(query.from_user.id)))
        return

    for product in result.products or []:
        await log_purchase(
            "purchase_successful",
            product_title=product["title"],
            phone=product["phone"],
            account_name=product["first_name"] or product["username"],
            user_id=query.from_user.id,
            price=float(product["price"]),
        )
    skipped_text = ""
    if removed_dead:
        skipped_text = (
            f"\nМёртвых не выдано: <b>{len(removed_dead)}</b>\n"
            f"За них вернулось/не списано: <b>{fmt_money(removed_dead_total)}</b>\n"
        )
    if unavailable_after_check:
        skipped_text += f"Недоступных пропущено: <b>{len(unavailable_after_check)}</b>\n"
    if replaced_unavailable:
        skipped_text += f"Заменено недоступных из корзины: <b>{len(replaced_unavailable)}</b>\n"
    if removed_unavailable:
        skipped_text += f"Не удалось заменить: <b>{len(removed_unavailable)}</b>\n"
    delivered_products = result.products or []
    if len(delivered_products) > 1 and result.batch_id:
        purchases = await list_user_batch_purchases(query.from_user.id, result.batch_id)
        account_rows = []
        for idx, item in enumerate(purchases, 1):
            account_rows.append([
                inline_button(
                    text=f"{idx}. {item['title']} • {item['phone'] or '—'}",
                    callback_data=f"my_purchase:{item['purchase_id']}:0:{result.batch_id}",
                )
            ])
        await safe_edit(
            query.message,
            f"{ICON_SUCCESS} <b>Корзина оплачена</b>\n\n"
            f"Выдано товаров: <b>{len(delivered_products)}</b>\n"
            f"Списано: <b>{fmt_money(result.total)}</b>\n"
            f"{skipped_text}"
            f"{ICON_COIN} Новый баланс: <b>{fmt_money(result.balance)}</b>\n\n"
            "Выберите формат получения аккаунтов. К0d по каждому аккаунту можно получить в <b>Моих покупках</b> внутри этого заказа.",
            purchase_batch_kb(batch_id=result.batch_id, page=0, account_rows=account_rows, can_bulk_download=True),
        )
        return

    await safe_edit(
        query.message,
        f"{ICON_SUCCESS} <b>Корзина оплачена</b>\n\n"
        f"Выдано товаров: <b>{len(delivered_products)}</b>\n"
        f"Списано: <b>{fmt_money(result.total)}</b>\n"
        f"{skipped_text}"
        f"{ICON_COIN} Новый баланс: <b>{fmt_money(result.balance)}</b>\n\n"
        "Данные товаров отправляю ниже.",
        purchase_success_kb(is_admin(query.from_user.id)),
    )
    for product in delivered_products:
        await _send_purchase_delivery(query, product)


@dp.callback_query(F.data == "menu_purchases")
@dp.callback_query(F.data.startswith("menu_purchases:"))
async def menu_purchases(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    page = 0
    if ":" in query.data:
        try:
            page = max(0, int(query.data.split(":", 1)[1]))
        except ValueError:
            page = 0

    groups = await list_user_purchase_groups(query.from_user.id)
    if not groups:
        text = f"{ICON_SPARKLE} <b>Мои покупки</b>\n\nПока пусто."
        await safe_edit(query.message, text, menu_only_kb(is_admin(query.from_user.id)))
        return

    total = len(groups)
    total_pages = -(-total // PURCHASE_PAGE_SIZE)
    page = min(page, total_pages - 1)
    offset = page * PURCHASE_PAGE_SIZE
    page_items = groups[offset:offset + PURCHASE_PAGE_SIZE]
    rows = []
    for group in page_items:
        count = int(group["items_count"])
        title = group["first_title"] or "Аkkаунт"
        phone = group["first_phone"] or "—"
        if count > 1:
            label = f"{count} аккаунтов • {short_date(group['created_at'])} • {fmt_money(float(group['total_price']))}"
        else:
            label = f"{title} • {phone} • {short_date(group['created_at'])}"
        rows.append([
            inline_button(
                text=label,
                callback_data=f"purchase_batch:{group['group_id']}:{page}",
            )
        ])

    text = (
        f"{ICON_SPARKLE} <b>Мои покупки</b>\n\n"
        f"Всего заказов: <b>{total}</b>\n"
        "Выберите заказ, чтобы открыть детали."
    )
    await safe_edit(
        query.message,
        text,
        purchases_nav_kb(purchase_rows=rows, page=page, total_pages=total_pages),
    )


@dp.callback_query(F.data.startswith("purchase_batch:"))
async def purchase_batch_detail(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    try:
        _, batch_id, page_raw = query.data.split(":", 2)
        page = max(0, int(page_raw))
    except ValueError:
        await query.answer("Покупка не найдена.", show_alert=True)
        return

    purchases = await list_user_batch_purchases(query.from_user.id, batch_id)
    if not purchases:
        await query.answer("Покупка не найдена.", show_alert=True)
        return

    total = sum(float(item["price"]) for item in purchases)
    date = short_date(purchases[0]["created_at"])
    rows = []
    for idx, item in enumerate(purchases, 1):
        rows.append([
            inline_button(
                text=f"{idx}. {item['title']} • {item['phone'] or '—'}",
                callback_data=f"my_purchase:{item['purchase_id']}:{page}:{batch_id}",
            )
        ])
    downloadable = sum(1 for item in purchases if product_session_file(item))
    text = (
        f"{ICON_SPARKLE} <b>Заказ</b>\n\n"
        f"Дата: <b>{date}</b>\n"
        f"Аккаунтов: <b>{len(purchases)}</b>\n"
        f"Сумма: <b>{fmt_money(total)}</b>\n\n"
        "Выберите аккаунт:"
    )
    await safe_edit(
        query.message,
        text,
        purchase_batch_kb(
            batch_id=batch_id,
            page=page,
            account_rows=rows,
            can_bulk_download=downloadable > 0,
        ),
    )


@dp.callback_query(F.data.startswith("my_purchase:"))
async def my_purchase_detail(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    try:
        parts = query.data.split(":")
        purchase_id = int(parts[1])
        page = max(0, int(parts[2]))
        batch_id = parts[3] if len(parts) > 3 else ""
    except (IndexError, ValueError):
        await query.answer("Покупка не найдена.", show_alert=True)
        return

    purchases = await list_user_purchases(query.from_user.id)
    purchase = next((item for item in purchases if int(item["purchase_id"]) == purchase_id), None)
    if not purchase:
        await query.answer("Покупка не найдена.", show_alert=True)
        return

    account_name = purchase["first_name"] or purchase["username"] or "—"
    text = (
        f"{ICON_SPARKLE} <b>Покупка</b>\n\n"
        f"<b>Ник:</b> {html.escape(str(account_name))}\n"
        f"<b>Номер:</b> <code>{html.escape(purchase['phone'] or '—')}</code>\n"
        f"<b>Облачный пароль:</b> <code>{html.escape(purchase['twofa_password'] or 'нет')}</code>\n"
        f"<b>Заметка:</b> {render_rich_text(purchase['extra_code'] or 'нет')}\n"
        f"<b>Дата:</b> {short_date(purchase['created_at'])}\n"
        f"<b>Сумма:</b> {fmt_money(float(purchase['price']))}"
    )
    await safe_edit(query.message, text, purchase_history_detail_kb(page, int(purchase["product_id"]), batch_id or None))


async def _render_product_list(query: CallbackQuery, *, country_id: int | None, page: int) -> None:
    country = None
    active_country_names: set[str] | None = None
    if country_id is not None:
        country_row = await get_catalog_country(country_id)
        if not country_row:
            await query.answer("Кнопка страны не найдена.", show_alert=True)
            return
        country = country_row["name"]
    else:
        active_country_names = {row["name"] for row in await list_catalog_countries()}
    offset = page * PAGE_SIZE
    if active_country_names is None:
        items = await list_product_departments(country=country, offset=offset, limit=PAGE_SIZE)
        total = await count_product_departments(country=country)
    else:
        all_items = [item for item in await list_product_departments(limit=1000) if item["country"] in active_country_names]
        total = len(all_items)
        items = all_items[offset:offset + PAGE_SIZE]
    total_accounts = await count_products(country=country)
    total_pages = -(-total // PAGE_SIZE) if total > 0 else 1
    rows = []
    for item in items:
        count = int(item["stock_count"] or 0)
        count_suffix = f" • {count} шт" if count else " • 0 шт"
        rows.append([
            inline_button(
                text=f"{item['title']} • {fmt_money(float(item['price']))}{count_suffix}",
                callback_data=f"product_group:{item['sample_product_id']}:{'all' if country_id is None else f'c{country_id}'}:{page}",
            )
        ])
    if country:
        text = f"{ICON_COUNTRY} <b>Страна:</b> {html.escape(country)}\n\n{ICON_TAG} <b>Аккаунтов:</b> {total_accounts}"
    else:
        text = f"<b>Все товары</b>\n\n{ICON_TAG} <b>Аккаунтов:</b> {total_accounts}"
    prefix = "catalog_all" if country_id is None else f"catalog_country:{country_id}:"
    await safe_edit(query.message, text, product_list_kb(prefix=prefix, product_rows=rows, page=page, total_pages=total_pages, back_callback="catalog_accounts"))


@dp.callback_query(F.data.startswith("catalog_country:"))
async def catalog_country(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    _, country_id_raw, page_raw = query.data.split(":", 2)
    await _render_product_list(query, country_id=int(country_id_raw), page=int(page_raw))


@dp.callback_query(F.data.startswith("catalog_all_"))
async def catalog_all(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    await _render_product_list(query, country_id=None, page=int(query.data.rsplit("_", 1)[1]))


@dp.callback_query(F.data.startswith("product_group:"))
async def product_group_view(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    try:
        _, sample_id_raw, scope, page_raw = query.data.split(":", 3)
        sample_product_id = int(sample_id_raw)
    except (ValueError, IndexError):
        await query.answer("Товар не найден.", show_alert=True)
        return
    group = await get_product_department(sample_product_id)
    if not group:
        await query.answer("Товар не найден.", show_alert=True)
        return
    if scope == "all":
        back_callback = f"catalog_all_{page_raw}"
    elif scope.startswith("c") and scope[1:].isdigit():
        back_callback = f"catalog_country:{scope[1:]}:{page_raw}"
    else:
        back_callback = "catalog_accounts"
    await state.update_data(cart_picker_back_callback=f"product_group:{sample_product_id}:{scope}:{page_raw}")
    _stock_count, in_cart_count, available_to_add = await get_department_cart_capacity(query.from_user.id, int(group["sample_product_id"]))
    await safe_edit(
        query.message,
        product_group_public_text(group, in_cart_count=in_cart_count, available_to_add=available_to_add),
        product_group_detail_kb(
            int(group["sample_product_id"]),
            can_buy=available_to_add > 0,
            back_callback=back_callback,
        ),
    )


@dp.callback_query(F.data.startswith("cart_picker_open:"))
@dp.callback_query(F.data.startswith("cart_picker:"))
async def cart_picker_update(query: CallbackQuery, state: FSMContext):
    try:
        await ensure_known_user(query)
        parts = (query.data or "").split(":")
        sample_product_id = int(parts[1])
        selected_qty = int(parts[2])
    except (ValueError, IndexError):
        await query.answer("Некорректное количество.", show_alert=True)
        return

    await query.answer()
    try:
        await render_group_cart_picker(query, state, sample_product_id, selected_qty)
    except Exception as exc:
        logger.exception("Cart picker failed: data=%s user=%s", query.data, query.from_user.id)
        await query.message.answer(
            f"{ICON_BLOCK} <b>Не удалось открыть выбор количества</b>\n\n"
            f"<code>{html.escape(str(exc) or type(exc).__name__)}</code>"
        )


async def render_group_cart_picker(
    query: CallbackQuery,
    state: FSMContext,
    sample_product_id: int,
    selected_qty: int,
    *,
    back_callback: str | None = None,
) -> None:
    group = await get_product_department(sample_product_id)
    if not group:
        await query.answer("Товар не найден.", show_alert=True)
        return
    stock_count, in_cart_count, available_to_add = await get_department_cart_capacity(query.from_user.id, sample_product_id)
    if back_callback is None:
        data = await state.get_data()
        back_callback = data.get("cart_picker_back_callback") or "catalog_accounts"
    if available_to_add <= 0:
        await safe_edit(
            query.message,
            product_group_public_text(group) + "\n\nВсе доступные аккаунты уже в корзине или закончились.",
            product_group_detail_kb(sample_product_id, can_buy=False, back_callback=back_callback),
        )
        return
    selected_qty = max(1, min(selected_qty, available_to_add))
    text = (
        f"<b>Добавить в корзину</b>\n\n"
        f"<b>{render_rich_text(group['title'])}</b>\n"
        f"<b>Страна:</b> {html.escape(group['country'])}\n"
        f"<b>В наличии:</b> {stock_count}\n"
        f"<b>Уже в корзине:</b> {in_cart_count}\n"
        f"<b>Можно добавить ещё:</b> {available_to_add}\n"
        f"<b>Цена за 1:</b> {fmt_money(float(group['price']))}\n"
        f"<b>Итого:</b> {fmt_money(float(group['price']) * selected_qty)}"
    )
    markup = product_group_cart_kb(
        sample_product_id,
        selected_qty=selected_qty,
        max_qty=available_to_add,
        back_callback=back_callback,
    )
    await safe_edit(
        query.message,
        text,
        markup,
    )


@dp.callback_query(F.data.startswith("product_"))
async def product_view(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    body = query.data[len("product_"):]
    product_id_raw, scope, page_raw = body.split(":", 2)
    product = await get_product(int(product_id_raw))
    if not product:
        await query.answer("Товар не найден.", show_alert=True)
        return
    if scope == "all":
        back_callback = f"catalog_all_{page_raw}"
    elif scope.startswith("c") and scope[1:].isdigit():
        back_callback = f"catalog_country:{scope[1:]}:{page_raw}"
    else:
        back_callback = "catalog_accounts"
    await safe_edit(
        query.message,
        product_public_text(product),
        product_detail_kb(
            product['product_id'],
            can_buy=product['status'] == 'available',
            can_claim=is_admin(query.from_user.id) and product['status'] in {'available', 'waiting_code'},
            back_callback=back_callback,
        ),
    )


async def _send_purchase_delivery(target: Message | CallbackQuery, product) -> None:
    """Отправляет пользователю информацию о купленном товаре с инструкциями."""
    caption = purchase_delivery_text(product)
    message = target.message if isinstance(target, CallbackQuery) else target
    await message.answer(caption, reply_markup=purchase_waiting_kb(product['product_id'], "menu_catalog"))


def purchase_delivery_text(product) -> str:
    return (
        f"{ICON_PARTY} <b>Покупка оформлена</b>\n\n"
        f"{ICON_PURCHASE_TAG} Товар: <b>{render_rich_text(product['title'])}</b>\n"
        f"Телефон: <code>{html.escape(product['phone'] or '—')}</code>\n"
        f"Облачный пароль: <code>{html.escape(product['twofa_password'] or 'нет')}</code>\n"
        f"Заметка: {render_rich_text(product['extra_code'] or 'нет')}\n\n"
        "К0d можно запросить кнопкой ниже. Доступ к к0d также останется в истории покупок."
    )


@dp.callback_query(F.data.startswith("purchase_waiting:"))
async def purchase_waiting_back(query: CallbackQuery):
    await ensure_known_user(query)
    product_id = int(query.data.split(":", 1)[1])
    product = await get_product(product_id)
    if not product or product["sold_to"] != query.from_user.id:
        await query.answer("Товар не найден.", show_alert=True)
        return
    await query.answer()
    await safe_edit(query.message, purchase_delivery_text(product), purchase_waiting_kb(product_id, "menu_catalog"))


@dp.callback_query(F.data.startswith("request_code:"))
async def request_code(query: CallbackQuery):
    await ensure_known_user(query)
    product_id = int(query.data.split(":")[1])
    user_id = query.from_user.id
    
    product = await get_product(product_id)
    if not product:
        await query.answer("Товар не найден.", show_alert=True)
        return
    
    if product["sold_to"] != user_id:
        await query.answer("Этот товар не ваш.", show_alert=True)
        return
    
    if product["status"] != "sold":
        await query.answer("Статус товара не позволяет получить к0d.", show_alert=True)
        return
    
    await query.answer("Получаю к0D...")
    status_msg = await query.message.edit_text(f"{ICON_KEYBOARD} Получаю к0D...")
    
    try:
        code = await session_manager.fetch_code_from_telegram(product_id)
        
        # Логируем отправку кода
        await log_purchase("code_sent",
            product_title=product["title"],
            phone=product["phone"],
            account_name=product["first_name"] or product["username"],
            code=code,
            user_id=user_id
        )
        
        # Формируем сообщение с кодом и облачным паролем
        twofa_password = product["twofa_password"] if product["twofa_password"] else None
        twofa_text = f"\nОблачный пароль: <code>{html.escape(twofa_password)}</code>" if twofa_password else ""
        
        await status_msg.edit_text(
            f"{ICON_CHECK} <b>К0d получен</b>\n\n"
            f"{ICON_PURCHASE_TAG} <b>Товар:</b> {render_rich_text(product['title'])}\n"
            f"Телефон: <code>{html.escape(product['phone'] or '—')}</code>\n"
            f"К0d: <code>{code}</code>{twofa_text}\n\n"
            "Используйте эти данные для входа.",
            reply_markup=code_received_kb(product_id)
        )
        
    except Exception as e:
        # Логируем ошибку
        await log_purchase("purchase_error",
            product_title=product["title"],
            phone=product["phone"],
            account_name=product["first_name"] or product["username"],
            user_id=user_id,
            error=str(e)
        )
        
        twofa_text = (
            f"<b>0блачный пароль:</b> <code>{html.escape(product['twofa_password'])}</code>\n"
            if product["twofa_password"]
            else ""
        )
        await status_msg.edit_text(
            f"{ICON_BLOCK} <b>Не удалось получить к0d</b>\n\n"
            f"{ICON_PURCHASE_TAG} <b>Товар:</b> {render_rich_text(product['title'])}\n"
            f"<b>Телефон:</b> <code>{html.escape(product['phone'] or '—')}</code>\n"
            f"{twofa_text}\n"
            f"<b>Ошибка:</b> {html.escape(str(e))}\n\n"
            "Попробуйте запросить к0d немного позже. Если проблема повторится, напишите в поддержку.",
            reply_markup=code_received_kb(product_id)
        )


@dp.callback_query(F.data.startswith("user_download_session:"))
async def user_download_session(query: CallbackQuery):
    await ensure_known_user(query)
    try:
        product_id = int(query.data.split(":", 1)[1])
    except ValueError:
        await query.answer("Некорректный товар.", show_alert=True)
        return

    product = await get_owned_sold_product(query.from_user.id, product_id)
    if not product:
        await query.answer("Покупка не найдена.", show_alert=True)
        return

    session_file = product_session_file(product)
    if not session_file:
        await query.answer("Файл .session не найден.", show_alert=True)
        return

    await query.answer("Готовлю session+json...")
    with tempfile.TemporaryDirectory(prefix=f"session_json_{product_id}_") as tmp:
        tmp_dir = Path(tmp)
        try:
            archive_path, errors = build_session_json_archive([product], tmp_dir, str(product_id))
        except Exception as exc:
            await query.message.answer(
                f"{ICON_BLOCK} <b>Не удалось подготовить session+json</b>\n\n"
                f"<code>{html.escape(str(exc) or type(exc).__name__)}</code>"
            )
            return
        caption = (
            f"{ICON_PURCHASE_TAG} <b>{render_rich_text(product['title'])}</b>\n"
            f"Телефон: <code>{html.escape(product['phone'] or '—')}</code>\n"
            "Архив session+json"
        )
        if errors:
            caption += "\n\nЧасть файлов пропущена."
        await query.message.answer_document(
            FSInputFile(archive_path, filename=f"account_{product_account_label(product)}_session_json.zip"),
            caption=caption,
        )


@dp.callback_query(F.data.startswith("user_download_tdata:"))
async def user_download_tdata(query: CallbackQuery):
    await ensure_known_user(query)
    try:
        product_id = int(query.data.split(":", 1)[1])
    except ValueError:
        await query.answer("Некорректный товар.", show_alert=True)
        return

    product = await get_owned_sold_product(query.from_user.id, product_id)
    if not product:
        await query.answer("Покупка не найдена.", show_alert=True)
        return

    session_file = product_session_file(product)
    if not session_file:
        await query.answer("Файл .session не найден.", show_alert=True)
        return

    await query.answer("Готовлю tdata...")
    progress = await query.message.answer(
        f"{ICON_FOLDER} <b>Готовлю tdata</b>\n\n"
        f"{ICON_PURCHASE_TAG} <b>Товар:</b> {render_rich_text(product['title'])}\n"
        f"Телефон: <code>{html.escape(product['phone'] or '—')}</code>"
    )

    with tempfile.TemporaryDirectory(prefix=f"tdata_{product_id}_") as tmp:
        tmp_dir = Path(tmp)
        try:
            archive_path = await build_tdata_archive(product, tmp_dir)
        except Exception as exc:
            logger.exception("Could not build tdata for product #%s", product_id)
            await progress.edit_text(
                f"{ICON_BLOCK} <b>Не удалось подготовить tdata</b>\n\n"
                f"<code>{html.escape(str(exc) or type(exc).__name__)}</code>"
            )
            return

        await progress.delete()
        await query.message.answer_document(
            FSInputFile(archive_path, filename=archive_path.name),
            caption=(
                f"{ICON_PURCHASE_TAG} <b>{render_rich_text(product['title'])}</b>\n"
                f"Телефон: <code>{html.escape(product['phone'] or '—')}</code>\n"
                "Архив tdata"
            ),
        )


@dp.callback_query(F.data.startswith("batch_download_ask:"))
async def batch_download_ask(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    try:
        _, file_type, batch_id, page_raw = query.data.split(":", 3)
        page = max(0, int(page_raw))
    except ValueError:
        await query.answer("Покупка не найдена.", show_alert=True)
        return
    if file_type not in {"session", "tdata"}:
        await query.answer("Неверный формат.", show_alert=True)
        return

    purchases = await list_user_batch_purchases(query.from_user.id, batch_id)
    downloadable = [item for item in purchases if product_session_file(item)]
    if not downloadable:
        await query.answer("Нет доступных session-файлов.", show_alert=True)
        return

    fmt = "session+json ZIP" if file_type == "session" else "tdata ZIP"
    text = (
        f"{ICON_FOLDER} <b>Скачать аккаунты?</b>\n\n"
        f"Формат: <b>{html.escape(fmt)}</b>\n"
        f"Аккаунтов: <b>{len(downloadable)}</b>\n\n"
        "Бот подготовит и отправит один ZIP-архив."
    )
    await safe_edit(query.message, text, batch_download_confirm_kb(batch_id, page, file_type))


@dp.callback_query(F.data.startswith("batch_download:"))
async def batch_download(query: CallbackQuery):
    await ensure_known_user(query)
    try:
        _, file_type, batch_id, page_raw = query.data.split(":", 3)
        page = max(0, int(page_raw))
    except ValueError:
        await query.answer("Покупка не найдена.", show_alert=True)
        return
    if file_type not in {"session", "tdata"}:
        await query.answer("Неверный формат.", show_alert=True)
        return

    purchases = await list_user_batch_purchases(query.from_user.id, batch_id)
    downloadable = [item for item in purchases if product_session_file(item)]
    if not downloadable:
        await query.answer("Нет доступных session-файлов.", show_alert=True)
        return

    await query.answer("Готовлю архив...")
    await safe_edit(
        query.message,
        f"{ICON_FOLDER} <b>Готовлю архив</b>\n\n"
        f"Формат: <b>{html.escape('session+json ZIP' if file_type == 'session' else 'tdata ZIP')}</b>\n"
        f"Аккаунтов: <b>{len(downloadable)}</b>",
        purchase_batch_kb(batch_id=batch_id, page=page, can_bulk_download=True),
    )

    with tempfile.TemporaryDirectory(prefix=f"batch_{safe_filename_part(batch_id, 'batch')}_") as tmp:
        tmp_dir = Path(tmp)
        try:
            if file_type == "session":
                archive_path, errors = build_session_json_archive(downloadable, tmp_dir, batch_id)
                caption_format = "session+json"
            else:
                archive_path, errors = await build_batch_tdata_archive(downloadable, tmp_dir, batch_id)
                caption_format = "tdata"
        except Exception as exc:
            logger.exception("Could not build batch archive %s for %s", file_type, batch_id)
            await query.message.answer(
                f"{ICON_BLOCK} <b>Архив не готов</b>\n\n"
                f"<code>{html.escape(str(exc) or type(exc).__name__)}</code>"
            )
            return

        caption = (
            f"{ICON_FOLDER} <b>Архив аккаунтов</b>\n\n"
            f"Формат: <b>{html.escape(caption_format)}</b>\n"
            f"Аккаунтов в заказе: <b>{len(downloadable)}</b>"
        )
        if errors:
            caption += "\n\nПропущено:\n" + "\n".join(f"  • {html.escape(error)}" for error in errors[:5])
            if len(errors) > 5:
                caption += f"\n  ... и ещё {len(errors) - 5}"
        await query.message.answer_document(
            FSInputFile(archive_path, filename=archive_path.name),
            caption=caption,
        )

    await safe_edit(
        query.message,
        f"{ICON_SUCCESS} <b>Архив отправлен</b>\n\n"
        "К0d по каждому аккаунту можно получить в списке заказа ниже.",
        purchase_batch_kb(batch_id=batch_id, page=page, account_rows=[
            [
                inline_button(
                    text=f"{idx}. {item['title']} • {item['phone'] or '—'}",
                    callback_data=f"my_purchase:{item['purchase_id']}:{page}:{batch_id}",
                )
            ]
            for idx, item in enumerate(downloadable, 1)
        ], can_bulk_download=True),
    )


@dp.callback_query(F.data.startswith("buy_group:"))
async def product_group_buy(query: CallbackQuery):
    await ensure_known_user(query)
    try:
        sample_product_id = int(query.data.split(":", 1)[1])
    except Exception:
        await query.answer("Некорректный товар.", show_alert=True)
        return

    await query.answer("Проверяю аккаунт...")
    candidates = await list_available_products_in_department(sample_product_id, limit=1000)
    if not candidates:
        await safe_edit(query.message, f"{ICON_BLOCK} <b>В этом отделе сейчас 0 в наличии.</b>", back_to_main_kb(is_admin(query.from_user.id)))
        return

    await safe_edit(
        query.message,
        f"{ICON_KEYBOARD} <b>Формируем заказ</b>\n\n"
        "Проверяю аккаунт перед выдачей. Обычно это занимает несколько секунд.",
    )
    result = None
    product = None
    for candidate in candidates:
        if not await verify_product_alive_for_sale(candidate, context=f"Покупка из отдела пользователем {query.from_user.id}"):
            continue
        try:
            current_result = await purchase_product(query.from_user.id, int(candidate["product_id"]))
        except Exception as exc:
            logger.exception("Group purchase failed for user=%s product=%s", query.from_user.id, candidate["product_id"])
            await safe_edit(query.message, f"❌ <b>Ошибка покупки</b>\n\n{html.escape(str(exc))}", back_to_main_kb(is_admin(query.from_user.id)))
            return
        if current_result.ok:
            result = current_result
            product = await get_product(int(current_result.product_id))
            break
        if current_result.reason == "insufficient_funds":
            text = (
                f"❌ <b>Недостаточно средств</b>\n\n"
                f"<b>Цена товара:</b> {fmt_money(float(candidate['price']))}\n"
                f"{ICON_COIN} <b>Ваш баланс:</b> {fmt_money(current_result.balance)}\n\n"
                f"Пополните {ICON_COIN} баланс и попробуйте снова."
            )
            await safe_edit(query.message, text, back_to_main_kb(is_admin(query.from_user.id)))
            return
        if current_result.reason == "not_available":
            continue

    if not result or not product:
        await safe_edit(
            query.message,
            f"{ICON_BLOCK} <b>Нет живых аккаунтов</b>\n\n"
            "Я проверил доступные аккаунты этого отдела, но живых для выдачи сейчас нет. Админы уже получили уведомления по проблемным аккаунтам.",
            back_to_main_kb(is_admin(query.from_user.id)),
        )
        return

    await log_purchase("purchase_successful",
        product_title=product["title"],
        phone=product["phone"],
        account_name=product["first_name"] or product["username"],
        price=float(product["price"]),
        user_id=query.from_user.id
    )

    await _send_purchase_delivery(query, product)
    await safe_edit(
        query.message,
        product_public_text(product) + f"\n\n{ICON_SUCCESS} Покупка оформлена.\n{ICON_COIN} <b>Новый баланс:</b> {fmt_money(result.balance)}",
        back_to_main_kb(is_admin(query.from_user.id)),
    )


@dp.callback_query(F.data.startswith("buy_"))
async def product_buy(query: CallbackQuery):
    await ensure_known_user(query)
    try:
        product_id = int(query.data.split("_", 1)[1])
    except Exception:
        await query.answer("Некорректный товар.", show_alert=True)
        return

    await query.answer("Оформляю покупку...")
    candidate = await get_product(product_id)
    if not candidate:
        await safe_edit(query.message, f"{ICON_BLOCK} <b>Товар не найден.</b>", back_to_main_kb(is_admin(query.from_user.id)))
        return
    if candidate["status"] != "available":
        await safe_edit(query.message, f"{ICON_BLOCK} <b>Товар уже недоступен.</b>", back_to_main_kb(is_admin(query.from_user.id)))
        return
    await safe_edit(
        query.message,
        f"{ICON_KEYBOARD} <b>Формируем заказ</b>\n\n"
        "Проверяю аккаунт перед выдачей. Обычно это занимает несколько секунд.",
    )
    if not await verify_product_alive_for_sale(candidate, context=f"Покупка конкретного товара пользователем {query.from_user.id}"):
        await safe_edit(
            query.message,
            f"{ICON_BLOCK} <b>Аккаунт снят с продажи</b>\n\n"
            "Перед выдачей проверил аккаунт, он оказался мёртвым. Админы уже получили уведомление, деньги не списаны.",
            back_to_main_kb(is_admin(query.from_user.id)),
        )
        return
    try:
        result = await purchase_product(query.from_user.id, product_id)
    except Exception as exc:
        logger.exception("Purchase failed for user=%s product=%s", query.from_user.id, product_id)
        await safe_edit(query.message, f"❌ <b>Ошибка покупки</b>\n\n{html.escape(str(exc))}", back_to_main_kb(is_admin(query.from_user.id)))
        return

    if not result.ok:
        if result.reason == "insufficient_funds":
            product = await get_product(product_id)
            price = fmt_money(float(product["price"])) if product else "—"
            text = (
                f"❌ <b>Недостаточно средств</b>\n\n"
                f"<b>Цена товара:</b> {price}\n"
                f"{ICON_COIN} <b>Ваш баланс:</b> {fmt_money(result.balance)}\n\n"
                f"Пополните {ICON_COIN} баланс и попробуйте снова."
            )
            await safe_edit(query.message, text, back_to_main_kb(is_admin(query.from_user.id)))
            return
        if result.reason == "not_available":
            await safe_edit(query.message, f"{ICON_BLOCK} <b>Товар уже недоступен.</b>", back_to_main_kb(is_admin(query.from_user.id)))
            return
        await safe_edit(query.message, f"{ICON_BLOCK} <b>Покупка не удалась.</b>", back_to_main_kb(is_admin(query.from_user.id)))
        return

    product = await get_product(product_id)
    if not product:
        await safe_edit(query.message, "✅ Покупка оформлена, но карточка товара не найдена.", back_to_main_kb(is_admin(query.from_user.id)))
        return
    
    await log_purchase("purchase_successful",
        product_title=product["title"],
        phone=product["phone"],
        account_name=product["first_name"] or product["username"],
        price=float(product["price"]),
        user_id=query.from_user.id
    )
    
    await _send_purchase_delivery(query, product)
    await safe_edit(
        query.message,
        product_public_text(product) + f"\n\n{ICON_SUCCESS} Покупка оформлена.\n{ICON_COIN} <b>Новый баланс:</b> {fmt_money(result.balance)}",
        back_to_main_kb(is_admin(query.from_user.id)),
    )


@dp.callback_query(F.data.startswith("admin_verify_account:"))
async def admin_verify_account(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    
    try:
        product_id = int(query.data.split(":", 1)[1])
    except Exception:
        await query.answer("Некорректный товар.", show_alert=True)
        return
    
    product = await get_product(product_id)
    if not product:
        await query.answer("Товар не найден.", show_alert=True)
        return
    
    await query.answer("Проверяю аkkаунт...")
    
    try:
        result = await session_manager.verify_account_alive(product_id)
        
        if result["alive"]:
            status_text = (
                f"<b>Аккаунт живой</b>\n\n"
                f"User ID: <code>{result['user_id']}</code>\n"
                f"Телефон: <code>{result['phone']}</code>\n"
                f"Username: <b>{result['username']}</b>"
            )
        else:
            status_text = f"<b>Аккаунт не прошел проверку</b>\n\n{result['error']}"
        
        await safe_edit(query.message, status_text, admin_product_detail_kb(product_id))
    
    except Exception as e:
        logger.exception("Ошибка при проверке аккаунта")
        await query.answer(f"Ошибка: {str(e)}", show_alert=True)


@dp.callback_query(F.data == "admin_scan_accounts")
async def admin_scan_accounts(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id): return
    await query.answer()
    data = await state.get_data()
    interval = int(data.get("scan_interval") or 60)
    limit = int(data.get("scan_limit") or 5)
    await state.update_data(scan_interval=interval, scan_limit=limit)
    text = (
        "<b>Глубокая проверка аkkаунтов</b>\n\n"
        "Эта проверка подключается к сессuu каждого аkkаунта и осторожно проверяет, валидна ли она. "
        "Используйте только когда это действительно нужно.\n\n"
        f"<b>Интервал:</b> {interval} сек\n"
        f"<b>Лимит:</b> {limit} аkkаунтов за запуск\n\n"
        "Рекомендуемо: 60-120 секунд и небольшой лимит."
    )
    await safe_edit(query.message, text, admin_scan_settings_kb(interval, limit))


@dp.callback_query(F.data.startswith("admin_scan_interval:"))
async def admin_scan_interval(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id): return
    interval = int(query.data.split(":", 1)[1])
    await state.update_data(scan_interval=interval)
    await admin_scan_accounts(query, state)


@dp.callback_query(F.data.startswith("admin_scan_limit:"))
async def admin_scan_limit(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id): return
    limit = int(query.data.split(":", 1)[1])
    await state.update_data(scan_limit=limit)
    await admin_scan_accounts(query, state)


@dp.callback_query(F.data == "admin_scan_interval_custom")
async def admin_scan_interval_custom(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id): return
    await query.answer()
    await state.set_state(AdminScanStates.waiting_interval)
    await safe_edit(
        query.message,
        "<b>Интервал проверки</b>\n\nВведите задержку между аккаунтами в секундах.\nМинимум: <b>20</b>, максимум: <b>600</b>.",
        cancel_flow_kb("admin_scan_accounts"),
    )


@dp.message(AdminScanStates.waiting_interval)
async def admin_scan_interval_message(message: Message, state: FSMContext):
    await ensure_known_user(message)
    if not is_admin(message.from_user.id): return
    try:
        interval = int((message.text or "").strip())
    except ValueError:
        await message.answer("Введите число секунд.")
        return
    if interval < 20 or interval > 600:
        await message.answer("Интервал должен быть от 20 до 600 секунд.")
        return
    await state.update_data(scan_interval=interval)
    await state.set_state(None)
    data = await state.get_data()
    limit = int(data.get("scan_limit") or 5)
    await message.answer(
        "<b>Глубокая проверка аккаунтов</b>\n\n"
        f"<b>Интервал:</b> {interval} сек\n"
        f"<b>Лимит:</b> {limit} аккаунтов за запуск",
        reply_markup=admin_scan_settings_kb(interval, limit),
    )


@dp.callback_query(F.data == "admin_scan_limit_custom")
async def admin_scan_limit_custom(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id): return
    await query.answer()
    await state.set_state(AdminScanStates.waiting_limit)
    await safe_edit(
        query.message,
        "<b>Лимит проверки</b>\n\nВведите сколько аккаунтов проверить за один запуск.\nМинимум: <b>1</b>, максимум: <b>20</b>.",
        cancel_flow_kb("admin_scan_accounts"),
    )


@dp.message(AdminScanStates.waiting_limit)
async def admin_scan_limit_message(message: Message, state: FSMContext):
    await ensure_known_user(message)
    if not is_admin(message.from_user.id): return
    try:
        limit = int((message.text or "").strip())
    except ValueError:
        await message.answer("Введите число аккаунтов.")
        return
    if limit < 1 or limit > 20:
        await message.answer("Лимит должен быть от 1 до 20 аккаунтов.")
        return
    await state.update_data(scan_limit=limit)
    await state.set_state(None)
    data = await state.get_data()
    interval = int(data.get("scan_interval") or 60)
    await message.answer(
        "<b>Глубокая проверка аккаунтов</b>\n\n"
        f"<b>Интервал:</b> {interval} сек\n"
        f"<b>Лимит:</b> {limit} аккаунтов за запуск",
        reply_markup=admin_scan_settings_kb(interval, limit),
    )


@dp.callback_query(F.data == "admin_scan_confirm")
async def admin_scan_confirm(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id): return
    data = await state.get_data()
    interval = int(data.get("scan_interval") or 60)
    limit = int(data.get("scan_limit") or 5)
    await query.answer()
    await safe_edit(
        query.message,
        "<b>Подтвердите глубокую проверку</b>\n\n"
        "Бот будет по очереди подключаться к сеccuям аkkаунTов. "
        "Это не локальный скан файлов, поэтому запускать нужно осторожно.\n\n"
        f"<b>Интервал:</b> {interval} сек\n"
        f"<b>Лимит:</b> {limit} аккаунтов",
        admin_scan_confirm_kb(),
    )


@dp.callback_query(F.data == "admin_scan_start")
async def admin_scan_start(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id): return
    data = await state.get_data()
    interval = max(20, min(600, int(data.get("scan_interval") or 60)))
    limit = max(1, min(20, int(data.get("scan_limit") or 5)))
    await query.answer("Запускаю осторожную проверку.")
    status_msg = query.message
    await safe_edit(
        status_msg,
        "<b>Глубокая проверка аккаунтов</b>\n\n"
        "Запускаю осторожную проверку. Не закрывайте бот до завершения.",
    )

    products = await list_products(status="available", limit=limit)
    total = len(products)
    success = 0
    failed_data = {}
    
    for idx, p in enumerate(products, 1):
        await safe_edit(
            status_msg,
            f"<b>Глубокая проверка аккаунтов</b>\n\n"
            f"Проверяю: <b>{idx}/{total}</b>\n"
            f"Товар: <code>#{p['product_id']}</code>\n"
            f"Успешно: <b>{success}</b>\n"
            f"Ошибок: <b>{len(failed_data)}</b>\n\n"
            f"Интервал между аккаунтами: <b>{interval} сек</b>",
        )
        res = await session_manager.verify_account_alive(p["product_id"])
        if res.get("alive"):
            await update_product_info(p["product_id"], 
                phone=res["phone"],
                telegram_id=res["user_id"],
                username=res["username"],
                first_name=res["first_name"]
            )
            success += 1
        else:
            failed_data[p["product_id"]] = res.get("error", "Unknown error")
        if idx < total:
            await asyncio.sleep(interval)
    await state.update_data(last_scan_errors=failed_data)
    
    kb = []
    if failed_data:
        kb.append([InlineKeyboardButton(text="Посмотреть ошибки", callback_data="admin_scan_failed_list")])
    kb.append([InlineKeyboardButton(text="В админку", callback_data="admin_home", icon_custom_emoji_id=BTN_ICON_ADMIN)])
    
    await status_msg.edit_text(
        f"<b>Сканирование завершено</b>\n\n"
        f"Проверено: <b>{total}</b>\n"
        f"Успешно: <b>{success}</b>\n"
        f"Ошибок: <b>{len(failed_data)}</b>\n\n"
        f"Интервал: <b>{interval} сек</b>\n"
        f"Лимит: <b>{limit}</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb)
    )

@dp.callback_query(F.data == "admin_scan_failed_list")
async def admin_scan_failed_list(query: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    errors = data.get("last_scan_errors", {})
    if not errors:
        await query.answer("Ошибок не найдено.", show_alert=True)
        return
    
    rows = []
    for pid in errors.keys():
        rows.append([InlineKeyboardButton(text=f"Товар #{pid}", callback_data=f"scan_err_det:{pid}")])
    
    rows.append([InlineKeyboardButton(text="Назад", callback_data="admin_home", icon_custom_emoji_id=BTN_ICON_BACK)])
    await safe_edit(query.message, "<b>Ошибки сканирования</b>\n\nВыберите товар:", InlineKeyboardMarkup(inline_keyboard=rows))

@dp.callback_query(F.data.startswith("scan_err_det:"))
async def admin_scan_failed_detail(query: CallbackQuery, state: FSMContext):
    pid = int(query.data.split(":")[1])
    data = await state.get_data()
    errors = data.get("last_scan_errors", {})
    err_text = errors.get(pid, "No error saved")
    
    product = await get_product(pid)
    if not product:
        await query.answer("Товар не найден.")
        return
        
    text = (
        f"<b>Ошибка сканирования товара #{pid}</b>\n\n"
        f"Ошибка: <code>{html.escape(err_text)}</code>\n\n"
        f"{product_admin_text(product)}"
    )
    
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Удалить товар", callback_data=f"admin_remove_{pid}")],
        [InlineKeyboardButton(text="К списку ошибок", callback_data="admin_scan_failed_list")]
    ])
    await safe_edit(query.message, text, kb)


@dp.message(AdminSearchUserStates.waiting_user_query)
async def admin_user_search_process(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    query = (message.text or "").strip()
    if not query.isdigit():
        await message.answer("Введите числовой User ID.")
        return

    user_id = int(query)
    user = await get_user(user_id)
    if not user:
        await message.answer("Пользователь не найден в базе.")
        return

    await state.clear()
    balance = float(user["balance"])
    text = (
        f"<b>Пользователь</b>\n\n"
        f"ID: <code>{user['user_id']}</code>\n"
        f"<b>Ник:</b> @{html.escape(user['username'] or '—')}\n"
        f"<b>Имя:</b> {html.escape(user['first_name'] or '—')}\n"
        f"{ICON_COIN} <b>Баланс:</b> {fmt_money(balance)}\n"
        f"<b>Первый вход:</b> {user['joined_at'][:10]}"
    )
    await message.answer(text, reply_markup=admin_user_manage_kb(user_id))


@dp.callback_query(F.data == "admin_product_search")
async def admin_product_search_start(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    await state.set_state(AdminSearchProductStates.waiting_product_query)
    await safe_edit(
        query.message,
        "<b>Поиск товара</b>\n\n"
        "Отправьте номер телефона аkkаунта или ID товара.",
        cancel_flow_kb("admin_home"),
    )


@dp.message(AdminSearchProductStates.waiting_product_query)
async def admin_product_search_process(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    query_text = (message.text or "").strip()
    if len(query_text) < 2:
        await message.answer("Отправьте номер телефона или ID товара.")
        return
    products = await search_products(query_text, limit=20)
    if not products:
        await message.answer("Товар не найден.", reply_markup=admin_home_kb())
        await state.clear()
        return
    await state.clear()
    if len(products) == 1:
        product = products[0]
        has_session = has_server_session(product)
        can_fetch_code = has_session and product["status"] in {"waiting_code", "verifying", "sold"}
        await message.answer(
            product_admin_text(product),
            reply_markup=admin_product_detail_kb(
                int(product["product_id"]),
                back_callback="admin_product_search",
                can_terminate_sessions=has_session,
                can_fetch_code=can_fetch_code,
            ),
        )
        return

    rows = []
    for product in products:
        rows.append([
            InlineKeyboardButton(
                text=f"#{product['product_id']} {product['phone'] or '—'} • {product['title']}",
                callback_data=f"admin_stock_product:{product['product_id']}:search",
            )
        ])
    await message.answer(
        f"<b>Найдено товаров:</b> {len(products)}\n\nВыберите нужный:",
        reply_markup=admin_product_search_results_kb(rows),
    )


@dp.callback_query(F.data.startswith("admin_user_purchases:"))
async def admin_user_purchases_view(query: CallbackQuery):
    if not is_admin(query.from_user.id): return
    target_id = int(query.data.split(":")[1])
    purchases = await list_user_purchases(target_id)
    if not purchases:
        await query.answer("У этого пользователя нет покупок.", show_alert=True)
        return
    
    lines = [f"<b>История покупок пользователя</b> <code>{target_id}</code>\n"]
    lines.append(f"Всего покупок: <b>{len(purchases)}</b>\n")
    
    for idx, p in enumerate(purchases[:20], 1):
        product_id = p['product_id']
        title = p['title']
        country = p['country']
        price = fmt_money(float(p['price']))
        date = p['created_at'][:10] if p['created_at'] else "—"
        account_name = p['first_name'] or p['username'] or '—'
        lines.append(f"{idx}. <b>{account_name}</b> | {title} ({country}) - {price} [{date}]")
        lines.append(f"   ID товара: <code>{product_id}</code>")
    
    text = "\n".join(lines)
    
    # Добавляем кнопки для каждого товара
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for idx, p in enumerate(purchases[:5], 1):
        kb.inline_keyboard.append([
            InlineKeyboardButton(
                text=f"Детали товара #{idx}: {p['title'][:20]}",
                callback_data=f"admin_purchase_detail:{p['product_id']}"
            )
        ])
    
    kb.inline_keyboard.append([InlineKeyboardButton(text="Назад", callback_data=f"admin_home", icon_custom_emoji_id=BTN_ICON_BACK)])
    
    await query.message.answer(text, reply_markup=kb)
    await query.answer()


@dp.callback_query(F.data.startswith("admin_purchase_detail:"))
async def admin_purchase_detail_view(query: CallbackQuery):
    """Показывает полную информацию о товаре из истории покупок админу (с session и tdata)"""
    if not is_admin(query.from_user.id): return
    
    product_id = int(query.data.split(":")[1])
    product = await get_product(product_id)
    
    if not product:
        await query.answer("Товар не найден.", show_alert=True)
        return
    
    # Читаем session файл если существует
    session_info = "—"
    if product['session_path']:
        try:
            session_file = Path(product['session_path'])
            if session_file.exists():
                session_info = f"Существует\n{product['session_path']}"
            else:
                session_info = "Файл не найден"
        except:
            session_info = "Ошибка чтения"
    
    # Читаем tdata если существует
    tdata_info = "—"
    if product['session_path']:
        try:
            session_file = Path(product['session_path'])
            tdata_path = session_file.parent / "tdata"
            if tdata_path.exists():
                tdata_info = f"Существует\n{tdata_path}"
            else:
                tdata_info = "Папка не найдена"
        except:
            tdata_info = "Ошибка чтения"
    
    text = (
        "<b>Полная информация о товаре</b>\n\n"
        f"<b>Основная информация:</b>\n"
        f"<b>Товар:</b> {render_rich_text(product['title'])}\n"
        f"🆔 <b>ID:</b> <code>{product['product_id']}</code>\n"
        f"{ICON_COUNTRY} <b>Страна:</b> {html.escape(product['country'])}\n"
        f"<b>Цена:</b> {fmt_money(float(product['price']))}\n"
        f"<b>Статус:</b> {html.escape(product['status'])}\n\n"
        
        f"<b>Данные аkkаунта:</b>\n"
        f"<b>Телефон:</b> <code>{html.escape(product['phone'] or '—')}</code>\n"
        f"<b>Username:</b> {html.escape(product['username'] or '—')}\n"
        f"<b>Имя:</b> {html.escape(product['first_name'] or '—')}\n"
        f"🔐 <b>2FA пароль:</b> {html.escape(product['twofa_password'] or 'нет')}\n"
        f"<b>К0d/заметка:</b> {render_rich_text(product['extra_code'] or 'нет')}\n"
        f"<b>Описание:</b> {render_rich_text(product['description'] or '—')}\n\n"
        
        f"<b>Файлы сессии:</b>\n"
        f"Session: {session_info}\n"
        f"tdata: {tdata_info}\n\n"
        
        f"<b>История продажи:</b>\n"
        f"Покупатель: <code>{product['sold_to'] or '—'}</code>\n"
        f"Дата продажи: <b>{product['sold_at'][:10] if product['sold_at'] else '—'}</b>\n"
        f"Цена продажи: <b>{fmt_money(float(product['sold_price'])) if product['sold_price'] else '—'}</b>"
    )
    
    await query.message.answer(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Назад к покупкам", callback_data="admin_home", icon_custom_emoji_id=BTN_ICON_BACK)]
    ]))
    await query.answer()


@dp.callback_query(F.data.startswith("admin_user_topup:"))
async def admin_user_topup_btn(query: CallbackQuery, state: FSMContext):
    if not is_admin(query.from_user.id): return
    target_id = int(query.data.split(":")[1])
    await state.update_data(target_user_id=target_id)
    await state.set_state(AdminTopUpStates.waiting_amount)
    await safe_edit(query.message, f"<b>{ICON_COIN} Выдача баланса</b>\n\nПользователь: <code>{target_id}</code>\nВведите сумму:", cancel_flow_kb("admin_home"))
    await query.answer()


@dp.callback_query(F.data.startswith("admin_user_reset:"))
async def admin_user_reset_btn(query: CallbackQuery):
    if not is_admin(query.from_user.id): return
    target_id = int(query.data.split(":")[1])
    
    # Обнуляем баланс
    await add_balance(
        target_id,
        -await get_balance(target_id),
        "admin_reset",
        f"Баланс обнулён администратором {query.from_user.id}"
    )
    
    user = await get_user(target_id)
    text = (
        f"<b>Пользователь</b>\n\n"
        f"ID: <code>{user['user_id']}</code>\n"
        f"<b>Ник:</b> @{html.escape(user['username'] or '—')}\n"
        f"{ICON_COIN} <b>Баланс:</b> 0.00 {settings.currency} (обнулен)\n"
    )
    
    try:
        await query.message.edit_text(text, reply_markup=admin_user_manage_kb(target_id))
    except:
        await query.message.answer(text, reply_markup=admin_user_manage_kb(target_id))
        
    await query.answer("Баланс обнулен.")
    try:
        await bot.send_message(target_id, f"{ICON_COIN} Ваш баланс был обнулен администратором.")
    except:
        pass


@dp.callback_query(F.data.startswith("admin_claim_ask:"))
async def admin_claim_ask(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not await require_admin(query):
        return
    try:
        product_id = int(query.data.split(":", 1)[1])
    except Exception:
        await query.answer("Некорректный товар.", show_alert=True)
        return
    product = await get_product(product_id)
    if not product:
        await query.answer("Товар не найден.", show_alert=True)
        return
    if product["status"] not in {"available", "waiting_code"}:
        await query.answer("Этот аккаунт уже не в наличии.", show_alert=True)
        return
    back_callback = "admin_home" if query.data.endswith(":admin") else "admin_stock_product:%s" % product_id
    text = (
        "<b>Подтвердите действие</b>\n\n"
        f"Забрать аkkаунт со склада?\n\n"
        f"{ICON_PURCHASE_TAG} <b>Товар:</b> {render_rich_text(product['title'])}\n"
        f"<b>Телефон:</b> <code>{html.escape(product['phone'] or '—')}</code>"
    )
    await safe_edit(query.message, text, admin_claim_confirm_kb(product_id, back_callback))


@dp.callback_query(F.data.startswith("admin_claim:"))
async def admin_claim_product(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not await require_admin(query):
        return
    try:
        parts = query.data.split(":", 2)
        product_id = int(parts[1])
        back_callback = parts[2] if len(parts) > 2 else "admin_home"
    except Exception:
        await query.answer("Некорректный товар.", show_alert=True)
        return
    product = await get_product(product_id)
    if not product:
        await query.answer("Товар не найден.", show_alert=True)
        return
    if product["status"] not in {"available", "waiting_code"}:
        await query.answer("Этот аккаунт уже не в наличии.", show_alert=True)
        return
    if not await claim_product_for_admin(query.from_user.id, product_id):
        await query.answer("Не удалось забрать аккаунт.", show_alert=True)
        return
    product = await get_product(product_id)
    await log_purchase("admin_action", action=f"Админ забрал аккаунт #{product_id} ({product['phone']})", admin_id=query.from_user.id)
    await safe_edit(query.message, "Аккаунт забран со склада.\n\n" + product_admin_text(product), back_to_main_kb(True))

@dp.callback_query(F.data == "admin_home")
async def admin_home(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        await query.answer("Нет доступа.", show_alert=True)
        return
    await safe_edit(query.message, "<b>Админка</b>", admin_home_kb())


def broadcast_confirm_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Отправить всем", callback_data="admin_broadcast_send", icon_custom_emoji_id=BTN_ICON_CHECK)],
        [InlineKeyboardButton(text="Отменить", callback_data="cancel_flow:admin_home", icon_custom_emoji_id=BTN_ICON_CANCEL)],
    ])


@dp.callback_query(F.data == "admin_broadcast")
async def admin_broadcast_start(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    await state.set_state(AdminBroadcastStates.waiting_text)
    await safe_edit(
        query.message,
        "<b>Рассылка</b>\n\n"
        "Отправьте текст рассылки. Premium emoji можно отправить обычным сообщением или вставить HTML вида:\n"
        '<code>&lt;tg-emoji emoji-id="6028338546736107668"&gt;⭐️&lt;/tg-emoji&gt;</code>',
        cancel_flow_kb("admin_home"),
    )


@dp.message(AdminBroadcastStates.waiting_text)
async def admin_broadcast_text(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    text = sanitize_admin_text(message)
    if not text:
        await message.answer("Текст рассылки пустой.")
        return
    await state.update_data(broadcast_text=text)
    await message.answer(
        "<b>Предпросмотр рассылки</b>\n\n"
        f"{text}",
        reply_markup=broadcast_confirm_kb(),
    )


async def send_broadcast_to_user(user_id: int, text: str) -> str:
    try:
        await bot.send_message(user_id, text)
        return "sent"
    except TelegramRetryAfter as exc:
        await asyncio.sleep(float(getattr(exc, "retry_after", 1)) + 1)
        try:
            await bot.send_message(user_id, text)
            return "sent"
        except TelegramForbiddenError:
            return "unavailable"
        except TelegramBadRequest:
            return "unavailable"
        except TelegramAPIError:
            return "error"
        except Exception:
            logger.exception("Unexpected broadcast retry error for user %s", user_id)
            return "error"
    except TelegramForbiddenError:
        return "unavailable"
    except TelegramBadRequest:
        return "unavailable"
    except TelegramAPIError:
        return "error"
    except Exception:
        logger.exception("Unexpected broadcast error for user %s", user_id)
        return "error"


@dp.callback_query(F.data == "admin_broadcast_send")
async def admin_broadcast_send(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer("Запускаю рассылку...")
    if not is_admin(query.from_user.id):
        return
    data = await state.get_data()
    text = data.get("broadcast_text")
    if not text:
        await safe_edit(query.message, "Текст рассылки не найден.", admin_home_kb())
        await state.clear()
        return
    user_ids = await list_user_ids()
    sent = 0
    unavailable = 0
    errors = 0
    await safe_edit(
        query.message,
        f"{ICON_KEYBOARD} <b>Рассылка запущена</b>\n\n"
        f"Пользователей: <b>{len(user_ids)}</b>",
    )
    for user_id in user_ids:
        status = await send_broadcast_to_user(user_id, text)
        if status == "sent":
            sent += 1
        elif status == "unavailable":
            unavailable += 1
        else:
            errors += 1
        await asyncio.sleep(0.06)
    await state.clear()
    await safe_edit(
        query.message,
        f"{ICON_SUCCESS} <b>Рассылка завершена</b>\n\n"
        f"Всего пользователей: <b>{len(user_ids)}</b>\n"
        f"Отправлено: <b>{sent}</b>\n"
        f"Недоступно/заблокировали: <b>{unavailable}</b>\n"
        f"Ошибок: <b>{errors}</b>",
        admin_home_kb(),
    )


@dp.callback_query(F.data == "admin_stats")
async def admin_stats(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    stats = await get_stats()
    
    text = (
        "<b>Статистика</b>\n\n"
        f"Пользователей: <b>{stats['users']}</b>\n"
        f"<b>В наличии:</b> {stats['available']}\n"
        f"<b>Сумма склада:</b> {fmt_money(stats['stock_value'])}\n"
        f"Ожидают к0d: <b>{stats.get('waiting', 0)}</b>\n"
        f"Продано: <b>{stats['sold']}</b>\n"
        f"Выручка: <b>{fmt_money(stats['revenue'])}</b>\n"
        f"Аккаунты: <b>{fmt_money(stats.get('accounts_revenue', 0))}</b>\n"
        f"Услуги: <b>{fmt_money(stats.get('services_revenue', 0))}</b>"
    )
    await safe_edit(query.message, text, admin_stats_kb())


@dp.callback_query(F.data == "admin_export_database")
async def admin_export_database(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    
    await query.message.edit_text("Создаю Excel файл...")
    
    try:
        filepath = await create_full_database_excel()
        if filepath:
            await query.message.answer_document(
                FSInputFile(filepath),
                caption="<b>Экспорт базы данных</b>\n\nПользователи, покупки и товары.",
                parse_mode=ParseMode.HTML
            )
            await query.message.edit_text("Excel файл отправлен.")
        else:
            await query.message.edit_text("Не удалось создать Excel файл.")
    except Exception as e:
        logger.error(f"Ошибка при создании Excel: {e}")
        await query.message.edit_text(f"Ошибка: {html.escape(str(e))}")


@dp.callback_query(F.data == "admin_reset_stats")
async def admin_reset_stats(query: CallbackQuery):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id):
        return
    
    # Подтверждение перед сбросом
    confirm_kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="Да, сбросить", callback_data="admin_reset_stats_confirm"),
            InlineKeyboardButton(text="Отменить", callback_data="admin_stats", icon_custom_emoji_id=BTN_ICON_CANCEL),
        ]
    ])
    
    await query.answer()
    await safe_edit(query.message, "<b>Подтвердите сброс</b>\n\nИстория покупок и балансы будут очищены. Товары сохранятся.", confirm_kb)


@dp.callback_query(F.data == "admin_reset_revenue")
async def admin_reset_revenue(query: CallbackQuery):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id):
        return
    confirm_kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="Да, сбросить", callback_data="admin_reset_revenue_confirm"),
            InlineKeyboardButton(text="Отменить", callback_data="admin_stats", icon_custom_emoji_id=BTN_ICON_CANCEL),
        ]
    ])
    await query.answer()
    await safe_edit(
        query.message,
        "<b>Сбросить выручку?</b>\n\nИстория продаж сохранится, но сумма выручки в статистике начнет считаться заново.",
        confirm_kb,
    )


@dp.callback_query(F.data == "admin_reset_revenue_confirm")
async def admin_reset_revenue_confirm(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    reset_at = await reset_revenue_stats()
    await log_purchase("admin_action", action="Сброшена выручка в статистике", admin_id=query.from_user.id)
    stats = await get_stats()
    text = (
        "<b>Выручка сброшена</b>\n\n"
        f"<b>Сумма склада:</b> {fmt_money(stats['stock_value'])}\n"
        f"Выручка: <b>{fmt_money(stats['revenue'])}</b>\n"
        f"Аккаунты: <b>{fmt_money(stats.get('accounts_revenue', 0))}</b>\n"
        f"Услуги: <b>{fmt_money(stats.get('services_revenue', 0))}</b>\n"
        f"Считается с: <code>{html.escape(reset_at[:19])}</code>"
    )
    await safe_edit(query.message, text, admin_stats_kb())


@dp.callback_query(F.data == "admin_reset_stats_confirm")
async def admin_reset_stats_confirm(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    
    try:
        await reset_stats()
        await log_purchase("admin_action", action="Сброс статистики и балансов", admin_id=query.from_user.id)
        await query.message.edit_text(
            "<b>Статистика сброшена</b>\n\n"
            "Товары сохранены.\n"
            "История покупок очищена.\n"
            "Балансы очищены.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="В админку", callback_data="admin_home", icon_custom_emoji_id=BTN_ICON_ADMIN)]
            ])
        )
    except Exception as e:
        logger.error(f"Ошибка при сбросе статистики: {e}")
        await query.message.edit_text(
            f"<b>Ошибка при сбросе статистики</b>\n\n"
            f"<code>{str(e)[:100]}</code>\n\n"
            "Попробуйте позже.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="В админку", callback_data="admin_home", icon_custom_emoji_id=BTN_ICON_ADMIN)]
            ])
        )




@dp.callback_query(F.data == "admin_catalog")
async def admin_catalog(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    text = (
        "<b>Страны каталога</b>\n\n"
        "Эти страны показываются в каталоге отдельными кнопками. "
        "При добавлении аккаунта товар можно положить в одну из них."
    )
    await safe_edit(query.message, text, admin_catalog_kb(await build_admin_country_rows()))


async def render_admin_country(message: Message, country_id: int) -> None:
    country = await get_catalog_country(country_id)
    if not country:
        await safe_edit(message, "Кнопка страны не найдена.", admin_catalog_kb(await build_admin_country_rows()))
        return
    total = await count_products(country=country["name"])
    groups = await list_product_departments(country=country["name"], limit=20)
    group_rows = []
    for group in groups:
        stock_count = int(group["stock_count"] or 0)
        stock_text = f"{stock_count} шт" if stock_count else "нет в наличии"
        group_rows.append([
            inline_button(
                text=f"{group['title']} • {fmt_money(float(group['price']))} • {stock_text}",
                callback_data=f"admin_product_group:{group['sample_product_id']}:{country_id}",
            )
        ])
    text = (
        "<b>Страна каталога</b>\n\n"
        f"Название: <b>{html.escape(country['name'])}</b>\n"
        f"<b>Premium-флаг:</b> <code>{html.escape(country['icon_custom_emoji_id'] or 'не задан')}</code>\n"
        f"<b>В наличии:</b> {total}\n"
        f"<b>Типов товара:</b> {len(groups)}"
    )
    await safe_edit(message, text, admin_country_kb(country_id, group_rows))


@dp.callback_query(F.data.startswith("admin_country:"))
async def admin_country_view(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    country_id = int(query.data.rsplit(":", 1)[1])
    await render_admin_country(query.message, country_id)


@dp.callback_query(F.data.startswith("admin_department_create:"))
async def admin_department_create_start(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    try:
        country_id = int(query.data.split(":", 1)[1])
    except ValueError:
        await query.answer("Страна не найдена.", show_alert=True)
        return
    country = await get_catalog_country(country_id)
    if not country:
        await query.answer("Кнопка страны не найдена.", show_alert=True)
        return
    await state.update_data(department_country_id=country_id, department_country_name=country["name"])
    await state.set_state(AdminCatalogStates.waiting_department_title)
    await safe_edit(
        query.message,
        "<b>Создать отдел</b>\n\n"
        f"{ICON_COUNTRY} <b>Страна:</b> {html.escape(country['name'])}\n\n"
        "Отправьте название отдела, например <code>+7 RU | Рег имо</code>.",
        cancel_flow_kb(f"admin_country:{country_id}"),
    )


@dp.message(AdminCatalogStates.waiting_department_title)
async def admin_department_create_title(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    title = sanitize_admin_text(message)
    if len(plain_button_text(title)) < 3:
        await message.answer("Название отдела слишком короткое.")
        return
    await state.update_data(department_title=title)
    await state.set_state(AdminCatalogStates.waiting_department_price)
    await message.answer(
        f"Название отдела: <b>{render_rich_text(title)}</b>\n\n"
        f"Теперь отправьте цену в {html.escape(settings.currency)}. Пример: <code>25</code>.",
        reply_markup=cancel_flow_kb(f"admin_country:{(await state.get_data()).get('department_country_id', 0)}"),
    )


@dp.message(AdminCatalogStates.waiting_department_price)
async def admin_department_create_price(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    price = parse_float(message.text or "")
    if price is None or price < 0:
        await message.answer("Нужна корректная цена от 0.")
        return
    await state.update_data(department_price=price)
    await state.set_state(AdminCatalogStates.waiting_department_description)
    await message.answer(
        "Отправьте описание для карточки отдела. Если не нужно — отправьте <code>-</code>.",
        reply_markup=cancel_flow_kb(f"admin_country:{(await state.get_data()).get('department_country_id', 0)}"),
    )


@dp.message(AdminCatalogStates.waiting_department_description)
async def admin_department_create_description(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    description = sanitize_admin_text(message)
    await state.update_data(department_description="" if description == "-" else description)
    await state.set_state(AdminCatalogStates.waiting_department_extra_code)
    await message.answer(
        "Отправьте заметку после покупки. Если не нужно — отправьте <code>-</code>.",
        reply_markup=cancel_flow_kb(f"admin_country:{(await state.get_data()).get('department_country_id', 0)}"),
    )


@dp.message(AdminCatalogStates.waiting_department_extra_code)
async def admin_department_create_finish(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    country_id = int(data.get("department_country_id") or 0)
    title = str(data.get("department_title") or "")
    price = float(data.get("department_price") or 0)
    description = str(data.get("department_description") or "")
    extra_code_raw = sanitize_admin_text(message)
    extra_code = "" if extra_code_raw == "-" else extra_code_raw
    try:
        department_id = await create_catalog_department(
            country_id=country_id,
            title=title,
            price=price,
            description=description,
            extra_code=extra_code,
            created_by=message.from_user.id,
        )
    except Exception as exc:
        await message.answer(f"Не удалось создать отдел: {html.escape(str(exc))}")
        return
    await state.clear()
    country_name = data.get("department_country_name") or ""
    await log_purchase(
        "admin_action",
        action=f"Создан отдел #{department_id}: {country_name} / {title} / {fmt_money(price)}",
        admin_id=message.from_user.id,
    )
    await message.answer(
        f"{ICON_SUCCESS} <b>Отдел создан</b>\n\n"
        f"<b>Название:</b> {render_rich_text(title)}\n"
        f"{ICON_COUNTRY} <b>Страна:</b> {html.escape(str(country_name))}\n"
        f"<b>Цена:</b> {fmt_money(price)}",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="К стране", callback_data=f"admin_country:{country_id}", icon_custom_emoji_id=BTN_ICON_BACK)],
        ]),
    )


@dp.callback_query(F.data.startswith("admin_product_group:"))
async def admin_product_group_view(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    try:
        parts = query.data.split(":")
        sample_id_raw = parts[1]
        country_id_raw = parts[2]
        sample_product_id = int(sample_id_raw)
        country_id = int(country_id_raw)
        page = max(0, int(parts[3])) if len(parts) > 3 else 0
    except (ValueError, IndexError):
        await query.answer("Тип товара не найден.", show_alert=True)
        return
    group = await get_product_department(sample_product_id)
    if not group:
        await query.answer("Этот тип товара закончился.", show_alert=True)
        return
    total_accounts = await count_products_in_department(sample_product_id)
    total_pages = -(-total_accounts // PAGE_SIZE) if total_accounts > 0 else 1
    page = min(page, total_pages - 1)
    products = await list_products_in_department(sample_product_id, limit=PAGE_SIZE, offset=page * PAGE_SIZE)
    account_rows = []
    status_labels = {
        "available": "в наличии",
        "sold": "продан",
        "dead": "мёртвый",
        "waiting_code": "ожидает",
        "verifying": "проверка",
    }
    for product in products:
        status = status_labels.get(product["status"], product["status"])
        phone = product["phone"] or "—"
        account_rows.append([
            InlineKeyboardButton(
                text=f"#{product['product_id']} • {phone} • {status}",
                callback_data=f"admin_stock_product:{product['product_id']}:group:{sample_product_id}:{country_id}:{page}",
            )
        ])
    text = product_group_admin_text(group)
    if total_accounts:
        text += f"\n\n<b>Аккаунты отдела:</b> {total_accounts}"
    else:
        text += "\n\n<b>Аккаунтов в отделе:</b> 0"
    await safe_edit(
        query.message,
        text,
        admin_product_group_kb(
            int(group["sample_product_id"]),
            country_id,
            account_rows=account_rows,
            page=page,
            total_pages=total_pages,
        ),
    )


@dp.callback_query(F.data.startswith("admin_edit_group:"))
async def admin_edit_group_choice(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    try:
        _, sample_id_raw, country_id_raw = query.data.split(":", 2)
        sample_product_id = int(sample_id_raw)
        country_id = int(country_id_raw)
    except (ValueError, IndexError):
        await query.answer("Тип товара не найден.", show_alert=True)
        return
    group = await get_product_department(sample_product_id)
    if not group:
        await query.answer("Этот тип товара закончился.", show_alert=True)
        return
    await state.update_data(edit_group_sample_id=sample_product_id, edit_group_country_id=country_id)
    await safe_edit(
        query.message,
        product_group_admin_text(group) + "\n\nЧто редачим?",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Название", callback_data="edit_group_field:title")],
            [InlineKeyboardButton(text="Цена", callback_data="edit_group_field:price")],
            [InlineKeyboardButton(text="Описание", callback_data="edit_group_field:description")],
            [InlineKeyboardButton(text="Заметка после покупки", callback_data="edit_group_field:extra_code")],
            [InlineKeyboardButton(text="Отменить", callback_data=f"admin_product_group:{sample_product_id}:{country_id}", icon_custom_emoji_id=BTN_ICON_CANCEL)],
        ]),
    )


@dp.callback_query(F.data.startswith("admin_remove_group_ask:"))
async def admin_remove_group_ask(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    try:
        _, sample_id_raw, country_id_raw = query.data.split(":", 2)
        sample_product_id = int(sample_id_raw)
        country_id = int(country_id_raw)
    except (ValueError, IndexError):
        await query.answer("Отдел не найден.", show_alert=True)
        return
    group = await get_product_department(sample_product_id)
    if not group:
        await query.answer("Отдел не найден.", show_alert=True)
        return
    if int(group["stock_count"] or 0) > 0:
        text = (
            "<b>Отдел нельзя удалить</b>\n\n"
            f"<b>Название:</b> {render_rich_text(group['title'])}\n"
            f"{ICON_COUNTRY} <b>Страна:</b> {html.escape(group['country'])}\n"
            f"<b>В наличии:</b> {int(group['stock_count'] or 0)}\n\n"
            "Сначала продайте, перенесите или снимите с продажи все аккаунты этого отдела."
        )
        await safe_edit(query.message, text, admin_product_group_kb(sample_product_id, country_id))
        return
    text = (
        "<b>Удалить отдел?</b>\n\n"
        f"<b>Название:</b> {render_rich_text(group['title'])}\n"
        f"{ICON_COUNTRY} <b>Страна:</b> {html.escape(group['country'])}\n"
        f"<b>Цена:</b> {fmt_money(float(group['price']))}\n"
        f"<b>В наличии:</b> {int(group['stock_count'] or 0)}\n"
        f"<b>Всего аккаунтов:</b> {int(group['total_count'] or 0)}\n\n"
        "Отдел пропадёт из каталога и админского списка страны. Купленные аккаунты и история покупателей останутся."
    )
    await safe_edit(query.message, text, admin_product_group_remove_confirm_kb(sample_product_id, country_id))


@dp.callback_query(F.data.startswith("admin_remove_group:"))
async def admin_remove_group(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    try:
        _, sample_id_raw, country_id_raw = query.data.split(":", 2)
        sample_product_id = int(sample_id_raw)
        country_id = int(country_id_raw)
    except (ValueError, IndexError):
        await query.answer("Отдел не найден.", show_alert=True)
        return
    result = await remove_product_department(sample_product_id, query.from_user.id)
    if not result.get("ok"):
        if result.get("reason") == "has_available":
            await query.answer("Нельзя удалить отдел, пока в нём есть аккаунты в наличии.", show_alert=True)
            return
        await query.answer("Отдел не найден.", show_alert=True)
        return
    await log_purchase(
        "admin_action",
        action=(
            f"Удалён отдел: {result['country']} / {result['title']} / "
            f"{fmt_money(float(result['price']))}; товаров: {result['total']}; "
            f"из корзин убрано: {result['cart_removed']}"
        ),
        admin_id=query.from_user.id,
    )
    text = (
        f"{ICON_SUCCESS} <b>Отдел удалён</b>\n\n"
        f"<b>Название:</b> {render_rich_text(result['title'])}\n"
        f"{ICON_COUNTRY} <b>Страна:</b> {html.escape(str(result['country']))}\n"
        f"<b>Цена:</b> {fmt_money(float(result['price']))}\n"
        f"<b>Аккаунтов в отделе:</b> {int(result['total'])}\n"
        f"<b>Из корзин убрано:</b> {int(result['cart_removed'])}"
    )
    await safe_edit(
        query.message,
        text,
        InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="К стране", callback_data=f"admin_country:{country_id}", icon_custom_emoji_id=BTN_ICON_BACK)]]),
    )


@dp.callback_query(F.data.startswith("edit_group_field:"))
async def admin_edit_group_field(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    field = query.data.split(":", 1)[1]
    prompts = {
        "title": "Введите новое название кнопки товара:",
        "price": "Введите новую цену:",
        "description": "Введите новое описание (или - для удаления):",
        "extra_code": "Введите заметку после покупки (или - для удаления):",
    }
    if field not in prompts:
        await query.answer("Поле не найдено.", show_alert=True)
        return
    await state.update_data(edit_group_field=field)
    await state.set_state(AdminEditProductGroupStates.waiting_new_value)
    data = await state.get_data()
    country_id = int(data.get("edit_group_country_id") or 0)
    await safe_edit(query.message, prompts[field], cancel_flow_kb(f"admin_country:{country_id}"))


@dp.message(AdminEditProductGroupStates.waiting_new_value)
async def admin_edit_group_new_value(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    sample_product_id = int(data.get("edit_group_sample_id") or 0)
    country_id = int(data.get("edit_group_country_id") or 0)
    field = data.get("edit_group_field")
    raw_value = (message.text or "").strip()
    if field == "price":
        value = parse_float(raw_value)
        if value is None or value < 0:
            await message.answer("Цена должна быть числом от 0.")
            return
    elif field in {"description", "extra_code"}:
        rich_value = sanitize_admin_text(message)
        value = "" if rich_value == "-" else rich_value
    elif field == "title":
        value = sanitize_admin_text(message)
        if len(plain_button_text(value)) < 3:
            await message.answer("Название слишком короткое.")
            return
    else:
        await message.answer("Поле не найдено.")
        await state.clear()
        return

    changed = await update_product_group_info(sample_product_id, **{field: value})
    await state.clear()
    if changed <= 0:
        await message.answer("Тип товара не найден или уже закончился.", reply_markup=admin_catalog_kb(await build_admin_country_rows()))
        return
    group = await get_product_department(sample_product_id)
    await log_purchase("admin_action", action=f"Обновлен тип товара sample #{sample_product_id}, поле {field}, товаров: {changed}", admin_id=message.from_user.id)
    if group:
        await message.answer(
            f"Обновлено аккаунтов: <b>{changed}</b>\n\n" + product_group_admin_text(group),
            reply_markup=admin_product_group_kb(int(group["sample_product_id"]), country_id),
        )
    else:
        await message.answer(
            f"Обновлено аккаунтов: <b>{changed}</b>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="К стране", callback_data=f"admin_country:{country_id}")]]),
        )


@dp.callback_query(F.data == "admin_country_add")
async def admin_country_add_start(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    await state.set_state(AdminCatalogStates.waiting_country_name)
    await safe_edit(
        query.message,
        "<b>Новая страна</b>\n\n"
        "Отправьте название страны с обычным или premium-флагом — ID бот определит сам.\n\n"
        "Например: <code>🇺🇿 Узбекистан</code>",
        cancel_flow_kb("admin_catalog"),
    )


@dp.message(AdminCatalogStates.waiting_country_name)
async def admin_country_add_name(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    name, icon_id = parse_country_name_and_icon(message)
    if len(name) < 2:
        await message.answer("Добавьте название страны после флага.")
        return
    try:
        country_id = await add_catalog_country(name, icon_id)
    except Exception as exc:
        await message.answer(f"Не удалось добавить страну: {html.escape(str(exc))}")
        return
    await state.clear()
    country = await get_catalog_country(country_id)
    await log_purchase("admin_action", action=f"Добавлена страна: {country['name']}", admin_id=message.from_user.id)
    await message.answer(
        f"Страна добавлена: <b>{html.escape(country['name'])}</b>",
        reply_markup=admin_catalog_kb(await build_admin_country_rows()),
    )


@dp.callback_query(F.data.startswith("admin_country_rename:"))
async def admin_country_rename_start(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    country_id = int(query.data.split(":", 1)[1])
    country = await get_catalog_country(country_id)
    if not country:
        await query.answer("Кнопка страны не найдена.", show_alert=True)
        return
    await state.update_data(
        rename_country_id=country_id,
        rename_country_old_name=country["name"],
        rename_country_old_icon=country["icon_custom_emoji_id"],
    )
    await state.set_state(AdminCatalogStates.waiting_country_rename)
    await safe_edit(
        query.message,
        "<b>Переименовать страну</b>\n\n"
        f"Сейчас: <b>{html.escape(country['name'])}</b>\n"
        "Отправьте новое название без флага.",
        cancel_flow_kb("admin_catalog"),
    )


@dp.message(AdminCatalogStates.waiting_country_rename)
async def admin_country_rename_name(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    new_name = " ".join((message.text or "").strip().split())
    if len(new_name) < 2:
        await message.answer("Название страны слишком короткое.")
        return
    await state.update_data(rename_country_new_name=new_name)
    await state.set_state(AdminCatalogStates.waiting_country_rename_icon)
    current_icon = data.get("rename_country_old_icon") or "не задан"
    await message.answer(
        "<b>Premium-флаг</b>\n\n"
        f"Текущий ID: <code>{html.escape(current_icon)}</code>\n\n"
        "Отправьте новый premium emoji, HTML с <code>emoji-id</code>, сам ID.\n"
        "Чтобы оставить текущий флаг, отправьте <code>-</code>.",
        reply_markup=cancel_flow_kb("admin_catalog"),
    )


@dp.message(AdminCatalogStates.waiting_country_rename_icon)
async def admin_country_rename_finish(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    country_id = int(data.get("rename_country_id") or 0)
    old_name = data.get("rename_country_old_name") or ""
    new_name = data.get("rename_country_new_name") or ""
    raw_text = (message.text or "").strip()
    keep_icon = raw_text == "-"
    icon_id = None if keep_icon else extract_custom_emoji_id(message)
    if not keep_icon and not icon_id:
        await message.answer("Не вижу premium emoji ID. Отправьте premium emoji, HTML с <code>emoji-id</code>, сам ID или <code>-</code>.")
        return
    try:
        ok = await rename_catalog_country(country_id, new_name, icon_id, keep_icon=keep_icon)
    except Exception as exc:
        await message.answer(f"Не удалось переименовать страну: {html.escape(str(exc))}")
        return
    await state.clear()
    if not ok:
        await message.answer("Кнопка страны не найдена.", reply_markup=admin_catalog_kb(await build_admin_country_rows()))
        return
    await log_purchase("admin_action", action=f"Переименована страна: {old_name} -> {new_name}", admin_id=message.from_user.id)
    await message.answer(
        f"Страна переименована: <b>{html.escape(old_name)}</b> → <b>{html.escape(new_name)}</b>",
        reply_markup=admin_catalog_kb(await build_admin_country_rows()),
    )


@dp.callback_query(F.data.startswith("admin_country_remove_ask:"))
async def admin_country_remove_ask(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    country_id = int(query.data.split(":", 1)[1])
    country = await get_catalog_country(country_id)
    if not country:
        await query.answer("Кнопка страны не найдена.", show_alert=True)
        return
    total = await count_products(country=country["name"])
    if total > 0:
        text = (
            "<b>Страну нельзя удалить</b>\n\n"
            f"{ICON_COUNTRY} <b>Страна:</b> {html.escape(country['name'])}\n"
            f"<b>В наличии:</b> {total}\n\n"
            "Сначала продайте, перенесите или снимите с продажи все аккаунты этой страны."
        )
        await safe_edit(query.message, text, admin_country_kb(country_id))
        return
    text = (
        "<b>Удалить страну из каталога?</b>\n\n"
        f"{ICON_COUNTRY} <b>Страна:</b> {html.escape(country['name'])}\n"
        f"<b>Товаров в наличии:</b> {total}\n\n"
        "Товары не удалятся, кнопка просто пропадет из каталога."
    )
    await safe_edit(query.message, text, admin_country_remove_confirm_kb(country_id))


@dp.callback_query(F.data.startswith("admin_country_remove:"))
async def admin_country_remove(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    country_id = int(query.data.split(":", 1)[1])
    country = await get_catalog_country(country_id)
    if country:
        total = await count_products(country=country["name"])
        if total > 0:
            await query.answer("Нельзя удалить страну, пока в ней есть аккаунты в наличии.", show_alert=True)
            await admin_country_view(query)
            return
    if await remove_catalog_country(country_id):
        await query.answer("Кнопка убрана из каталога.")
    else:
        await query.answer("Не удалось убрать кнопку.", show_alert=True)
    await safe_edit(query.message, "<b>Страны каталога</b>", admin_catalog_kb(await build_admin_country_rows()))


@dp.callback_query(F.data == "admin_proxy")
async def admin_proxy_menu(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    await safe_edit(query.message, build_proxy_text(), proxy_menu_kb(bool(load_global_proxy())))


@dp.callback_query(F.data == "admin_proxy_set")
async def admin_proxy_set(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    await state.set_state(AdminProxyStates.waiting_proxy_input)
    await safe_edit(
        query.message,
        "🌐 <b>Установка общего прокси</b>\n\n"
        "Отправь прокси одним сообщением. Поддерживаются:\n"
        "• <code>host:port</code>\n"
        "• <code>host:port:login:password</code>\n"
        "• <code>socks5://login:password@host:port</code>\n"
        "• <code>http://login:password@host:port</code>\n"
        "• <code>host:port:secret</code>\n"
        "• ссылка <code>t.me/proxy</code> или <code>tg://proxy</code>",
        cancel_flow_kb("admin_proxy"),
    )


@dp.callback_query(F.data == "admin_proxy_clear")
async def admin_proxy_clear(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    save_global_proxy(None)
    await log_purchase("admin_action", action="Удален общий прокси", admin_id=query.from_user.id)
    await safe_edit(query.message, build_proxy_text(), proxy_menu_kb(False))


@dp.callback_query(F.data == "admin_proxy_ping")
async def admin_proxy_ping(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer("Проверяю прокси...")
    if not is_admin(query.from_user.id):
        return
    proxy_settings = load_global_proxy()
    if not proxy_settings:
        await safe_edit(query.message, build_proxy_text(), proxy_menu_kb(False))
        return

    await safe_edit(
        query.message,
        "<b>Проверяю прокси...</b>\n\n"
        f"<code>{html.escape(format_proxy_summary(proxy_settings))}</code>",
        proxy_menu_kb(True),
    )
    try:
        result = await check_proxy_latency(proxy_settings)
    except Exception as exc:
        await safe_edit(
            query.message,
            "<b>Прокси не прошел проверку</b>\n\n"
            f"<code>{html.escape(format_proxy_summary(proxy_settings))}</code>\n\n"
            f"Ошибка: <code>{html.escape(str(exc) or type(exc).__name__)}</code>",
            proxy_menu_kb(True),
        )
        return

    await safe_edit(
        query.message,
        "<b>Прокси работает</b>\n\n"
        f"<code>{html.escape(format_proxy_summary(proxy_settings))}</code>\n\n"
        f"<b>Пинг:</b> {int(result.get('latency_ms', 0))} мс",
        proxy_menu_kb(True),
    )


@dp.message(AdminProxyStates.waiting_proxy_input)
async def admin_proxy_input(message: Message, state: FSMContext):
    try:
        proxy_settings = parse_proxy_input(message.text or "")
    except Exception as exc:
        await message.answer(f"Не удалось распознать прокси: {html.escape(str(exc))}")
        return
    progress = await message.answer("Проверяю прокси...")
    try:
        result = await check_proxy_latency(proxy_settings)
    except Exception as exc:
        await progress.edit_text(
            "Прокси не прошел проверку и не был сохранен.\n\n"
            f"<code>{html.escape(format_proxy_summary(proxy_settings))}</code>\n\n"
            f"Ошибка: <code>{html.escape(str(exc) or type(exc).__name__)}</code>"
        )
        return
    save_global_proxy(proxy_settings)
    await log_purchase("admin_action", action="Обновлен общий прокси", admin_id=message.from_user.id)
    await state.clear()
    await progress.edit_text(
        "Прокси сохранен.\n\n"
        f"<code>{html.escape(format_proxy_summary(proxy_settings))}</code>\n\n"
        f"<b>Пинг:</b> {int(result.get('latency_ms', 0))} мс",
        reply_markup=proxy_menu_kb(True),
    )


@dp.callback_query(F.data == "admin_stock")
@dp.callback_query(F.data == "admin_stock_catalog")
async def admin_stock_catalog(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    
    # Показываем только доступные страны (с available товарами)
    countries = await get_available_countries()
    
    rows = []
    for country_row in countries:
        country = country_row["country"]
        count = country_row["count"]
        rows.append([InlineKeyboardButton(text=f"{country} ({count})", callback_data=f"admin_stock_country:{country}")])
    
    if not rows:
        rows = [[InlineKeyboardButton(text="Нет товаров в наличии", callback_data="noop")]]
    
    text = "<b>Товары по странам</b>\n\nВыберите страну:"
    await safe_edit(query.message, text, admin_countries_available_kb(rows))


@dp.callback_query(F.data.startswith("admin_stock_country:"))
async def admin_stock_country(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    
    parts = query.data.split(":")
    country = parts[1]
    page = int(parts[2]) if len(parts) > 2 else 0
    
    offset = page * PAGE_SIZE
    products = await list_products(country=country, offset=offset, limit=PAGE_SIZE, status="available")
    total = await count_products(country=country, status="available")
    total_pages = -(-total // PAGE_SIZE) if total > 0 else 1
    
    rows = []
    for product in products:
        rows.append([InlineKeyboardButton(
            text=f"{product['title']} • {fmt_money(float(product['price']))}",
            callback_data=f"admin_stock_product:{product['product_id']}"
        )])
    
    if not rows:
        rows = [[InlineKeyboardButton(text="Нет товаров", callback_data="noop")]]
    
    text = f"<b>{country}</b>\n\nТовары в наличии: <b>{total}</b>"
    await safe_edit(query.message, text, admin_products_by_country_kb(rows, country=country, page=page, total_pages=total_pages))


@dp.callback_query(F.data == "admin_stock_sold_list")
@dp.callback_query(F.data.startswith("admin_stock_sold_list:"))
async def admin_stock_sold_list(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return

    page = 0
    if ":" in query.data:
        try:
            page = max(0, int(query.data.split(":", 1)[1]))
        except ValueError:
            page = 0

    total = await count_products(status="sold")
    total_pages = -(-total // ADMIN_SOLD_PAGE_SIZE) if total > 0 else 1
    page = min(page, total_pages - 1)
    offset = page * ADMIN_SOLD_PAGE_SIZE
    products = await list_products(status="sold", limit=ADMIN_SOLD_PAGE_SIZE, offset=offset)
    rows = []
    for p in products:
        title = p["title"] or f"Товар #{p['product_id']}"
        phone = p["phone"] or "—"
        rows.append([InlineKeyboardButton(
            text=f"#{p['product_id']} {title} • {phone}",
            callback_data=f"admin_stock_product:{p['product_id']}:sold:{page}",
        )])

    if not rows:
        rows = [[InlineKeyboardButton(text="Нет проданных товаров", callback_data="noop")]]

    text = (
        "<b>История продаж</b>\n\n"
        f"Продано аkkаунтов: <b>{total}</b>\n"
        f"На странице: <b>{len(products)}</b>"
    )
    await safe_edit(
        query.message,
        text,
        admin_sold_history_kb(product_rows=rows, page=page, total_pages=total_pages),
    )


@dp.callback_query(F.data.startswith("admin_edit_product:"))
async def admin_edit_product_choice(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id): return
    product_id = int(query.data.split(":")[1])
    await state.update_data(edit_product_id=product_id)
    
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Название", callback_data="edit_field:title")],
        [InlineKeyboardButton(text="Цена", callback_data="edit_field:price")],
        [InlineKeyboardButton(text="Описание", callback_data="edit_field:description")],
        [InlineKeyboardButton(text="Доп. к0d/заметка", callback_data="edit_field:extra_code")],
        [InlineKeyboardButton(text="Отменить", callback_data=f"admin_stock_product:{product_id}", icon_custom_emoji_id=BTN_ICON_CANCEL)]
    ])
    await safe_edit(query.message, "<b>Что редактируем?</b>", kb)


@dp.callback_query(F.data.startswith("edit_field:"))
async def admin_edit_field_start(query: CallbackQuery, state: FSMContext):
    field = query.data.split(":")[1]
    await state.update_data(edit_field=field)
    await state.set_state(AdminEditProductStates.waiting_new_value)
    
    prompts = {
        "title": "Введите новое название:",
        "price": "Введите новую цену:",
        "description": "Введите новое описание (или - для удаления):",
        "extra_code": "Введите новую заметку/к0d (или - для удаления):"
    }
    await safe_edit(query.message, prompts[field], cancel_flow_kb("admin_home"))


@dp.message(AdminEditProductStates.waiting_new_value)
async def admin_edit_field_finish(message: Message, state: FSMContext):
    data = await state.get_data()
    product_id = data["edit_product_id"]
    field = data["edit_field"]
    val = message.text.strip()
    
    update_data = {}
    if field == "price":
        f_val = parse_float(val)
        if f_val is None:
            await message.answer("Введите число.")
            return
        update_data["price"] = f_val
    else:
        update_data[field] = "" if val == "-" else val

    await update_product_info(product_id, **update_data)
    await state.clear()
    product = await get_product(product_id)
    await message.answer("Данные обновлены.", reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="К товару", callback_data=f"admin_stock_product:{product_id}")]]))


@dp.callback_query(F.data.startswith("admin_download_session:"))
async def admin_download_session(query: CallbackQuery):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id): return
    product_id = int(query.data.split(":")[1])
    product = await get_product(product_id)
    
    if not product or not product["session_path"]:
        await query.answer("Файл сессии не найден.", show_alert=True)
        return
    
    session_file = Path(product["session_path"])
    if not session_file.exists():
        await query.answer("Файл на диске отсутствует.", show_alert=True)
        return
    
    await query.answer("Отправляю файл...")
    await query.message.answer_document(
        FSInputFile(session_file, filename=f"account_{product['phone'] or product_id}.session"),
        caption=f"Файл сессии для аkkаунта <code>{product['phone']}</code>"
    )


@dp.callback_query(F.data.startswith("admin_stock_product:"))
async def admin_stock_product_detail(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return

    parts = query.data.split(":")
    try:
        product_id = int(parts[1])
    except (IndexError, ValueError):
        await query.answer("Товар не найден.", show_alert=True)
        return

    back_callback = "admin_stock_catalog"
    if len(parts) >= 4 and parts[2] == "sold":
        try:
            sold_page = max(0, int(parts[3]))
        except ValueError:
            sold_page = 0
        back_callback = f"admin_stock_sold_list:{sold_page}"
    elif len(parts) >= 3 and parts[2] == "search":
        back_callback = "admin_product_search"
    elif len(parts) >= 6 and parts[2] == "group":
        try:
            sample_product_id = int(parts[3])
            country_id = int(parts[4])
            group_page = max(0, int(parts[5]))
            back_callback = f"admin_product_group:{sample_product_id}:{country_id}:{group_page}"
        except ValueError:
            back_callback = "admin_stock_catalog"

    product = await get_product(product_id)

    if not product:
        await query.answer("Товар не найден.", show_alert=True)
        return

    has_session = has_server_session(product)
    can_fetch_code = has_session and product["status"] in {"waiting_code", "verifying", "sold"}
    await safe_edit(
        query.message,
        product_admin_text(product),
        admin_product_detail_kb(
            product_id,
            back_callback=back_callback,
            can_terminate_sessions=has_session,
            can_fetch_code=can_fetch_code,
        ),
    )


@dp.callback_query(F.data.startswith("admin_get_code:"))
async def admin_get_code(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer("Получаю к0d...")
    if not is_admin(query.from_user.id):
        return
    product_id = int(query.data.split(":", 1)[1])
    product = await get_product(product_id)
    if not product:
        await query.answer("Товар не найден.", show_alert=True)
        return
    if not product["session_path"]:
        await query.answer("У товара нет серверной сессии.", show_alert=True)
        return

    await safe_edit(
        query.message,
        "<b>Получаю секретное число</b>\n\n"
        f"{ICON_PURCHASE_TAG} <b>Товар:</b> {render_rich_text(product['title'])}\n"
        f"<b>Телефон:</b> <code>{html.escape(product['phone'] or '—')}</code>\n\n"
        "Получаю секретное слово",
    )
    try:
        code = await session_manager.fetch_code_from_telegram(product_id)
    except Exception as exc:
        logger.exception("Admin code fetch failed for product #%s", product_id)
        await safe_edit(
            query.message,
            f"{ICON_BLOCK} <b>Не удалось получить к0d</b>\n\n"
            f"{ICON_PURCHASE_TAG} <b>Товар:</b> {render_rich_text(product['title'])}\n"
            f"<b>Телефон:</b> <code>{html.escape(product['phone'] or '—')}</code>\n\n"
            f"<b>Ошибка:</b> {html.escape(str(exc))}",
            admin_product_detail_kb(
                product_id,
                back_callback=f"admin_stock_product:{product_id}",
                can_terminate_sessions=has_server_session(product),
                can_fetch_code=has_server_session(product),
            ),
        )
        return

    twofa_text = (
        f"\n<b>Облачный пароль:</b> <code>{html.escape(product['twofa_password'])}</code>"
        if product["twofa_password"]
        else ""
    )
    await log_purchase("admin_action", action=f"Админ получил к0D товара #{product_id}", admin_id=query.from_user.id)
    await safe_edit(
        query.message,
        f"{ICON_CHECK} <b>К0D получен</b>\n\n"
        f"{ICON_PURCHASE_TAG} <b>Товар:</b> {render_rich_text(product['title'])}\n"
        f"<b>Телефон:</b> <code>{html.escape(product['phone'] or '—')}</code>\n"
        f"<b>К0D:</b> <code>{html.escape(str(code))}</code>{twofa_text}",
        admin_product_detail_kb(
            product_id,
            back_callback=f"admin_stock_product:{product_id}",
            can_terminate_sessions=has_server_session(product),
            can_fetch_code=has_server_session(product),
        ),
    )


@dp.callback_query(F.data.startswith("admin_terminate_sessions_ask:"))
async def admin_terminate_sessions_ask(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    product_id = int(query.data.split(":", 1)[1])
    product = await get_product(product_id)
    if not product:
        await query.answer("Товар не найден.", show_alert=True)
        return
    if not has_server_session(product):
        await query.answer("У товара нет серверной сессии.", show_alert=True)
        return
    text = (
        "<b>Завершить с3ccuu?</b>\n\n"
        f"{ICON_PURCHASE_TAG} <b>Товар:</b> {render_rich_text(product['title'])}\n"
        f"<b>Телефон:</b> <code>{html.escape(product['phone'] or '—')}</code>\n\n"
        "Бот завершит все активные с3ccuu аккаунта, кроме своей серверной с3ccuu.\n\n"
        f"{ICON_NOTICE} Это действие нельзя отменить."
    )
    await safe_edit(query.message, text, admin_terminate_sessions_step1_kb(product_id))


@dp.callback_query(F.data.startswith("admin_terminate_sessions_step2:"))
async def admin_terminate_sessions_step2(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    product_id = int(query.data.split(":", 1)[1])
    product = await get_product(product_id)
    if not product:
        await query.answer("Товар не найден.", show_alert=True)
        return
    text = (
        "<b>Финальное подтверждение</b>\n\n"
        f"Вы точно хотите завершить все чужие сессии у товара <code>#{product_id}</code>?\n\n"
        f"{ICON_NOTICE} Бот сохранит свою сессию и не будет удалять её автоочисткой."
    )
    await safe_edit(query.message, text, admin_terminate_sessions_step2_kb(product_id))


@dp.callback_query(F.data.startswith("admin_terminate_sessions_confirm:"))
async def admin_terminate_sessions_confirm(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer("Завершаю сессии...")
    if not is_admin(query.from_user.id):
        return
    product_id = int(query.data.split(":", 1)[1])
    product = await get_product(product_id)
    if not product:
        await query.answer("Товар не найден.", show_alert=True)
        return
    if not has_server_session(product):
        await query.answer("У товара нет серверной сессии.", show_alert=True)
        return

    await safe_edit(
        query.message,
        "<b>Завершение с3cсий</b>\n\n"
        "Подключаюсь к серверной сессии и закрываю остальные авторизации...",
    )
    result = await session_manager.terminate_other_sessions(product_id)
    if not result.get("ok"):
        await safe_edit(
            query.message,
            f"{ICON_BLOCK} <b>Не удалось завершить сессии</b>\n\n"
            f"<code>{html.escape(str(result.get('error') or 'Unknown error'))}</code>",
            admin_product_detail_kb(
                product_id,
                back_callback=f"admin_stock_product:{product_id}",
                can_terminate_sessions=has_server_session(product),
                can_fetch_code=has_server_session(product),
            ),
        )
        return

    await mark_product_session_cleanup_disabled(product_id)
    await log_purchase(
        "admin_action",
        action=(
            f"Завершены все сторонние сессии товара #{product_id} | "
            f"закрыто: {result.get('terminated', 0)}, ошибок: {result.get('failed', 0)}"
        ),
        admin_id=query.from_user.id,
    )
    product = await get_product(product_id)
    text = (
        f"{ICON_SUCCESS} <b>Сессии завершены</b>\n\n"
        f"<b>Всего с3ссuй:</b> {result.get('total', 0)}\n"
        f"<b>Серверная с3ссuя:</b> сохранена\n"
        f"<b>Завершено:</b> {result.get('terminated', 0)}\n"
        f"<b>Ошибок:</b> {result.get('failed', 0)}\n\n"
        "Авто-выход и удаление серверной с3ссuu для этого товара отключены.\n\n"
        f"{product_admin_text(product)}"
    )
    await safe_edit(
        query.message,
        text,
        admin_product_detail_kb(
            product_id,
            back_callback=f"admin_stock_product:{product_id}",
            can_terminate_sessions=has_server_session(product),
            can_fetch_code=has_server_session(product) and product["status"] in {"waiting_code", "verifying", "sold"},
        ),
    )


@dp.callback_query(F.data == "admin_stuck_products")
async def admin_stuck_products(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    
    stuck = await get_stuck_products()
    
    rows = []
    for product in stuck:
        product_id = product["product_id"]
        title = product["title"]
        status = product["status"]
        rows.append([InlineKeyboardButton(
            text=f"#{product_id} {title} ({status})",
            callback_data=f"admin_stuck_detail:{product_id}"
        )])
    
    if not rows:
        rows = [[InlineKeyboardButton(text="Нет застрявших товаров", callback_data="noop")]]
    
    text = f"<b>Застрявшие товары</b>\n\nТовары, которые не завершили покупку.\n\nКоличество: <b>{len(rows)}</b>"
    await safe_edit(query.message, text, stuck_products_kb(rows))


@dp.callback_query(F.data.startswith("admin_stuck_detail:"))
async def admin_stuck_detail(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    product_id = int(query.data.split(":", 1)[1])
    product = await get_product(product_id)
    if not product:
        await query.answer("Товар не найден.", show_alert=True)
        return
    text = (
        "<b>Застрявший товар</b>\n\n"
        f"<b>ID:</b> <code>{product_id}</code>\n"
        f"<b>Название:</b> {render_rich_text(product['title'])}\n"
        f"<b>Статус:</b> {html.escape(product['status'])}\n"
        f"<b>Покупатель:</b> <code>{product['sold_to'] or '—'}</code>\n"
        f"<b>Сумма:</b> {fmt_money(float(product['sold_price'] or product['price'] or 0))}\n\n"
        "Выберите действие:"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Вернуть в каталог с возвратом", callback_data=f"admin_stuck_return:{product_id}")],
        [InlineKeyboardButton(text="Удалить без возврата", callback_data=f"admin_stuck_remove:{product_id}", icon_custom_emoji_id=BTN_ICON_CANCEL)],
    ])
    if has_server_session(product):
        kb.inline_keyboard.append([
            InlineKeyboardButton(
                text="Завершить с3ccuu",
                callback_data=f"admin_terminate_sessions_ask:{product_id}",
                icon_custom_emoji_id=BTN_ICON_CANCEL,
            )
        ])
    kb.inline_keyboard.append([InlineKeyboardButton(text="Назад", callback_data="admin_stuck_products", icon_custom_emoji_id=BTN_ICON_BACK)])
    await safe_edit(query.message, text, kb)


@dp.callback_query(F.data.startswith("admin_stuck_return:"))
async def admin_stuck_return(query: CallbackQuery):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id):
        return
    product_id = int(query.data.split(":", 1)[1])
    product = await get_product(product_id)
    buyer_id = int(product["sold_to"]) if product and product["sold_to"] else None
    refund_amount = float((product["sold_price"] or product["price"] or 0)) if product else 0
    if await return_product_to_catalog(product_id):
        await log_purchase("admin_action", action=f"Застрявший товар #{product_id} возвращен в каталог с возвратом средств", admin_id=query.from_user.id)
        if buyer_id and refund_amount > 0:
            try:
                await bot.send_message(
                    buyer_id,
                    f"{ICON_NOTICE} <b>Покупка отменена</b>\n\n"
                    f"{ICON_PURCHASE_TAG} <b>Товар:</b> {render_rich_text(product['title'])}\n"
                    f"{ICON_COIN} <b>Возврат:</b> {fmt_money(refund_amount)}\n\n"
                    "Товар вернулся в каталог, средства возвращены на баланс.",
                    reply_markup=support_kb(),
                )
            except Exception:
                logger.exception("Could not notify user %s about admin refund for product #%s", buyer_id, product_id)
        await query.answer("Товар вернулся в каталог, деньги возвращены.")
    else:
        await query.answer("Не удалось вернуть товар.", show_alert=True)
    await admin_stuck_products(query)


@dp.callback_query(F.data.startswith("admin_stuck_remove:"))
async def admin_stuck_remove(query: CallbackQuery):
    await ensure_known_user(query)
    if not is_admin(query.from_user.id):
        return
    product_id = int(query.data.split(":", 1)[1])
    if await force_remove_product(product_id):
        await log_purchase("admin_action", action=f"Застрявший товар #{product_id} удален без возврата средств", admin_id=query.from_user.id)
        await query.answer("Товар удален без возврата средств.")
    else:
        await query.answer("Не удалось удалить товар.", show_alert=True)
    await admin_stuck_products(query)


@dp.callback_query(F.data.startswith("admin_remove_"))
async def admin_remove(query: CallbackQuery):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    product_id = int(query.data.split("_")[-1])
    product = await get_product(product_id)
    if not product:
        await query.answer("Товар не найден.", show_alert=True)
        await safe_edit(query.message, "<b>Админка</b>", admin_home_kb())
        return

    if product["status"] == "sold":
        session_path = (product["session_path"] or "").strip()
        if session_path:
            await session_manager.logout_and_delete_product_session(product_id)
        delete_result = await delete_sold_product_with_history(
            product_id,
            session_path,
            allow_session_cleared=True,
        )
        if delete_result == "deleted":
            await query.answer("Проданный товар удален.")
            await log_purchase(
                "admin_action",
                action=f"Проданный товар #{product_id} удален из карточки с очисткой сессии",
                admin_id=query.from_user.id,
            )
        else:
            await query.answer(f"Не удалось удалить: {delete_result}", show_alert=True)
        await safe_edit(query.message, "<b>Админка</b>", admin_home_kb())
        return
    
    # Пытаемся обычное удаление, если не сработает - принудительное
    if await remove_product(product_id):
        await query.answer("Товар снят с продажи.")
        await log_purchase("admin_action", action=f"Товар #{product_id} снят с продажи", admin_id=query.from_user.id)
    elif await force_remove_product(product_id):
        await query.answer("Товар удален.")
        await log_purchase("admin_action", action=f"Товар #{product_id} принудительно удален", admin_id=query.from_user.id)
    else:
        await query.answer("Товар не найден.", show_alert=True)
    
    await safe_edit(query.message, "<b>Админка</b>", admin_home_kb())


@dp.callback_query(F.data == "admin_topup")
async def admin_topup_start(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    await state.set_state(AdminTopUpStates.waiting_user_id)
    await safe_edit(query.message, f"<b>{ICON_COIN} Выдача баланса</b>\n\nОтправьте <code>user_id</code> пользователя.", cancel_flow_kb("admin_home"))


@dp.message(AdminTopUpStates.waiting_user_id)
async def admin_topup_user_id(message: Message, state: FSMContext):
    text = (message.text or "").strip()
    if not text.isdigit():
        await message.answer("Нужен числовой user_id.")
        return
    await state.update_data(target_user_id=int(text))
    await state.set_state(AdminTopUpStates.waiting_amount)
    await message.answer("Отправьте сумму. Пример: <code>25</code>")


@dp.message(AdminTopUpStates.waiting_amount)
async def admin_topup_amount(message: Message, state: FSMContext):
    amount = parse_float(message.text or "")
    if amount is None:
        await message.answer("Нужна корректная сумма.")
        return
    data = await state.get_data()
    target_user_id = int(data["target_user_id"])
    new_balance = await add_balance(target_user_id, amount, kind="admin_topup", note=f"Пополнение от администратора {message.from_user.id}", actor_id=message.from_user.id)
    await log_purchase("admin_topup", user_id=target_user_id, amount=amount, admin_id=message.from_user.id)
    await state.clear()
    await message.answer(f"{ICON_COIN} Баланс выдан.\n\n<b>User ID:</b> <code>{target_user_id}</code>\n<b>Сумма:</b> {fmt_money(amount)}\n{ICON_COIN} <b>Новый баланс:</b> {fmt_money(new_balance)}", reply_markup=admin_home_kb())
    try:
        await bot.send_message(target_user_id, f"{ICON_COIN} Баланс пополнен.\n\n<b>Сумма:</b> {fmt_money(amount)}\n<b>Текущий баланс:</b> {fmt_money(new_balance)}")
    except Exception:
        pass


@dp.callback_query(F.data == "admin_add")
async def admin_add_start(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    await state.clear()
    await session_manager.cleanup(query.from_user.id)
    
    # Выбор способа добавления товара
    await state.set_state(AdminAddProductStates.waiting_add_method)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="По номеру + к0d", callback_data="admin_add_by_phone")],
        [InlineKeyboardButton(text="Загрузить .session файл", callback_data="admin_add_by_session")],
        [InlineKeyboardButton(text="Отменить", callback_data="cancel_flow:admin_home", icon_custom_emoji_id=BTN_ICON_CANCEL)],
    ])
    await safe_edit(
        query.message,
        "<b>Добавление товара</b>\n\n"
        "Выберите способ добавления:\n\n"
        "<b>По номеру + к0d</b> - вход через номер телефона.\n"
        "<b>По файлу</b> - загрузка готового .session файла.",
        kb
    )


@dp.callback_query(F.data == "admin_add_by_phone")
async def admin_add_by_phone(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    
    current_proxy = load_global_proxy()
    await state.set_state(AdminAddProductStates.waiting_phone)
    await safe_edit(
        query.message,
        "<b>Добавление по номеру телефона</b>\n\n"
        f"Текущий общий прокси:\n<code>{html.escape(format_proxy_summary(current_proxy))}</code>\n\n"
        "Отправь номер телефона аккаунта.\nФормат: <code>+998901234567</code>",
        cancel_flow_kb("admin_home"),
    )


@dp.callback_query(F.data == "admin_add_by_session")
async def admin_add_by_session(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    
    # Выбор: одна или несколько сессий
    await state.set_state(AdminAddProductStates.waiting_session_count)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Одна сессuя", callback_data="session_count_single")],
        [InlineKeyboardButton(text="Несколько сессий", callback_data="session_count_bulk")],
        [InlineKeyboardButton(text="Отменить", callback_data="cancel_flow:admin_home", icon_custom_emoji_id=BTN_ICON_CANCEL)],
    ])
    await safe_edit(
        query.message,
        "<b>Добавление по .session файлам</b>\n\n"
        "Выберите режим:\n\n"
        "<b>Одна сессия</b> - один файл и одна карточка.\n"
        "<b>Несколько сессий</b> - много файлов с общими данными.",
        kb
    )


@dp.callback_query(F.data == "session_count_single")
async def session_count_single(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    
    await state.set_state(AdminAddProductStates.waiting_session_file)
    await safe_edit(
        query.message,
        "<b>Загрузка .session файла</b>\n\n"
        "Отправь файл сессии Telethon (.session файл)\n\n"
        "После загрузки я попрошу страну и отдел товара.",
        cancel_flow_kb("admin_home"),
    )


@dp.callback_query(F.data == "session_count_bulk")
async def session_count_bulk(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    
    await state.set_state(AdminAddProductStates.waiting_bulk_sessions)
    await state.update_data(
        bulk_sessions=[],
        bulk_session_hashes=[],
        bulk_session_names={},
        bulk_metadata_by_session={},
        bulk_pending_metadata={},
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Готово", callback_data="bulk_sessions_done")],
        [InlineKeyboardButton(text="Отменить", callback_data="cancel_flow:admin_home", icon_custom_emoji_id=BTN_ICON_CANCEL)],
    ])
    await safe_edit(
        query.message,
        "<b>Загрузка нескольких .session файлов</b>\n\n"
        "Отправляй .session, .json или .zip с парами session/json.\n\n"
        "Когда загрузите все файлы, нажмите <b>Готово</b>.\n\n"
        "Затем заполнишь данные один раз для всех сессий.",
        kb,
    )


@dp.message(AdminAddProductStates.waiting_bulk_sessions)
async def admin_add_bulk_sessions(message: Message, state: FSMContext):
    if not message.document:
        await message.answer("Отправьте .session, .json или .zip.")
        return
    
    doc = message.document
    file_name = doc.file_name or "account.session"
    lower_name = file_name.lower()
    if not lower_name.endswith((".session", ".json", ".zip")):
        await message.answer(f"Нужен .session, .json или .zip. Получен: {doc.file_name}")
        return
    
    try:
        file_info = await bot.get_file(doc.file_id)
        file_content = await bot.download(file_info)
        content = file_content.read()

        data = await state.get_data()
        bulk_hashes = data.get("bulk_session_hashes", [])
        bulk_sessions = data.get("bulk_sessions", [])
        bulk_session_names = data.get("bulk_session_names", {}) or {}
        metadata_by_session = data.get("bulk_metadata_by_session", {}) or {}
        pending_metadata = data.get("bulk_pending_metadata", {}) or {}

        if lower_name.endswith(".json"):
            metadata = parse_session_metadata_bytes(content, file_name)
            metadata_lookup_put(pending_metadata, file_name, metadata)
            await state.update_data(bulk_pending_metadata=pending_metadata)
            matched = await apply_bulk_metadata_to_existing_sessions(state, metadata, file_name)
            await message.answer(
                "JSON принят.\n"
                f"Файл: <code>{html.escape(file_name)}</code>\n"
                f"Совпало с уже загруженными сессиями: <b>{matched}</b>\n\n"
                "Можно отправить .session/.zip или нажать <b>Готово</b>.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Готово", callback_data="bulk_sessions_done")],
                    [InlineKeyboardButton(text="Отменить", callback_data="cancel_flow:admin_home", icon_custom_emoji_id=BTN_ICON_CANCEL)],
                ])
            )
            return

        if lower_name.endswith(".zip"):
            imported = 0
            duplicates = 0
            json_count = 0
            matched = 0
            errors = []
            json_lookup = dict(pending_metadata)
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                members = [item for item in archive.infolist() if not item.is_dir()]
                session_members = []
                for item in members:
                    inner_name = Path(item.filename).name
                    if not inner_name:
                        continue
                    inner_lower = inner_name.lower()
                    try:
                        if inner_lower.endswith(".json"):
                            metadata = parse_session_metadata_bytes(archive.read(item), inner_name)
                            metadata_lookup_put(json_lookup, inner_name, metadata)
                            json_count += 1
                        elif inner_lower.endswith(".session"):
                            session_members.append((item, inner_name))
                    except Exception as exc:
                        errors.append(f"{inner_name}: {exc}")

                for item, inner_name in session_members:
                    try:
                        session_bytes = archive.read(item)
                        file_hash = session_file_sha256(session_bytes)
                        if file_hash in bulk_hashes:
                            duplicates += 1
                            continue
                        session_path = unique_uploaded_session_path("bulk", file_hash, inner_name)
                        with open(session_path, "wb") as f:
                            f.write(session_bytes)
                        metadata = find_metadata_for_session(session_path, json_lookup, inner_name)
                        write_uploaded_session_metadata(session_path, metadata, file_name=inner_name)
                        if metadata:
                            metadata_by_session[str(session_path)] = metadata
                            matched += 1
                        bulk_sessions.append(str(session_path))
                        bulk_hashes.append(file_hash)
                        bulk_session_names[str(session_path)] = inner_name
                        imported += 1
                    except Exception as exc:
                        errors.append(f"{inner_name}: {exc}")

            await state.update_data(
                bulk_sessions=bulk_sessions,
                bulk_session_hashes=bulk_hashes,
                bulk_session_names=bulk_session_names,
                bulk_metadata_by_session=metadata_by_session,
                bulk_pending_metadata=json_lookup,
            )
            text = (
                "<b>ZIP обработан</b>\n\n"
                f"Добавлено сессий: <b>{imported}</b>\n"
                f"JSON внутри: <b>{json_count}</b>\n"
                f"JSON применено: <b>{matched}</b>\n"
                f"Дублей пропущено: <b>{duplicates}</b>\n"
                f"Всего в пачке: <b>{len(bulk_sessions)}</b>"
            )
            if errors:
                text += "\n\nОшибки:\n" + "\n".join(f"  • {html.escape(error)}" for error in errors[:5])
                if len(errors) > 5:
                    text += f"\n  ... и ещё {len(errors) - 5}"
            await message.answer(
                text,
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Готово", callback_data="bulk_sessions_done")],
                    [InlineKeyboardButton(text="Отменить", callback_data="cancel_flow:admin_home", icon_custom_emoji_id=BTN_ICON_CANCEL)],
                ])
            )
            return

        session_bytes = content
        file_hash = session_file_sha256(session_bytes)
        if file_hash in bulk_hashes:
            await message.answer(
                "Этот .session уже есть в текущей пачке и повторно не добавлен.\n\n"
                f"Файл: <code>{html.escape(file_name)}</code>"
            )
            return

        session_path = unique_uploaded_session_path("bulk", file_hash, file_name)
        with open(session_path, "wb") as f:
            f.write(session_bytes)

        metadata = find_metadata_for_session(session_path, pending_metadata, file_name)
        write_uploaded_session_metadata(session_path, metadata, file_name=file_name)
        
        bulk_sessions.append(str(session_path))
        bulk_hashes.append(file_hash)
        bulk_session_names[str(session_path)] = file_name
        if metadata:
            metadata_by_session[str(session_path)] = metadata
        await state.update_data(
            bulk_sessions=bulk_sessions,
            bulk_session_hashes=bulk_hashes,
            bulk_session_names=bulk_session_names,
            bulk_metadata_by_session=metadata_by_session,
        )
        
        await message.answer(
            f"Загружено: <b>{len(bulk_sessions)}</b>\n"
            f"Последний файл: <code>{html.escape(session_path.name)}</code>\n"
            f"SHA-256: <code>{file_hash[:12]}</code>\n\n"
            f"JSON: {'применён' if metadata else 'сгенерирован desktop-профиль'}\n\n"
            "Отправьте следующие файлы или нажмите <b>Готово</b>.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Готово", callback_data="bulk_sessions_done")],
                [InlineKeyboardButton(text="Отменить", callback_data="cancel_flow:admin_home", icon_custom_emoji_id=BTN_ICON_CANCEL)],
            ])
        )
    except zipfile.BadZipFile:
        await message.answer("ZIP не читается. Проверьте архив и отправьте ещё раз.")
    except Exception as exc:
        logger.error(f"Ошибка при загрузке bulk сессии: {exc}")
        await message.answer(f"Ошибка: {html.escape(str(exc))}")


@dp.callback_query(F.data == "bulk_sessions_done")
async def bulk_sessions_done(query: CallbackQuery, state: FSMContext):
    await ensure_known_user(query)
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    
    data = await state.get_data()
    bulk_sessions = data.get("bulk_sessions", [])
    
    if not bulk_sessions:
        await query.answer("Загрузите хотя бы одну сессию.", show_alert=True)
        return

    metadata_by_session = data.get("bulk_metadata_by_session", {}) or {}
    sessions_without_json = [path for path in bulk_sessions if str(path) not in metadata_by_session]
    if not sessions_without_json:
        await state.update_data(twofa_password="")
        await prompt_admin_add_country(
            query.message,
            state,
            f"Загружено сессий: {len(bulk_sessions)}.\n\n"
            "JSON найден для каждой сессии, общий 2FA не нужен.\n"
            "Если в JSON нет пароля, для этого аккаунта считается, что 2FA нет.\n\n"
            "Выберите страну каталога для этих сессий.",
        )
        return

    await state.set_state(AdminAddProductStates.waiting_bulk_password)
    await safe_edit(
        query.message,
        f"<b>Загружено сессий: {len(bulk_sessions)}</b>\n\n"
        f"Без JSON: <b>{len(sessions_without_json)}</b>\n\n"
        "🔐 Отправь общий пароль 2FA для сессий без JSON.\n\n"
        "Для сессий с JSON будет использован их собственный пароль из JSON, "
        "а если его там нет — 2FA считается отсутствующим.\n\n"
        "Если у сессий без JSON 2FA не требуется - отправь <code>-</code>",
        cancel_flow_kb("admin_home"),
    )


@dp.message(AdminAddProductStates.waiting_bulk_password)
async def admin_add_bulk_password(message: Message, state: FSMContext):
    password = (message.text or "").strip()
    if not password:
        await message.answer("Отправьте пароль или <code>-</code>, если 2FA не нужен.")
        return
    
    # Если "-" - значит 2FA нет
    twofa_password = "" if password == "-" else password
    
    await state.update_data(twofa_password=twofa_password)
    await prompt_admin_add_country(message, state, "Пароль сохранен.\n\nВыберите страну каталога для этих сессий.")



@dp.message(AdminAddProductStates.waiting_session_file)
async def admin_add_session_file(message: Message, state: FSMContext):
    if not message.document:
        await message.answer("Отправьте файл.")
        return
    
    doc = message.document
    file_name = doc.file_name or "account.session"
    lower_name = file_name.lower()
    if not lower_name.endswith((".session", ".json")):
        await message.answer("Нужен файл с расширением .session или .json.")
        return
    
    try:
        file_info = await bot.get_file(doc.file_id)
        file_content = await bot.download(file_info)
        content = file_content.read()

        if lower_name.endswith(".json"):
            metadata = parse_session_metadata_bytes(content, file_name)
            data = await state.get_data()
            session_path = data.get("session_path")
            if session_path:
                write_uploaded_session_metadata(session_path, metadata, file_name=file_name)
                await state.update_data(single_session_metadata=metadata)
                await message.answer(
                    "JSON прикреплён к загруженной сессии.\n\n"
                    "🔐 Отправь пароль 2FA\n\n"
                    "Если 2FA не требуется - отправь <code>-</code>",
                    reply_markup=cancel_flow_kb("admin_home")
                )
                await state.set_state(AdminAddProductStates.waiting_password)
                return
            await state.update_data(single_session_metadata=metadata)
            await message.answer(
                "JSON принят.\n\n"
                "Теперь отправьте .session для этого аккаунта.",
                reply_markup=cancel_flow_kb("admin_home")
            )
            return

        session_bytes = content
        file_hash = session_file_sha256(session_bytes)
        session_path = unique_uploaded_session_path("single", file_hash, file_name)

        with open(session_path, "wb") as f:
            f.write(session_bytes)

        data = await state.get_data()
        metadata = data.get("single_session_metadata")
        write_uploaded_session_metadata(session_path, metadata, file_name=file_name)
        
        await state.update_data(
            session_path=str(session_path),
            session_sha256=file_hash,
            phone=str((metadata or {}).get("phone") or ""),
            telegram_id=(metadata or {}).get("user_id"),
            username=str((metadata or {}).get("username") or ""),
            first_name=str((metadata or {}).get("first_name") or ""),
            twofa_password="",
            single_session_metadata=metadata or {},
        )
        await state.set_state(AdminAddProductStates.waiting_password)
        await message.answer(
            f"Файл загружен: <code>{html.escape(session_path.name)}</code>\n"
            f"SHA-256: <code>{file_hash[:12]}</code>\n\n"
            f"JSON: {'применён' if metadata else 'сгенерирован desktop-профиль'}\n\n"
            "🔐 Отправь пароль 2FA\n\n"
            "Если 2FA не требуется - отправь <code>-</code>",
            reply_markup=cancel_flow_kb("admin_home")
        )
    except Exception as exc:
        logger.error(f"Ошибка при загрузке .session файла: {exc}")
        await message.answer(f"Ошибка при загрузке: {html.escape(str(exc))}")


@dp.message(AdminAddProductStates.waiting_phone)
async def admin_add_phone(message: Message, state: FSMContext):
    phone = parse_phone(message.text or "")
    if not phone:
        await message.answer("Введите корректный номер телефона.")
        return
    progress = await message.answer(
        "Отправляю секретное число...\n\n"
        f"Прокси:\n<code>{html.escape(format_proxy_summary(load_global_proxy()))}</code>"
    )
    try:
        await session_manager.send_code(message.from_user.id, phone)
    except Exception as exc:
        await progress.edit_text(f"Не удалось отправить секретное число:\n\n{html.escape(str(exc))}", reply_markup=admin_home_kb())
        return
    await state.update_data(phone=phone, code_input="")
    await state.set_state(AdminAddProductStates.waiting_code)
    await progress.edit_text(build_code_prompt_text(""), reply_markup=code_keypad_kb(False))


async def submit_admin_login_code(target: Message | CallbackQuery, state: FSMContext, code: str) -> None:
    if len(code) < 5:
        if isinstance(target, CallbackQuery):
            await target.answer("К0d слишком короткий.", show_alert=True)
        else:
            await target.answer("К0d слишком короткий. Отправьте 5-8 цифр.")
        return

    await state.update_data(code_input=code)
    status_message = target.message if isinstance(target, CallbackQuery) else target
    await safe_edit(status_message, "Проверяю к0d входа...", cancel_flow_kb("admin_home"))
    try:
        result = await session_manager.submit_code(target.from_user.id, code)
    except LoginExpiredError as exc:
        await state.clear()
        await safe_edit(status_message, f"{html.escape(str(exc))}", admin_home_kb())
        return
    except Exception as exc:
        await state.clear()
        await session_manager.cleanup(target.from_user.id)
        await safe_edit(status_message, f"Ошибка логина: {html.escape(str(exc))}", admin_home_kb())
        return
    if not result["ok"]:
        await state.update_data(code_input="")
        await safe_edit(status_message, "Неверный к0d. Попробуйте еще раз.\n\n" + build_code_prompt_text(""), code_keypad_kb(False))
        return
    if result["need_password"]:
        await state.set_state(AdminAddProductStates.waiting_password)
        await safe_edit(status_message, "🔐 Нужен пароль 2FA. Отправьте его одним сообщением.", cancel_flow_kb("admin_home"))
        return
    me = result["me"]
    await state.update_data(telegram_id=int(me.id), username=me.username or "", first_name=me.first_name or "User", phone=getattr(me, "phone", "") or "", twofa_password="")
    await prompt_admin_add_country(status_message, state, "Сеccuя подключена.\n\nВыберите страну каталога для аккаунта.")


@dp.callback_query(F.data.startswith("code_digit:"))
async def code_digit(query: CallbackQuery, state: FSMContext):
    if await state.get_state() != AdminAddProductStates.waiting_code.state:
        await query.answer("Сейчас секретное число не ждётся.", show_alert=True)
        return
    data = await state.get_data()
    code_input = str(data.get("code_input", ""))
    if len(code_input) < 8:
        code_input += query.data.split(":", 1)[1]
    await state.update_data(code_input=code_input)
    await query.answer()
    await safe_edit(query.message, build_code_prompt_text(code_input), code_keypad_kb(len(code_input) >= 5))


@dp.callback_query(F.data == "code_backspace")
async def code_backspace(query: CallbackQuery, state: FSMContext):
    if await state.get_state() != AdminAddProductStates.waiting_code.state:
        await query.answer()
        return
    data = await state.get_data()
    code_input = str(data.get("code_input", ""))[:-1]
    await state.update_data(code_input=code_input)
    await query.answer()
    await safe_edit(query.message, build_code_prompt_text(code_input), code_keypad_kb(len(code_input) >= 5))


@dp.callback_query(F.data == "code_clear")
async def code_clear(query: CallbackQuery, state: FSMContext):
    if await state.get_state() != AdminAddProductStates.waiting_code.state:
        await query.answer()
        return
    await state.update_data(code_input="")
    await query.answer()
    await safe_edit(query.message, build_code_prompt_text(""), code_keypad_kb(False))


@dp.callback_query(F.data == "code_submit")
async def code_submit(query: CallbackQuery, state: FSMContext):
    if await state.get_state() != AdminAddProductStates.waiting_code.state:
        await query.answer("Сейчас к0d не ждётся.", show_alert=True)
        return
    data = await state.get_data()
    code = str(data.get("code_input", ""))
    await query.answer()
    await submit_admin_login_code(query, state, code)


@dp.message(AdminAddProductStates.waiting_code)
async def admin_add_code_message(message: Message, state: FSMContext):
    code = normalize_login_code(message.text or "")
    if not 5 <= len(code) <= 8:
        await message.answer("Отправьте к0d одним сообщением: 5-8 цифр.", reply_markup=code_keypad_kb(False))
        return
    await submit_admin_login_code(message, state, code)


@dp.message(AdminAddProductStates.waiting_password)
async def admin_add_password(message: Message, state: FSMContext):
    data = await state.get_data()
    session_path = data.get("session_path")

    if session_path and message.document:
        doc = message.document
        file_name = doc.file_name or "account.json"
        if not file_name.lower().endswith(".json"):
            await message.answer("На этом шаге можно прикрепить только .json или отправить пароль/-.")
            return
        try:
            file_info = await bot.get_file(doc.file_id)
            file_content = await bot.download(file_info)
            metadata = parse_session_metadata_bytes(file_content.read(), file_name)
            write_uploaded_session_metadata(session_path, metadata, file_name=file_name)
            await state.update_data(single_session_metadata=metadata)
        except Exception as exc:
            await message.answer(f"Не удалось прочитать JSON: {html.escape(str(exc))}")
            return
        await message.answer(
            "JSON прикреплён.\n\n"
            "Теперь отправь пароль 2FA или <code>-</code>, если 2FA не нужен.",
            reply_markup=cancel_flow_kb("admin_home")
        )
        return

    password = (message.text or "").strip()
    if not password:
        await message.answer("Отправьте пароль или <code>-</code>, если 2FA не нужен.")
        return
    
    if session_path:
        metadata = data.get("single_session_metadata") or load_existing_session_metadata(session_path)
        metadata_password = str((metadata or {}).get("twofa_password") or "")
        twofa_password = metadata_password if password == "-" and metadata_password else ("" if password == "-" else password)
        await state.update_data(twofa_password=twofa_password)
        await prompt_admin_add_country(message, state, "Пароль сохранен.\n\nВыберите страну каталога для аккаунта.")
    else:
        # Это обычное добавление по номеру - проверяем пароль
        progress = await message.answer("Проверяю пароль 2FA...")
        try:
            me = await session_manager.submit_password(message.from_user.id, password)
        except Exception as exc:
            await progress.edit_text(f"Пароль не подошел:\n\n{html.escape(str(exc))}")
            return
        await state.update_data(telegram_id=int(me.id), username=me.username or "", first_name=me.first_name or "User", phone=getattr(me, "phone", "") or "", twofa_password=password)
        flow_id = make_add_flow_id()
        country_rows = await build_country_select_rows(flow_id)
        if not country_rows:
            await state.clear()
            await session_manager.cleanup(message.from_user.id)
            await progress.edit_text("2FA принята, но стран в каталоге нет. Сначала добавьте страну: Админка → Страны каталога.", reply_markup=admin_home_kb())
            return
        await state.set_state(AdminAddProductStates.waiting_country)
        await state.update_data(add_flow_id=flow_id)
        await progress.edit_text("2FA принята.\n\nВыберите страну каталога для аккаунта.", reply_markup=country_select_kb(country_rows))


@dp.message(AdminAddProductStates.waiting_title)
async def admin_add_title(message: Message, state: FSMContext):
    title = sanitize_admin_text(message)
    if len(plain_button_text(title)) < 3:
        await message.answer("Название слишком короткое.")
        return
    await state.update_data(title=title)
    await state.set_state(AdminAddProductStates.waiting_price)
    await message.answer(
        f"Название отдела: <b>{render_rich_text(title)}</b>\n\n"
        f"Теперь цена в {html.escape(settings.currency)}. Пример: <code>25</code>",
        reply_markup=cancel_flow_kb("admin_home"),
    )


@dp.callback_query(F.data.startswith("add_country:"))
async def admin_add_country(query: CallbackQuery, state: FSMContext):
    if await state.get_state() != AdminAddProductStates.waiting_country.state:
        await query.answer("Сейчас страна не выбирается.", show_alert=True)
        return
    flow_id, country_id = parse_add_country_callback(query.data or "")
    if country_id is None:
        await query.answer("Кнопка страны не найдена.", show_alert=True)
        return
    if not await check_add_flow_id(query, state, flow_id):
        return
    country = await get_catalog_country(country_id)
    if not country:
        await query.answer("Кнопка страны не найдена.", show_alert=True)
        return
    await query.answer()
    await state.update_data(country=country["name"])
    await state.set_state(AdminAddProductStates.waiting_department)
    await safe_edit(
        query.message,
        f"{ICON_COUNTRY} <b>Страна:</b> {html.escape(country['name'])}\n\n"
        "Выберите отдел товара или создайте новый.",
        InlineKeyboardMarkup(inline_keyboard=await build_department_select_rows(country_id, country["name"], flow_id)),
    )


@dp.callback_query(F.data.startswith("add_department_back"))
async def admin_add_department_back(query: CallbackQuery, state: FSMContext):
    if await state.get_state() != AdminAddProductStates.waiting_department.state:
        await query.answer()
        return
    flow_id = parse_optional_flow_callback(query.data or "")
    if not await check_add_flow_id(query, state, flow_id):
        return
    await query.answer()
    await state.set_state(AdminAddProductStates.waiting_country)
    await safe_edit(query.message, "Выберите страну каталога для аккаунта.", country_select_kb(await build_country_select_rows(flow_id)))


@dp.callback_query(F.data.startswith("add_department_new"))
async def admin_add_department_new(query: CallbackQuery, state: FSMContext):
    if await state.get_state() != AdminAddProductStates.waiting_department.state:
        await query.answer("Сейчас отдел не выбирается.", show_alert=True)
        return
    flow_id = parse_optional_flow_callback(query.data or "")
    if not await check_add_flow_id(query, state, flow_id):
        return
    await query.answer()
    await state.set_state(AdminAddProductStates.waiting_title)
    await safe_edit(
        query.message,
        "Отправьте название нового отдела.\n\nНапример: <code>+7 RU | Рег имо</code>",
        cancel_flow_kb("admin_home"),
    )


@dp.callback_query(F.data.startswith("add_department:"))
async def admin_add_department(query: CallbackQuery, state: FSMContext):
    if await state.get_state() != AdminAddProductStates.waiting_department.state:
        await query.answer("Сейчас отдел не выбирается.", show_alert=True)
        return
    flow_id, sample_product_id = parse_add_department_callback(query.data or "")
    if sample_product_id is None:
        await query.answer("Отдел не найден.", show_alert=True)
        return
    if not await check_add_flow_id(query, state, flow_id):
        return
    group = await get_product_department(sample_product_id)
    if not group:
        await query.answer("Отдел закончился или удалён.", show_alert=True)
        return
    data = await state.get_data()
    if data.get("country") != group["country"]:
        await query.answer("Этот отдел из другой страны.", show_alert=True)
        return
    await query.answer("Отдел выбран.")
    await state.update_data(
        title=group["title"],
        price=float(group["price"]),
        description=group["description"] or "",
        extra_code=group["extra_code"] or "",
    )
    await finish_admin_add_products(query.message, state, query.from_user.id)


@dp.message(AdminAddProductStates.waiting_price)
async def admin_add_price(message: Message, state: FSMContext):
    price = parse_float(message.text or "")
    if price is None:
        await message.answer("Нужна корректная цена.")
        return
    await state.update_data(price=price)
    await state.set_state(AdminAddProductStates.waiting_description)
    await message.answer("Отправьте описание для покупателя. Если не нужно — отправьте <code>-</code>.")


@dp.message(AdminAddProductStates.waiting_description)
async def admin_add_description(message: Message, state: FSMContext):
    description = sanitize_admin_text(message)
    await state.update_data(description="" if description == "-" else description)
    await state.set_state(AdminAddProductStates.waiting_extra_code)
    await message.answer("Отправьте дополнительный к0d или заметку. Если не нужно — отправьте <code>-</code>.")


async def finish_admin_add_products(message: Message, state: FSMContext, admin_id: int) -> None:
    data = await state.get_data()
    extra_code = data.get("extra_code", "")

    bulk_sessions = data.get("bulk_sessions", [])
    if bulk_sessions:
        await message.answer(f"Создаю товаров: {len(bulk_sessions)}...")

        created_products = []
        errors = []
        seen_telegram_ids: dict[int, int] = {}
        seen_phones: dict[str, int] = {}
        bulk_metadata_by_session = data.get("bulk_metadata_by_session", {}) or {}

        for idx, session_path in enumerate(bulk_sessions, 1):
            try:
                json_metadata = bulk_metadata_by_session.get(str(session_path))
                session_twofa = (
                    str((json_metadata or {}).get("twofa_password") or "")
                    if json_metadata is not None
                    else str(data.get("twofa_password", ""))
                )
                alive_check = await session_manager.verify_session_file_alive(session_path)
                if not alive_check.get("alive"):
                    error = alive_check.get("error", "Неизвестная ошибка")
                    discard_session_files(session_path)
                    errors.append(f"Сессия #{idx}: {error}")
                    logger.warning(f"Сессия {idx}/{len(bulk_sessions)} не принята: {error}")
                    continue

                phone = (alive_check.get("phone") or "").strip()
                telegram_id = alive_check.get("user_id")
                if telegram_id is not None:
                    telegram_id = int(telegram_id)
                    if telegram_id in seen_telegram_ids:
                        discard_session_files(session_path)
                        errors.append(f"Сессия #{idx}: дубль Telegram ID {telegram_id} в текущей пачке (первая: #{seen_telegram_ids[telegram_id]}).")
                        continue
                    seen_telegram_ids[telegram_id] = idx
                if phone:
                    if phone in seen_phones:
                        discard_session_files(session_path)
                        errors.append(f"Сессия #{idx}: дубль телефона {phone} в текущей пачке (первая: #{seen_phones[phone]}).")
                        continue
                    seen_phones[phone] = idx

                existing = await find_existing_product_identity(
                    session_path=str(session_path),
                    phone=phone,
                    telegram_id=telegram_id,
                )
                if existing:
                    discard_session_files(session_path)
                    errors.append(f"Сессия #{idx}: дубль в базе, {duplicate_product_text(existing)}.")
                    continue

                product_id = await create_product(
                    title=data["title"],
                    country=data["country"],
                    price=float(data["price"]),
                    description=data.get("description", ""),
                    extra_code=extra_code,
                    session_path=session_path,
                    phone=phone,
                    telegram_id=telegram_id,
                    username=alive_check.get("username") or "",
                    first_name=alive_check.get("first_name") or "",
                    twofa_password=session_twofa,
                    created_by=admin_id,
                )
                created_products.append(product_id)
                logger.info(f"Товар #{product_id} (сессия {idx}/{len(bulk_sessions)}) создан успешно")

            except Exception as exc:
                errors.append(f"Товар #{idx}: {html.escape(str(exc))}")
                logger.error(f"Ошибка при создании товара {idx}: {exc}")

        await state.clear()
        await session_manager.cleanup(admin_id)

        result_text = f"<b>Товары созданы</b>\n\n"
        result_text += f"Успешно: <b>{len(created_products)}</b>\n"
        if errors:
            result_text += f"Ошибок: <b>{len(errors)}</b>\n\n"
            for error in errors[:5]:
                result_text += f"  • {error}\n"
            if len(errors) > 5:
                result_text += f"  ... и ещё {len(errors) - 5} ошибок"

        await message.answer(result_text, reply_markup=admin_home_kb())
        return

    try:
        uploaded_session_path = data.get("session_path")
        uploaded_alive_check = None
        if uploaded_session_path:
            uploaded_alive_check = await session_manager.verify_session_file_alive(uploaded_session_path)
            if not uploaded_alive_check.get("alive"):
                error = uploaded_alive_check.get("error", "Неизвестная ошибка")
                discard_session_files(uploaded_session_path)
                await state.clear()
                await session_manager.cleanup(admin_id)
                await message.answer(
                    f"Сессия невалидна и не принята:\n\n{html.escape(str(error))}\n\n"
                    "Аккаунт не добавлен в магазин.",
                    reply_markup=admin_home_kb()
                )
                return
            existing = await find_existing_product_identity(
                session_path=str(uploaded_session_path),
                phone=(uploaded_alive_check.get("phone") or "").strip(),
                telegram_id=uploaded_alive_check.get("user_id"),
                extra_code=extra_code,
            )
            if existing:
                discard_session_files(uploaded_session_path)
                await state.clear()
                await session_manager.cleanup(admin_id)
                await message.answer(
                    "Дубль не добавлен.\n\n"
                    f"{html.escape(duplicate_product_text(existing))}",
                    reply_markup=admin_home_kb(),
                )
                return
        else:
            existing = await find_existing_product_identity(
                phone=(data.get("phone", "") or "").strip(),
                telegram_id=data.get("telegram_id"),
                extra_code=extra_code,
            )
            if existing:
                await state.clear()
                await session_manager.cleanup(admin_id)
                await message.answer(
                    "Дубль не добавлен.\n\n"
                    f"{html.escape(duplicate_product_text(existing))}",
                    reply_markup=admin_home_kb(),
                )
                return
        product_id = await create_product(
            title=data["title"],
            country=data["country"],
            price=float(data["price"]),
            description=data.get("description", ""),
            extra_code=extra_code,
            session_path=str(uploaded_session_path) if uploaded_session_path else "pending",
            phone=(uploaded_alive_check.get("phone") if uploaded_alive_check else data.get("phone", "")) or "",
            telegram_id=(uploaded_alive_check.get("user_id") if uploaded_alive_check else data.get("telegram_id")),
            username=(uploaded_alive_check.get("username") if uploaded_alive_check else data.get("username", "")) or "",
            first_name=(uploaded_alive_check.get("first_name") if uploaded_alive_check else data.get("first_name", "")) or "",
            twofa_password=data.get("twofa_password", ""),
            created_by=admin_id,
        )

        if not uploaded_session_path:
            final_session = await session_manager.finalize_product_session(admin_id, product_id)
            await update_product_session_path(product_id, str(final_session))

        if not uploaded_alive_check:
            await message.answer("Проверяю аккаунт...")
            alive_check = await session_manager.verify_account_alive(product_id)

            if not alive_check.get("alive"):
                error = alive_check.get("error", "Неизвестная ошибка")
                await force_remove_product(product_id)
                await state.clear()
                await session_manager.cleanup(admin_id)
                await message.answer(
                    f"Аккаунт не прошел проверку:\n\n{error}\n\n"
                    f"Товар не добавлен в магазин.",
                    reply_markup=admin_home_kb()
                )
                return

            await update_product_info(
                product_id,
                phone=alive_check["phone"],
                telegram_id=alive_check["user_id"],
                username=alive_check["username"],
                first_name=alive_check["first_name"]
            )

    except Exception as exc:
        await state.clear()
        await session_manager.cleanup(admin_id)
        await message.answer(f"Не удалось сохранить товар: {html.escape(str(exc))}", reply_markup=admin_home_kb())
        return

    await state.clear()
    product = await get_product(product_id)
    await message.answer(
        "Аккаунт проверен и добавлен в магазин.\n\n" + product_admin_text(product),
        reply_markup=admin_home_kb()
    )


@dp.message(AdminAddProductStates.waiting_extra_code)
async def admin_add_extra_code(message: Message, state: FSMContext):
    extra_code = sanitize_admin_text(message)
    await state.update_data(extra_code="" if extra_code == "-" else extra_code)
    await finish_admin_add_products(message, state, message.from_user.id)


@dp.callback_query(F.data.startswith("cancel_flow:"))
async def cancel_flow(query: CallbackQuery, state: FSMContext):
    await query.answer()
    current_state = await state.get_state()
    data = await state.get_data()
    is_admin_add_flow = is_admin(query.from_user.id) and current_state in {
        AdminAddProductStates.waiting_add_method.state,
        AdminAddProductStates.waiting_session_count.state,
        AdminAddProductStates.waiting_session_file.state,
        AdminAddProductStates.waiting_bulk_sessions.state,
        AdminAddProductStates.waiting_bulk_password.state,
        AdminAddProductStates.waiting_phone.state,
        AdminAddProductStates.waiting_code.state,
        AdminAddProductStates.waiting_password.state,
        AdminAddProductStates.waiting_country.state,
        AdminAddProductStates.waiting_department.state,
        AdminAddProductStates.waiting_title.state,
        AdminAddProductStates.waiting_price.state,
        AdminAddProductStates.waiting_description.state,
        AdminAddProductStates.waiting_extra_code.state,
    }
    if is_admin_add_flow:
        session_path = data.get("session_path")
        if session_path and session_path != "pending":
            discard_session_files(session_path)
        for bulk_session_path in data.get("bulk_sessions", []) or []:
            if bulk_session_path and bulk_session_path != "pending":
                discard_session_files(bulk_session_path)
        await session_manager.cleanup(query.from_user.id)
    await state.clear()
    back_callback = query.data.split(":", 1)[1]
    if back_callback == "admin_home":
        await safe_edit(query.message, "<b>Админка</b>", admin_home_kb())
    elif back_callback == "admin_catalog":
        await safe_edit(query.message, "<b>Страны каталога</b>", admin_catalog_kb(await build_admin_country_rows()))
    elif back_callback.startswith("admin_country:"):
        try:
            country_id = int(back_callback.rsplit(":", 1)[1])
        except ValueError:
            await safe_edit(query.message, "<b>Страны каталога</b>", admin_catalog_kb(await build_admin_country_rows()))
        else:
            await render_admin_country(query.message, country_id)
    elif back_callback == "admin_proxy":
        await safe_edit(query.message, build_proxy_text(), proxy_menu_kb(bool(load_global_proxy())))
    elif back_callback == "admin_scan_accounts":
        await state.update_data(scan_interval=60, scan_limit=5)
        await safe_edit(
            query.message,
            "<b>Глубокая проверка аkkаунтов</b>\n\n"
            "Эта проверка подключается к с3cсuи каждого аkkаунта и осторожно проверяет, валидна ли она.\n\n"
            "<b>Интервал:</b> 60 сек\n"
            "<b>Лимит:</b> 5 аkkаунтов за запуск",
            admin_scan_settings_kb(60, 5),
        )
    elif back_callback == "menu_balance":
        text = (
            f"{ICON_COIN} <b>Баланс:</b> {fmt_money(await get_balance(query.from_user.id))}"
        )
        await safe_edit(query.message, text, back_to_main_kb(is_admin(query.from_user.id)))
    elif back_callback == "menu_catalog":
        await safe_edit(query.message, f"{ICON_CATALOG_SECTIONS} <b>Каталог</b>\n\nВыберите раздел:", catalog_sections_keyboard())
    elif back_callback == "catalog_accounts":
        await safe_edit(query.message, f"{ICON_TG_ACCOUNTS} <b>ТГ</b>\n\nВыберите страну:", catalog_home_kb(await build_country_rows()))
    elif back_callback == "user_topup_methods":
        await safe_edit(
            query.message,
            f"<b>{ICON_WALLET} Пополнение баланса</b>\n\n"
            f"{ICON_COIN} <b>Текущий баланс:</b> {fmt_money(await get_balance(query.from_user.id))}\n\n"
            "Выберите метод пополнения:",
            topup_methods_keyboard(),
        )
    else:
        await show_home(query)


@dp.callback_query(F.data == "noop")
async def noop_callback(query: CallbackQuery):
    await query.answer()


async def on_startup() -> None:
    await init_db()
    logger.info("Shop bot initialized")


async def main_async() -> None:
    await on_startup()
    # Регистрация middleware
    dp.message.outer_middleware(SubscriptionMiddleware())
    dp.message.outer_middleware(AgreementMiddleware())
    dp.callback_query.outer_middleware(CallbackFSMGuardMiddleware())
    dp.callback_query.outer_middleware(SubscriptionMiddleware())
    dp.callback_query.outer_middleware(AgreementMiddleware())
    await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())


def main() -> None:
    asyncio.run(main_async())
